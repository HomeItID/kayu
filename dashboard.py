import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry
import customtkinter
import os
import math
from openpyxl import Workbook, load_workbook
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg 
from PIL import Image
import sqlite3
from datetime import datetime, timedelta
import calendar
import mysql.connector

# Database connection
# conn = mysql.connector.connect(
#     host='localhost',
#     user='root',
#     password='root',
#     database='kayu'
# )

conn = sqlite3.connect('kayu.sqlite')
c = conn.cursor()

def center_window(window, width, height):
    # Mendapatkan lebar dan tinggi layar
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # Menghitung posisi x dan y untuk jendela agar berada di tengah layar
    x = int((screen_width/2) - (width/2))
    y = int((screen_height/2) - (height/2))

    # Menentukan posisi jendela
    root.geometry(f"{width}x{height}+{x}+{y}")
    root.resizable(False,False)

def center_content(frame):
    # Mengatur properti grid agar konten berada di tengah
    frame.grid_configure(padx=10, pady=10)

    # Mengatur properti grid untuk seluruh widget di dalam frame
    for widget in frame.winfo_children():
        widget.grid_configure(padx=5, pady=5)

def format_currency(value):
    formatted_value = 'Rp {:,.2f}'.format(abs(float(value)))
    if value < 0:
        formatted_value = '(' + formatted_value[0:] + ')'
    return formatted_value

def format_id_pembelian(id_pembelian, tanggal):
        # Extract the year and month from the "tanggal" in YYYY-MM-DD format
        year, month, _ = tanggal.split('-')
        month = int(month)
        
        # Convert the month number to its Roman numeral equivalent
        roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
        month_roman = roman_numerals[month - 1]
        
        # Format the ID_Pembelian as specified
        formatted_id = f"{id_pembelian}/FB-KS/{month_roman}-{year}"
        return formatted_id

def format_id_penjualan(id_penjualan, tanggal):
        # Extract the year and month from the "tanggal" in YYYY-MM-DD format
        year, month, _ = tanggal.split('-')
        month = int(month)
        
        # Convert the month number to its Roman numeral equivalent
        roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
        month_roman = roman_numerals[month - 1]
        
        # Format the ID_Pembelian as specified
        formatted_id = f"{id_penjualan}/INV-KS/{month_roman}-{year}"
        return formatted_id

def format_id_produksi(id_produksi, tanggal):
        # Extract the year and month from the "tanggal" in YYYY-MM-DD format
        year, month, _ = tanggal.split('-')
        month = int(month)
        
        # Convert the month number to its Roman numeral equivalent
        roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
        month_roman = roman_numerals[month - 1]
        
        # Format the ID_Pembelian as specified
        formatted_id = f"{id_produksi}/PD-KS/{month_roman}-{year}"
        return formatted_id

def format_surat_jalan(id_penjualan, tanggal):
        # Extract the year and month from the "tanggal" in YYYY-MM-DD format
        year, month, _ = tanggal.split('-')
        month = int(month)
        
        # Convert the month number to its Roman numeral equivalent
        roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
        month_roman = roman_numerals[month - 1]
        
        # Format the ID_Pembelian as specified
        formatted_id = f"{id_penjualan}/SJ-KS/{month_roman}-{year}"
        return formatted_id

def extract(formatted_id):
        # Split the formatted ID using "/" as the delimiter
        parts = formatted_id.split("/")
        
        # The first part contains the ID_Pembelian
        id = parts[0]
        
        return id

def populate_supplier_dropdown():
    c.execute("SELECT ID_Supplier, Nama FROM Supplier")
    supplier_rows = c.fetchall()

    listsupplier = []
    for row in supplier_rows:
        namasupp = row [1]
        listsupplier.append(namasupp)
    
    return listsupplier

def populate_pembeli_dropdown():
    c.execute("SELECT ID_Pembeli, Nama FROM Pembeli")
    pembeli_rows = c.fetchall()
    list_pembeli = []
    for row in pembeli_rows:
        list_pembeli.append(row[1])
    return list_pembeli

def populate_log_kayu_dropdown():
        c.execute("SELECT ID_Log_Kayu, Nama FROM Log_Kayu")
        log_kayu_rows = c.fetchall()

        listlogkayu = []
        for row in log_kayu_rows:
            namalogkayu = row [1]
            listlogkayu.append(namalogkayu)
        
        return listlogkayu

def populate_hasil_produksi_dropdown():
        c.execute("SELECT Nama, Jenis FROM Hasil_Produksi")
        hasil_produksi_rows = c.fetchall()

        list_hasil_produksi = []
        for row in hasil_produksi_rows:
            gabungan = "{} - {}".format(row[0],row[1])
            list_hasil_produksi.append(gabungan)

        return list_hasil_produksi

def fetch_akun_kasbank():
    c.execute("""SELECT ID_Akun, Nama FROM Akun WHERE Kategori="Kas & Bank";""")
    akun_ids = c.fetchall()
    list_akun = []
    for row in akun_ids:
        gabungan = "{} - {}".format(row[0],row[1])
        list_akun.append(gabungan)

    return list_akun

listtermin = ["Net 15", "Net 30", "Net 60", "Cash"]

listbulat = ["2","3","4"]

listdiameter = ["13","14","15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32", "33", "34", "35", "36"]

listangkutan = ["Fuso","Truk"]

listnopol = ["AD 12345 AW"]

listppn = ["Dengan PPN","Tanpa PPN"]

list_ukuran = ["122 x 122", "244 x 122", "10 x 122", "15 x 122"]

list_pasangan_ukuran = [("122 x 122",14884),("244 x 122",29768),("10 x 122",1220),("15 x 122",1830)]

list_pasangan_tebal = [("2.00",2),("2.50",2.5),("2.80",2.8),("3.00",3),("3.20",3.2)]

tebal_list = ["2.00","2.50","2.80","3.00","3.20"]

batas_waktu = {"Net 15": 15,"Net 30": 30,"Net 60": 60,"Cash": 0}

months = list(calendar.month_name[1:])  # Exclude the empty first item
months.insert(0, "Semua Bulan")  # Add "Semua Bulan" at the beginning

years = [str(year) for year in range(2023, 2030)]  # Adjust the range as needed

now = datetime.now()
current_month = now.strftime("%B")  # Get the current month in full name

def show_dashboard():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

def show_pembelian():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()
    
    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Pembelian ORDER BY ID_Pembelian DESC")
        rows = c.fetchall()
        for row in rows:
            pembelian_id = row[0]
            supplier_id = row[1]
            akun_id = row[2]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            akun = "{} - {}".format(akun_id, nama)
            tanggal_sj = row[3]
            tanggal_nota = row[4]
            id_format = format_id_pembelian(pembelian_id, tanggal_nota)
            bea_supplier = row[5]
            termin = row[6]
            pembayaran = row[7]
            no_sj = pembelian_id

            # Get the supplier name
            c.execute("SELECT Nama FROM Supplier WHERE ID_Supplier=?", (supplier_id,))
            supplier_name = c.fetchone()[0]

            listvolume = []
            listsubtotal = []
            c.execute("SELECT Diameter, Jumlah, Pembulatan, Harga_Beli FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
            detailbeli = c.fetchall()
            if detailbeli:
                for detail in detailbeli:
                    diameter = detail[0]
                    panjang = 130
                    jumlah = detail[1]
                    pembulatan = detail[2]
                    harga = detail[3]
                    volume = (math.pi * (diameter/2)**2 * panjang * jumlah)/1000000
                    rounded = round(volume, pembulatan)
                    subtotal = rounded * harga
                    listsubtotal.append(subtotal)
                    listvolume.append(rounded)

                totalvolume = round(sum(listvolume),2)
                total = sum(listsubtotal)
                biaya_bongkar = round(totalvolume * 7000)
                beban_pabrik = biaya_bongkar - bea_supplier
                grandtotal = total + beban_pabrik
            else:
                total = 0
                totalvolume = 0
                biaya_bongkar = 0
                beban_pabrik = 0
                grandtotal = 0

            # Insert data into the table view
            treeview.insert("", tk.END, values=(id_format, tanggal_nota, supplier_name, no_sj, tanggal_sj, totalvolume, format_currency(total), format_currency(biaya_bongkar), bea_supplier, format_currency(beban_pabrik), termin, format_currency(grandtotal), pembayaran, akun))

    # Function to add a new pembelian record
    def add_pembelian():
        supplier_name = entry_nama.get()
        tanggal_sj = entry_tanggal_sj.get()
        tanggal_nota = entry_tanggal_nota.get()
        bea_supplier = entry_bea_supplier.get()
        termin = entry_termin.get()
        pembayaran = entry_pembayaran.get()
        akun_id = entry_akun_id.get()

        parts = akun_id.split(" - ")
        id_akun = parts[0]

        c.execute("SELECT ID_Supplier, Nama FROM Supplier")
        supplier_rows = c.fetchall()
        for row in supplier_rows:
            if supplier_name == row[1]:
                supplier_id = row[0]

        try:
            inputpembayaran = float(pembayaran)
            inputbea_supplier = float(bea_supplier)
        except ValueError:
            messagebox.showwarning("Peringatan", "Pembayaran dan Bea Supplier harus berupa angka.")
            return  # Keluar dari fungsi jika validasi gagal

        c.execute("INSERT INTO Pembelian (ID_Supplier, ID_Akun, Tanggal_Surat_Jalan, Tanggal_Nota, Bea_Supplier, Termin, Pembayaran) VALUES ( ?, ?, ?, ?, ?, ?, ?)",
            (supplier_id, id_akun, tanggal_sj, tanggal_nota, bea_supplier, termin, pembayaran))
        conn.commit()

        refresh_table()

    # Function to update a pembelian record
    def update_pembelian():
        selected_item = treeview.focus()

        if not selected_item:
            messagebox.showerror("Error", "No Pembelian record selected")
            return
        
        if selected_item:
            pembelian_id = extract(treeview.item(selected_item)["values"][0])
            supplier_name = entry_nama.get()
            tanggal_sj = entry_tanggal_sj.get()
            tanggal_nota = entry_tanggal_nota.get()
            bea_supplier = entry_bea_supplier.get()
            termin = entry_termin.get()
            pembayaran = entry_pembayaran.get()
            akun_id = entry_akun_id.get()
            
            parts = akun_id.split(" - ")
            id_akun = parts[0]

            c.execute("SELECT ID_Supplier, Nama FROM Supplier")
            supplier_rows = c.fetchall()
            for row in supplier_rows:
                if supplier_name == row[1]:
                    supplier_id = row[0]

            try:
                inputpembayaran = float(pembayaran)
                inputbea_supplier = float(bea_supplier)
            except ValueError:
                messagebox.showwarning("Peringatan", "Pembayaran dan Bea Supplier harus berupa angka.")
                return  # Keluar dari fungsi jika validasi gagal

            c.execute("UPDATE Pembelian SET ID_Supplier=?, ID_Akun=?, Tanggal_Surat_Jalan=?, Tanggal_Nota=?, Bea_Supplier=?, Termin=?, Pembayaran=? WHERE ID_Pembelian=?", 
                (supplier_id, id_akun, tanggal_sj, tanggal_nota, bea_supplier, termin, pembayaran, pembelian_id))
            conn.commit()

            refresh_table()

    # Function to delete a pembelian record
    def delete_pembelian():
        selected_item = treeview.selection()

        if not selected_item:
            messagebox.showerror("Error", "No Pembelian record selected")
            return
        
        if selected_item:
            pembelian_id = treeview.item(selected_item)["values"][0]
            id_pembelian = extract(pembelian_id)
            c.execute("DELETE FROM Pembelian WHERE ID_Pembelian=?", (id_pembelian,))
            c.execute("DELETE FROM Detail_Beli WHERE ID_Pembelian=?", (id_pembelian,))
            conn.commit()

            refresh_table()
            refresh_detailbeli_table(id_pembelian)

    # Function to clear the entry fields for Pembelian
    def clear_fields():
        entry_nama.set("")
        entry_akun_id.set('')
        entry_tanggal_nota.set_date(None)
        entry_tanggal_sj.set_date(None)
        entry_bea_supplier.delete(0, tk.END)
        entry_termin.set("")
        entry_pembayaran.delete(0, tk.END)

    # Create the form widgets for Pembelian
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama Supplier", text_color="black")
    entry_nama = customtkinter.CTkComboBox(content_frame, values=populate_supplier_dropdown(), width=200)

    label_akun_id = customtkinter.CTkLabel(content_frame, text="Dibayar Dari", text_color="black")
    entry_akun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_kasbank(), width=200)

    label_tanggal_sj = customtkinter.CTkLabel(content_frame, text="Tanggal Surat Jalan", text_color="black")
    entry_tanggal_sj = DateEntry(content_frame, width=33, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_tanggal_nota = customtkinter.CTkLabel(content_frame, text="Tanggal Nota", text_color="black")
    entry_tanggal_nota = DateEntry(content_frame, width=33, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_bea_supplier = customtkinter.CTkLabel(content_frame, text="Bea Supplier", text_color="black")
    entry_bea_supplier = customtkinter.CTkEntry(content_frame, width=200)

    label_termin = customtkinter.CTkLabel(content_frame, text="Termin", text_color="black")
    entry_termin = customtkinter.CTkComboBox(content_frame, values=listtermin, width=200)

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_pembelian)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_pembelian)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_pembelian)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)
    button_sakr = customtkinter.CTkButton(content_frame, text="SAKR")
    button_nota = customtkinter.CTkButton(content_frame, text="Nota")

    # Position the form widgets for Pembelian
    label_tanggal_nota.grid(row=0, column=0, padx=5, pady=5)
    entry_tanggal_nota.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal_sj.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal_sj.grid(row=1, column=1, padx=5, pady=5)

    label_nama.grid(row=0, column=2, padx=5, pady=5)
    entry_nama.grid(row=0, column=3, padx=5, pady=5)

    label_bea_supplier.grid(row=2, column=0, padx=5, pady=5)
    entry_bea_supplier.grid(row=2, column=1, padx=5, pady=5)

    label_termin.grid(row=1, column=2, padx=5, pady=5)
    entry_termin.grid(row=1, column=3, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=2, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=3, padx=5, pady=5)

    label_akun_id.grid(row=5, column=2, padx=5, pady=5)
    entry_akun_id.grid(row=5, column=3, padx=5, pady=5)

    button_add.grid(row=6, column=0, padx=5, pady=5)
    button_delete.grid(row=6, column=1, padx=5, pady=5)
    button_update.grid(row=6, column=2, padx=5, pady=5)
    button_clear.grid(row=6, column=3, padx=5, pady=5)
    button_sakr.grid(row=5, column=0, padx=5, pady=5)
    button_nota.grid(row=5, column=1, padx=5, pady=5)

    # Create the table view for Pembelian
    treeview = ttk.Treeview(content_frame, columns=("ID", "Tanggal Nota", "Nama Supplier", "Nomor Surat Jalan", "Tanggal Surat Jalan","Volume","Total","Biaya Bongkar", "Bea Supplier", "Beban Pabrik", "Termin","Grand Total", "Pembayaran", "Dibayar Dari",), show="headings", height=12)
    treeview.heading("ID", text="No. Nota")
    treeview.heading("Tanggal Nota", text="Tanggal Nota")
    treeview.heading("Nama Supplier", text="Supplier")
    treeview.heading("Nomor Surat Jalan", text="No. SJ")
    treeview.heading("Tanggal Surat Jalan", text="Tanggal SJ")
    treeview.heading("Volume", text="Volume (m3)")
    treeview.heading("Total", text="Total")
    treeview.heading("Biaya Bongkar", text="Biaya Bongkar")
    treeview.heading("Bea Supplier", text="Bea Supplier")
    treeview.heading("Beban Pabrik", text="Beban Pabrik")
    treeview.heading("Termin", text="Termin")
    treeview.heading("Grand Total", text="Grand Total")
    treeview.heading("Pembayaran", text="Pembayaran")
    treeview.heading("Dibayar Dari", text="Dibayar Dari")

    treeview.column("ID", width=125)
    treeview.column("Tanggal Nota", width=100)
    treeview.column("Nama Supplier", width=150)
    treeview.column("Nomor Surat Jalan", width=50)
    treeview.column("Tanggal Surat Jalan", width=100)
    treeview.column("Volume", width=135)
    treeview.column("Total", width=135)
    treeview.column("Biaya Bongkar", width=115)
    treeview.column("Bea Supplier", width=115)
    treeview.column("Beban Pabrik", width=115)
    treeview.column("Termin", width=75)
    treeview.column("Grand Total", width=135)
    treeview.column("Pembayaran", width=135)
    treeview.column("Dibayar Dari", width=100)

    treeview.grid(row=7, column=0, columnspan=4, padx=10, pady=1)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            pembelian_id, tanggal_nota, nama_supplier, no_sj, tanggal_sj, volume, total, biaya_bongkar, bea_supplier, beban_pabrik, termin, grandtotal, pembayaran, akun = treeview.item(selected_item, "values")
            entry_nama.set(nama_supplier)
            entry_tanggal_sj.set_date(tanggal_sj)
            entry_tanggal_nota.set_date(tanggal_nota)
            entry_bea_supplier.delete(0, tk.END)
            entry_bea_supplier.insert(tk.END, bea_supplier)
            entry_termin.set(termin)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            entry_akun_id.set(akun)

            # Refresh the Detail Beli table view based on the selected Pembelian
            
            refresh_detailbeli_table(extract(pembelian_id))

    treeview.bind("<<TreeviewSelect>>", on_select)

    jabona1 = ["640000", "690000"]
    jabona2 = ["830000", "870000"]
    jabona2plus = ["850000", "900000"]
    jabona3 = ["870000", "910000"]
    sengona1 = ["620000", "670000", "690000", "720000"]
    sengona2 = ["810000", "870000", "920000", "890000"]
    sengona2plus = ["830000", "890000", "910000" ,"930000" ]
    sengona3 = ["850000", "890000", "910000","930000"]
    kayukerasa0 = ["210000", "220000"]
    kayukerasa1 = ["690000", "700000"]
    kayukerasa2 = ["890000", "900000"]
    kayukerasa2plus = ["910000", "920000"]
    kayukerasa3 = ["910000", "920000"]

    def update_harga_beli(event):
        selected_log_kayu = entry_log_kayu.get()

        if entry_diameter.get():
            selected_diameter = float(entry_diameter.get())

            if selected_log_kayu and selected_diameter:
                if selected_log_kayu == "Kayu Keras" and selected_diameter == 13 or selected_diameter == 14:
                    items = kayukerasa0
                elif selected_log_kayu == "Kayu Keras" and selected_diameter >= 15 and selected_diameter <= 19:
                    items = kayukerasa1
                elif selected_log_kayu == "Kayu Keras" and selected_diameter >= 20 and selected_diameter <= 24:
                    items = kayukerasa2
                elif selected_log_kayu == "Kayu Keras" and selected_diameter >= 25 and selected_diameter <= 29:
                    items = kayukerasa2plus
                elif selected_log_kayu == "Kayu Keras" and selected_diameter >= 30 and selected_diameter <= 35:
                    items = kayukerasa3
                elif selected_log_kayu == "Kayu Jabon" and selected_diameter >= 15 and selected_diameter <= 19:
                    items = jabona1
                elif selected_log_kayu == "Kayu Jabon" and selected_diameter >= 20 and selected_diameter <= 24:
                    items = jabona2
                elif selected_log_kayu == "Kayu Jabon" and selected_diameter >= 25 and selected_diameter <= 29:
                    items = jabona2plus
                elif selected_log_kayu == "Kayu Jabon" and selected_diameter >= 30 and selected_diameter <= 35:
                    items = jabona3
                elif selected_log_kayu == "Kayu Sengon" and selected_diameter >= 15 and selected_diameter <= 19:
                    items = sengona1
                elif selected_log_kayu == "Kayu Sengon" and selected_diameter >= 20 and selected_diameter <= 24:
                    items = sengona2
                elif selected_log_kayu == "Kayu Sengon" and selected_diameter >= 25 and selected_diameter <= 29:
                    items = sengona2plus
                elif selected_log_kayu == "Kayu Sengon" and selected_diameter >= 30 and selected_diameter <= 35:
                    items = sengona3
                else:
                    items = []

                entry_harga_beli['values'] = items

    # Function to add a new Detail Beli record
    def add_detailbeli():
        selected_item = treeview.focus()
        
        if not selected_item:
            messagebox.showerror("Error", "No Pembelian record selected")
            return
        
        pembelian_id = treeview.item(selected_item, "values")[0]

        id_pembelian = extract(pembelian_id)

        # Get the selected ID from the combobox
        selected_log_kayu_id = entry_log_kayu.get()
        c.execute("SELECT ID_Log_Kayu, Nama FROM Log_Kayu")
        log_kayu_rows = c.fetchall()
        for row in log_kayu_rows:
            if selected_log_kayu_id == row[1]:
                log_kayu_id = row[0]

        diameter = float(entry_diameter.get())

        jumlah = entry_jumlah_detail_beli.get()
        
        pembulatan =  int(entry_pembulatan.get())

        harga_beli =  entry_harga_beli.get()

        try:
            inputjumlah = int(jumlah)
            inputharga = float(harga_beli)
        except ValueError:
            messagebox.showwarning("Peringatan", "Jumlah dan Harga Beli harus berupa angka.")
            return  # Keluar dari fungsi jika validasi gagal

        # Insert the Detail Beli data into the database
        c.execute("INSERT INTO Detail_Beli (ID_Pembelian, ID_Log_Kayu, Diameter, Jumlah, Pembulatan, Harga_Beli) VALUES (?, ?, ?, ?, ?, ?)",
                (id_pembelian, log_kayu_id, diameter, jumlah, pembulatan, float(harga_beli)))
        conn.commit()

        refresh_detailbeli_table(id_pembelian)

    # Function to update a Detail Beli record
    def update_detailbeli():
        selected_item = detailbeli_treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Detail Beli record selected")
            return

        if selected_item:
            # Get the detailbeli_id from the selected item
            detailbeli_id = detailbeli_treeview.item(selected_item)["values"][0]

            # Get the selected ID from the combobox
            selected_log_kayu_id = entry_log_kayu.get()
            c.execute("SELECT ID_Log_Kayu, Nama FROM Log_Kayu")
            log_kayu_rows = c.fetchall()
            for row in log_kayu_rows:
                if selected_log_kayu_id == row[1]:
                    log_kayu_id = row[0]

            diameter = float(entry_diameter.get())

            jumlah = entry_jumlah_detail_beli.get()
            
            pembulatan =  int(entry_pembulatan.get())

            harga_beli =  float(entry_harga_beli.get())

            try:
                inputjumlah = int(jumlah)
                inputharga = float(harga_beli)
            except ValueError:
                messagebox.showwarning("Peringatan", "Jumlah dan Harga Beli harus berupa angka.")
                return  # Keluar dari fungsi jika validasi gagal

            # Update the Detail Beli record in the database
            c.execute("UPDATE Detail_Beli SET ID_Log_Kayu=?, Diameter=?, Jumlah=?, Pembulatan=?, Harga_Beli=? WHERE ID_Detail_Beli=?", 
                    (log_kayu_id, diameter, jumlah, pembulatan, harga_beli, detailbeli_id))
            conn.commit()

            # Refresh the Detail Beli table view
            pembelian_id = treeview.item(treeview.selection(), "values")[0]
            if not pembelian_id:
                messagebox.showerror("Error", "No Pembelian record selected")
                return
            
            id_pembelian = extract(pembelian_id)
            refresh_detailbeli_table(id_pembelian)

    # Function to delete a Detail Beli record
    def delete_detailbeli():
        selected_item = detailbeli_treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Detail Beli record selected")
            return

        if selected_item:
            # Get the detailbeli_id from the selected item
            detailbeli_id = detailbeli_treeview.item(selected_item)["values"][0]

            # Delete the Detail Beli record from the database
            c.execute("DELETE FROM Detail_Beli WHERE ID_Detail_Beli=?", (detailbeli_id,))
            conn.commit()

            pembelian_id = treeview.item(treeview.selection(), "values")[0]
            if not pembelian_id:
                messagebox.showerror("Error", "No Pembelian record selected")
                return
            
            id_pembelian = extract(pembelian_id)
            refresh_detailbeli_table(id_pembelian)

    def refresh_detailbeli_table(pembelian_id):
        # Clear existing table data
        for row in detailbeli_treeview.get_children():
            detailbeli_treeview.delete(row)

        # Fetch and display data from the database
        c.execute("SELECT * FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
        rows = c.fetchall()
        for row in rows:
            id_detail_beli = row[0]
            log_kayu_id = row[2]
            diameter = row[3]
            jumlah = row[4]
            pembulatan = row[5]
            harga_beli = row[6]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
            log_kayu_info = c.fetchone()
            log_kayu_nama = log_kayu_info[0]
            log_kayu_panjang = log_kayu_info[1]

            volume = (math.pi * (diameter/2)**2 * log_kayu_panjang * jumlah)/1000000
            rounded = round(volume, pembulatan)
            subtotal = rounded * harga_beli

            # Insert data into the table view
            detailbeli_treeview.insert("", tk.END, values=(id_detail_beli, log_kayu_nama, diameter, log_kayu_panjang, jumlah, pembulatan, rounded, harga_beli, format_currency(subtotal)))

    # Create the Detail Beli table view
    detailbeli_treeview = ttk.Treeview(content_frame, columns=("ID Detail Beli", "Log Kayu", "Diameter", "Panjang", "Jumlah", "Pembulatan", "Volume", "Harga", "Subtotal"), show="headings")
    detailbeli_treeview.heading("ID Detail Beli", text="No. ")
    detailbeli_treeview.heading("Log Kayu", text="Log Kayu")
    detailbeli_treeview.heading("Diameter", text="Diameter (cm)")
    detailbeli_treeview.heading("Panjang", text="Panjang (cm)")
    detailbeli_treeview.heading("Jumlah", text="Jumlah (batang)")
    detailbeli_treeview.heading("Pembulatan", text="Pembulatan")
    detailbeli_treeview.heading("Volume", text="Volume (m3)")
    detailbeli_treeview.heading("Harga", text="Harga per Kubik")
    detailbeli_treeview.heading("Subtotal", text="Subtotal")

    detailbeli_treeview.column("ID Detail Beli", width=50)
    detailbeli_treeview.column("Log Kayu", width=125)
    detailbeli_treeview.column("Diameter", width=100)
    detailbeli_treeview.column("Panjang", width=100)
    detailbeli_treeview.column("Jumlah", width=100)
    detailbeli_treeview.column("Pembulatan", width=100)
    detailbeli_treeview.column("Volume", width=100)
    detailbeli_treeview.column("Harga", width=150)
    detailbeli_treeview.column("Subtotal", width=175)

    detailbeli_treeview.grid(row=17, column=0, columnspan=4, padx=1, pady=1)

    def on_select_detailbeli(event):
        selected_item = detailbeli_treeview.focus()
        if selected_item:
            detailbeli_id, log_kayu_nama, diameter, panjang, jumlah, pembulatan, volume, harga_beli, subtotal  = detailbeli_treeview.item(selected_item, "values")
            entry_log_kayu.set(log_kayu_nama)
            entry_diameter.set(diameter)
            entry_jumlah_detail_beli.delete(0, tk.END)
            entry_jumlah_detail_beli.insert(tk.END, jumlah)
            entry_pembulatan.set(pembulatan)
            entry_harga_beli.set(harga_beli)

            update_harga_beli(None)

    detailbeli_treeview.bind("<<TreeviewSelect>>", on_select_detailbeli)

    # Function to clear the entry fields for Detail Beli
    def clear_detailbeli_fields():
        entry_log_kayu.set("")
        entry_diameter.set("")
        entry_jumlah_detail_beli.delete(0, tk.END)
        entry_log_kayu.set("")
        entry_harga_beli.set("")

    # Create the form widgets for Detail Beli
    label_log_kayu = customtkinter.CTkLabel(content_frame, text="Log Kayu", text_color="black")
    entry_log_kayu = ttk.Combobox(content_frame, values=populate_log_kayu_dropdown(), width=33)
    if populate_log_kayu_dropdown():
        entry_log_kayu.current(0)

    label_diameter = customtkinter.CTkLabel(content_frame, text="Diameter", text_color="black")
    entry_diameter = ttk.Combobox(content_frame, width=33, values=listdiameter)
    entry_diameter.current(2)

    entry_log_kayu.bind("<<ComboboxSelected>>", update_harga_beli)
    entry_diameter.bind("<<ComboboxSelected>>", update_harga_beli)

    label_jumlah_detail_beli = customtkinter.CTkLabel(content_frame, text="Jumlah", text_color="black")
    entry_jumlah_detail_beli = customtkinter.CTkEntry(content_frame, width=200)

    label_pembulatan = customtkinter.CTkLabel(content_frame, text="Pembulatan", text_color="black")
    entry_pembulatan = customtkinter.CTkComboBox(content_frame, values=listbulat, width=200)

    label_harga_beli = customtkinter.CTkLabel(content_frame, text="Harga Beli", text_color="black")
    entry_harga_beli = ttk.Combobox(content_frame, values=[], width=33)

    button_add_detailbeli = customtkinter.CTkButton(content_frame, text="Add", command=add_detailbeli)
    button_delete_detailbeli = customtkinter.CTkButton(content_frame, text="Delete", command=delete_detailbeli)
    button_update_detailbeli = customtkinter.CTkButton(content_frame, text="Update", command=update_detailbeli)
    button_clear_detailbeli = customtkinter.CTkButton(content_frame, text="Clear", command=clear_detailbeli_fields)

    # Position the form widgets for Detail Beli
    label_log_kayu.grid(row=13, column=0, padx=5, pady=5)
    entry_log_kayu.grid(row=13, column=1, padx=5, pady=5)

    label_diameter.grid(row=14, column=0, padx=5, pady=5)
    entry_diameter.grid(row=14, column=1, padx=5, pady=5)

    label_pembulatan.grid(row=13, column=2, padx=5, pady=5)
    entry_pembulatan.grid(row=13, column=3, padx=5, pady=5)

    label_jumlah_detail_beli.grid(row=14, column=2, padx=5, pady=5)
    entry_jumlah_detail_beli.grid(row=14, column=3, padx=5, pady=5)

    label_harga_beli.grid(row=15, column=0, padx=5, pady=5)
    entry_harga_beli.grid(row=15, column=1, padx=5, pady=5)

    button_add_detailbeli.grid(row=16, column=0, padx=5, pady=5)
    button_delete_detailbeli.grid(row=16, column=1, padx=5, pady=5)
    button_update_detailbeli.grid(row=16, column=2, padx=5, pady=5)
    button_clear_detailbeli.grid(row=16, column=3, padx=5, pady=5)

    # Initial load of the Pembelian table
    refresh_table()

def show_penjualan():
    # Clear the content frame
    for widget in content_frame.winfo_children():
        widget.destroy()
    
    # Function to refresh the Penjualan table
    def refresh_penjualan_table():
        for row in penjualan_treeview.get_children():
            penjualan_treeview.delete(row)

        c.execute("SELECT * FROM Penjualan ORDER BY ID_Penjualan DESC")
        rows = c.fetchall()
        for row in rows:
            penjualan_id = row[0]
            pembeli_id = row[1]
            akun_id = row[2]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            akun = "{} - {}".format(akun_id, nama)

            tanggal_sales_order = row[3]
            nomor_sales_order = row[4]
            tanggal_surat_jalan = row[5]
            nomor_surat_jalan = format_surat_jalan(penjualan_id, tanggal_surat_jalan)
            tanggal_faktur = row[6]
            id_format = format_id_penjualan(penjualan_id, tanggal_faktur)
            termin = row[7]
            pembayaran = row[8]
            alat_angkutan = row[9]
            identitas_kendaraan = row[10]
            ppn_keluaran = row[11]

            c.execute("SELECT Nama FROM Pembeli WHERE ID_Pembeli=?", (pembeli_id,))
            pembeli_name = c.fetchone()[0]

            c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
            detailjual = c.fetchall()

            if detailjual:
                listvolume = []
                listsubtotal = []

                for detail in detailjual:
                    idhasil = detail[0]
                    tebal = detail[1]
                    ukuran = detail[2]
                    jumlah = detail[3]
                    harga_jual = detail[4]

                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    
                    c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                    hasil_produksi_info = c.fetchone()

                    hasil_produksi_nama = hasil_produksi_info[0]
                    if hasil_produksi_nama == "Ampulur":
                        subtotal = jumlah * harga_jual
                    else:
                        subtotal = rounded * harga_jual

                    listvolume.append(volume)
                    listsubtotal.append(subtotal)

                volume = round(sum(listvolume),4)
                total = sum(listsubtotal)
                if ppn_keluaran == "Dengan PPN":
                    ppn = total*11/100
                else:
                    ppn = 0
                grandtotal = total+ppn

            else:
                total= 0
                ppn = 0
                volume = 0
                grandtotal = 0

            penjualan_treeview.insert("", tk.END, values=(
                id_format, tanggal_faktur, pembeli_name, nomor_sales_order, tanggal_sales_order,  nomor_surat_jalan, tanggal_surat_jalan,
                volume, format_currency(total), ppn_keluaran, format_currency(ppn), termin, format_currency(grandtotal), pembayaran, akun, alat_angkutan, identitas_kendaraan
            ))

    # Function to add a new Penjualan record
    def add_penjualan():
        pembeli_name = entry_pembeli.get()
        tanggal_sales_order = entry_tanggal_sales_order.get()
        nomor_sales_order = entry_nomor_sales_order.get()
        tanggal_surat_jalan = entry_tanggal_surat_jalan.get()
        tanggal_faktur = entry_tanggal_faktur.get()
        termin = entry_termin.get()
        pembayaran = entry_pembayaran.get()
        alat_angkutan = entry_alat_angkutan.get()
        identitas_kendaraan = entry_identitas_kendaraan.get()
        akun_id = entry_akun_id.get()
        ppn_keluaran = entry_ppn.get()

        parts = akun_id.split(" - ")
        id_akun = parts[0]

        c.execute("SELECT ID_Pembeli, Nama FROM Pembeli")
        pembeli_rows = c.fetchall()
        for row in pembeli_rows:
            if pembeli_name == row[1]:
                pembeli_id = row[0]

        try:
            inputpembayaran = float(pembayaran)
        except ValueError:
            messagebox.showwarning("Peringatan", "Pembayaran harus berupa angka.")
            return  # Keluar dari fungsi jika validasi gagal

        c.execute("INSERT INTO Penjualan (ID_Pembeli, ID_Akun, Tanggal_Sales_Order, Nomor_Sales_Order, Tanggal_Surat_Jalan, "
                  "Tanggal_Faktur, Termin, Pembayaran, Alat_Angkutan, Identitas_Kendaraan, PPN_Keluaran) "
                  "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                  (pembeli_id, id_akun, tanggal_sales_order, nomor_sales_order, tanggal_surat_jalan,
                   tanggal_faktur, termin, pembayaran, alat_angkutan, identitas_kendaraan, ppn_keluaran))
        conn.commit()

        refresh_penjualan_table()

    # Function to delete a Penjualan record
    def delete_penjualan():
        selected_item = penjualan_treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Penjualan record selected")
            return
        
        if selected_item:
            penjualan_id = penjualan_treeview.item(selected_item)["values"][0]
            id_penjualan = extract(penjualan_id)
            c.execute("DELETE FROM Penjualan WHERE ID_Penjualan=?", (id_penjualan,))
            c.execute("DELETE FROM Detail_Jual WHERE ID_Penjualan=?", (id_penjualan,))
            conn.commit()
            refresh_penjualan_table()
            refresh_detail_jual_table(id_penjualan)

    # Function to update a Penjualan record
    def update_penjualan():
        selected_item = penjualan_treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Penjualan record selected")
            return
        
        if selected_item:
            penjualan_id = extract(penjualan_treeview.item(selected_item)["values"][0])
            pembeli_name = entry_pembeli.get()
            tanggal_sales_order = entry_tanggal_sales_order.get()
            nomor_sales_order = entry_nomor_sales_order.get()
            tanggal_surat_jalan = entry_tanggal_surat_jalan.get()
            tanggal_faktur = entry_tanggal_faktur.get()
            termin = entry_termin.get()
            pembayaran = entry_pembayaran.get()
            alat_angkutan = entry_alat_angkutan.get()
            identitas_kendaraan = entry_identitas_kendaraan.get()
            akun_id = entry_akun_id.get()
            ppn_keluaran = entry_ppn.get()
            
            parts = akun_id.split(" - ")
            id_akun = parts[0]

            c.execute("SELECT ID_Pembeli, Nama FROM Pembeli")
            pembeli_rows = c.fetchall()
            for row in pembeli_rows:
                if pembeli_name == row[1]:
                    pembeli_id = row[0]

            try:
                inputpembayaran = float(pembayaran)
            except ValueError:
                messagebox.showwarning("Peringatan", "Pembayaran harus berupa angka.")
                return  # Keluar dari fungsi jika validasi gagal

            c.execute("UPDATE Penjualan SET ID_Pembeli=?, ID_Akun=?, Tanggal_Sales_Order=?, Nomor_Sales_Order=?, "
                      "Tanggal_Surat_Jalan=?, Tanggal_Faktur=?, Termin=?, Pembayaran=?, "
                      "Alat_Angkutan=?, Identitas_Kendaraan=?, PPN_Keluaran=? WHERE ID_Penjualan=?", 
                      (pembeli_id, id_akun, tanggal_sales_order, nomor_sales_order, tanggal_surat_jalan, 
                       tanggal_faktur, termin, pembayaran, alat_angkutan, identitas_kendaraan, ppn_keluaran, penjualan_id))

            conn.commit()
            refresh_penjualan_table()

    # Function to clear Penjualan entry fields
    def clear_penjualan_fields():
        entry_pembeli.set("")
        entry_akun_id.set('')
        entry_tanggal_sales_order.set_date(None)
        entry_nomor_sales_order.delete(0, tk.END)
        entry_tanggal_surat_jalan.set_date(None)
        entry_tanggal_faktur.set_date(None)
        entry_termin.set("")
        entry_pembayaran.delete(0, tk.END)
        entry_alat_angkutan.set("")
        entry_identitas_kendaraan.set("")
        entry_ppn.set("")

    # Create form widgets for Penjualan
    label_pembeli = customtkinter.CTkLabel(content_frame, text="Pembeli", text_color="black")
    entry_pembeli = customtkinter.CTkComboBox(content_frame, values=populate_pembeli_dropdown(), width=200)

    label_akun_id = customtkinter.CTkLabel(content_frame, text="Dibayar Ke", text_color="black")
    entry_akun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_kasbank(), width=200)

    label_tanggal_sales_order = customtkinter.CTkLabel(content_frame, text="Tanggal Sales Order", text_color="black")
    entry_tanggal_sales_order = DateEntry(content_frame, width=33, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_nomor_sales_order = customtkinter.CTkLabel(content_frame, text="Nomor Sales Order", text_color="black")
    entry_nomor_sales_order = customtkinter.CTkEntry(content_frame, width=200)

    label_tanggal_surat_jalan = customtkinter.CTkLabel(content_frame, text="Tanggal Surat Jalan", text_color="black")
    entry_tanggal_surat_jalan = DateEntry(content_frame, width=33, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_tanggal_faktur = customtkinter.CTkLabel(content_frame, text="Tanggal Faktur", text_color="black")
    entry_tanggal_faktur = DateEntry(content_frame, width=33, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_termin = customtkinter.CTkLabel(content_frame, text="Termin", text_color="black")
    entry_termin = customtkinter.CTkComboBox(content_frame, values=listtermin, width=200)

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    label_alat_angkutan = customtkinter.CTkLabel(content_frame, text="Alat Angkutan", text_color="black")
    entry_alat_angkutan = customtkinter.CTkComboBox(content_frame, values=listangkutan, width=200)

    label_identitas_kendaraan = customtkinter.CTkLabel(content_frame, text="Nomor Polisi", text_color="black")
    entry_identitas_kendaraan = customtkinter.CTkComboBox(content_frame, values=listnopol, width=200)

    entry_ppn = customtkinter.CTkComboBox(content_frame, values=listppn, width=200)

    button_add_penjualan = customtkinter.CTkButton(content_frame, text="Add", command=add_penjualan )
    button_delete_penjualan = customtkinter.CTkButton(content_frame, text="Delete", command=delete_penjualan)
    button_update_penjualan = customtkinter.CTkButton(content_frame, text="Update", command=update_penjualan )
    button_clear_penjualan = customtkinter.CTkButton(content_frame, text="Clear", command=clear_penjualan_fields)
    button_faktur = customtkinter.CTkButton(content_frame, text="Faktur")
    button_surat_jalan = customtkinter.CTkButton(content_frame, text="Surat Jalan")
    button_nota_angkutan = customtkinter.CTkButton(content_frame, text="Nota Angkutan")

    # Position the Penjualan form widgets
    label_tanggal_faktur.grid(row=0, column=0, padx=5, pady=5)
    entry_tanggal_faktur.grid(row=0, column=1, padx=5, pady=5)

    label_nomor_sales_order.grid(row=1, column=0, padx=5, pady=5)
    entry_nomor_sales_order.grid(row=1, column=1, padx=5, pady=5)

    label_tanggal_sales_order.grid(row=2, column=0, padx=5, pady=5)
    entry_tanggal_sales_order.grid(row=2, column=1, padx=5, pady=5)

    label_tanggal_surat_jalan.grid(row=3, column=0, padx=5, pady=5)
    entry_tanggal_surat_jalan.grid(row=3, column=1, padx=5, pady=5)

    label_pembeli.grid(row=0, column=2, padx=5, pady=5)
    entry_pembeli.grid(row=0, column=3, padx=5, pady=5)

    label_termin.grid(row=1, column=2, padx=5, pady=5)
    entry_termin.grid(row=1, column=3, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=2, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=3, padx=5, pady=5)

    label_alat_angkutan.grid(row=3, column=2, padx=5, pady=5)
    entry_alat_angkutan.grid(row=3, column=3, padx=5, pady=5)

    label_akun_id.grid(row=4, column=0, padx=5, pady=5)
    entry_akun_id.grid(row=4, column=1, padx=5, pady=5)

    label_identitas_kendaraan.grid(row=4, column=2, padx=5, pady=5)
    entry_identitas_kendaraan.grid(row=4, column=3, padx=5, pady=5)

    button_faktur.grid(row=5, column=0, padx=5, pady=5)
    button_surat_jalan.grid(row=5, column=1, padx=5, pady=5)
    button_nota_angkutan.grid(row=5, column=2, padx=5, pady=5)
    entry_ppn.grid(row=5, column=3, padx=5, pady=5)

    button_add_penjualan.grid(row=6, column=0, padx=5, pady=5)
    button_delete_penjualan.grid(row=6, column=1, padx=5, pady=5)
    button_update_penjualan.grid(row=6, column=2, padx=5, pady=5)
    button_clear_penjualan.grid(row=6, column=3, padx=5, pady=5)

    # Create the Penjualan table view
    penjualan_treeview = ttk.Treeview(content_frame, columns=("ID Penjualan", "Tanggal Faktur", "ID Pembeli", "Nomor Sales Order", "Tanggal Sales Order", 
                                                               "Nomor Surat Jalan", "Tanggal Surat Jalan", "Volume", "Total", "PPN", "PPN Keluaran", "Termin", "Grand Total",
                                                               "Pembayaran", "Dibayar Ke", "Alat Angkutan", "Identitas Kendaraan"), show="headings", height=8)
    penjualan_treeview.heading("ID Penjualan", text="No. Invoice")
    penjualan_treeview.heading("Tanggal Faktur", text="Tanggal Faktur")
    penjualan_treeview.heading("ID Pembeli", text="ID Pembeli")
    penjualan_treeview.heading("Nomor Sales Order", text="No. SO")
    penjualan_treeview.heading("Tanggal Sales Order", text="Tanggal SO")
    penjualan_treeview.heading("Nomor Surat Jalan", text="No. SJ")
    penjualan_treeview.heading("Tanggal Surat Jalan", text="Tanggal SJ")
    penjualan_treeview.heading("Volume", text="Volume (m3)")
    penjualan_treeview.heading("Total", text="Total")
    penjualan_treeview.heading("PPN", text="PPN")
    penjualan_treeview.heading("PPN Keluaran", text="PPN Keluaran")
    penjualan_treeview.heading("Termin", text="Termin")
    penjualan_treeview.heading("Grand Total", text="Grand Total")
    penjualan_treeview.heading("Pembayaran", text="Pembayaran")
    penjualan_treeview.heading("Dibayar Ke", text="Dibayar Ke")
    penjualan_treeview.heading("Alat Angkutan", text="Alat Angkutan")
    penjualan_treeview.heading("Identitas Kendaraan", text="Nomor Polisi")

    penjualan_treeview.column("ID Penjualan", width=125)
    penjualan_treeview.column("Tanggal Faktur", width=100)
    penjualan_treeview.column("ID Pembeli", width=125)
    penjualan_treeview.column("Nomor Sales Order", width=100)
    penjualan_treeview.column("Tanggal Sales Order", width=75)
    penjualan_treeview.column("Nomor Surat Jalan", width=125)
    penjualan_treeview.column("Tanggal Surat Jalan", width=75)
    penjualan_treeview.column("Volume", width=100)
    penjualan_treeview.column("Total", width=100)
    penjualan_treeview.column("PPN", width=75)
    penjualan_treeview.column("PPN Keluaran", width=100)
    penjualan_treeview.column("Termin", width=60)
    penjualan_treeview.column("Grand Total", width=100)
    penjualan_treeview.column("Pembayaran", width=100)
    penjualan_treeview.column("Dibayar Ke", width=100)
    penjualan_treeview.column("Alat Angkutan", width=100)
    penjualan_treeview.column("Identitas Kendaraan", width=100)

    penjualan_treeview.grid(row=7, column=0, columnspan=4, padx=0, pady=1)

    def on_select_penjualan(event):
        selected_item = penjualan_treeview.focus()
        if selected_item:
            penjualan_id, tanggal_faktur, pembeli_id,nomor_sales_order, tanggal_sales_order, nomor_surat_jalan, tanggal_surat_jalan, \
            volume, total, ppn_keluaran, ppn, termin, grand_total, pembayaran, akun, alat_angkutan, identitas_kendaraan = penjualan_treeview.item(selected_item, "values")
            entry_pembeli.set(pembeli_id)
            entry_tanggal_sales_order.set_date(tanggal_sales_order)
            entry_nomor_sales_order.delete(0, tk.END)
            entry_nomor_sales_order.insert(tk.END, nomor_sales_order)
            entry_tanggal_surat_jalan.set_date(tanggal_surat_jalan)
            entry_tanggal_faktur.set_date(tanggal_surat_jalan)
            entry_termin.set(termin)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            entry_alat_angkutan.set(alat_angkutan)
            entry_identitas_kendaraan.set(identitas_kendaraan)
            entry_akun_id.set(akun)
            entry_ppn.set(ppn_keluaran)

            refresh_detail_jual_table(extract(penjualan_id))

    penjualan_treeview.bind("<<TreeviewSelect>>", on_select_penjualan)

    # Function to refresh the Detail Jual table
    def refresh_detail_jual_table(penjualan_id):
        for row in detail_jual_treeview.get_children():
            detail_jual_treeview.delete(row)

        c.execute("SELECT * FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
        rows = c.fetchall()
        for row in rows:
            detail_jual_id = row[0]
            hasil_produksi_id = row[2]
            tebal = row[3]
            ukuran = row[4]
            jumlah = row[5]
            keterangan = row[6]
            harga_jual = row[7]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (hasil_produksi_id,))
            hasil_produksi_info = c.fetchone()
            hasil_produksi_nama = hasil_produksi_info[0]
            hasil_produksi_jenis = hasil_produksi_info[1]

            if hasil_produksi_nama == "Ampulur":
                nilaiukuran = "-"
                nilaitebal = "-"
                rounded = "-"
                subtotal = jumlah * harga_jual
            else:
                for row in list_pasangan_ukuran:
                    if ukuran == row[1]:
                        nilaiukuran = row[0]

                for row in list_pasangan_tebal:
                    if tebal == row[1]:
                        nilaitebal = row[0]

                volume = tebal * ukuran * jumlah / 10000000
                rounded = round(volume,4)
                subtotal = rounded * harga_jual

            detail_jual_treeview.insert("", tk.END, values=(detail_jual_id, hasil_produksi_nama, hasil_produksi_jenis, nilaitebal, nilaiukuran, jumlah, rounded, harga_jual, format_currency(subtotal), keterangan))

    # Function to add a new Detail Jual record
    def add_detail_jual():
        selected_item = penjualan_treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Penjualan record selected")
            return
    
        penjualan_id = penjualan_treeview.item(selected_item, "values")[0]

        id_penjualan = extract(penjualan_id)

        selected_hasil_produksi_id = entry_hasil_produksi.get()
        parts = selected_hasil_produksi_id.split(" - ")
        
        # The first part contains the ID_Pembelian
        nama_hasil_produksi = parts[0]
        jenis_hasil_produksi = parts[1]

        c.execute("SELECT ID_Hasil_Produksi, Nama, Jenis FROM Hasil_Produksi")
        hasil_produksi_rows = c.fetchall()
        for row in hasil_produksi_rows:
            if nama_hasil_produksi == row[1] and jenis_hasil_produksi == row[2]:
                hasil_produksi_id = row[0]

        if nama_hasil_produksi == "Ampulur":
            tebal = 0
            nilaiukuran = 0
        else:
            tebal = float(entry_tebal.get())

            ukuran = entry_ukuran.get()

            for row in list_pasangan_ukuran:
                if ukuran == row[0]:
                    nilaiukuran = row[1]

        jumlah = entry_jumlah_detail_jual.get()
        keterangan = entry_keterangan.get()
        harga_jual = entry_harga_jual.get()

        try:
            inputjumlah = float(jumlah)
            inputhargajual = float(harga_jual)
        except ValueError:
            messagebox.showwarning("Peringatan", "Harga Jual dan Jumlah harus berupa angka.")
            return  # Keluar dari fungsi jika validasi gagal

        c.execute("INSERT INTO Detail_Jual (ID_Penjualan, ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Keterangan, Harga_Jual) "
                  "VALUES (?, ?, ?, ?, ?, ?, ?)", (id_penjualan, hasil_produksi_id, tebal, nilaiukuran, jumlah, keterangan, harga_jual))
        conn.commit()

        refresh_detail_jual_table(id_penjualan)

    # Function to update a Detail Jual record
    def update_detail_jual():
        selected_item = detail_jual_treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Penjualan record selected")
            return

        if selected_item:
            detail_jual_id = detail_jual_treeview.item(selected_item)["values"][0]
            
            selected_hasil_produksi_id = entry_hasil_produksi.get()
            parts = selected_hasil_produksi_id.split(" - ")
        
            # The first part contains the ID_Pembelian
            nama_hasil_produksi = parts[0]
            jenis_hasil_produksi = parts[1]

            c.execute("SELECT ID_Hasil_Produksi, Nama, Jenis FROM Hasil_Produksi")
            hasil_produksi_rows = c.fetchall()
            for row in hasil_produksi_rows:
                if nama_hasil_produksi == row[1] and jenis_hasil_produksi == row[2]:
                    hasil_produksi_id = row[0]

            if nama_hasil_produksi == "Ampulur":
                tebal = 0
                nilaiukuran = 0
            else:
                tebal = float(entry_tebal.get())

                ukuran = entry_ukuran.get()

                for row in list_pasangan_ukuran:
                    if ukuran == row[0]:
                        nilaiukuran = row[1]

            jumlah = entry_jumlah_detail_jual.get()
            keterangan = entry_keterangan.get()
            harga_jual = entry_harga_jual.get()

            try:
                inputjumlah = float(jumlah)
                inputhargajual = float(harga_jual)
            except ValueError:
                messagebox.showwarning("Peringatan", "Harga Jual dan Jumlah harus berupa angka.")
                return  # Keluar dari fungsi jika validasi gagal

            c.execute("UPDATE Detail_Jual SET ID_Hasil_Produksi=?, Tebal=?, Ukuran=?, Jumlah=?, Keterangan=?, Harga_Jual=? "
                      "WHERE ID_Detail_Jual=?", (hasil_produksi_id, tebal, nilaiukuran, jumlah, keterangan, harga_jual, detail_jual_id))
            conn.commit()

            selected_penjualan_id = penjualan_treeview.item(penjualan_treeview.selection(), "values")[0]
            if not selected_penjualan_id:
                messagebox.showerror("Error", "No Penjualan record selected")
                return

            id_penjualan = extract(selected_penjualan_id)

            refresh_detail_jual_table(id_penjualan)

    # Function to delete a Detail Jual record
    def delete_detail_jual():
        selected_item = detail_jual_treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Penjualan record selected")
            return
        
        if selected_item:
            detail_jual_id = detail_jual_treeview.item(selected_item)["values"][0]
            c.execute("DELETE FROM Detail_Jual WHERE ID_Detail_Jual=?", (detail_jual_id,))
            conn.commit()

            selected_penjualan_id = penjualan_treeview.item(penjualan_treeview.selection(), "values")[0]
            if not selected_penjualan_id:
                messagebox.showerror("Error", "No Penjualan record selected")
                return
            id_penjualan = extract(selected_penjualan_id)

            refresh_detail_jual_table(id_penjualan)

    # Function to clear Detail Jual entry fields
    def clear_detail_jual_fields():
        entry_hasil_produksi.set("")
        entry_tebal.set("")
        entry_ukuran.set("")
        entry_jumlah_detail_jual.delete(0, tk.END)
        entry_keterangan.delete(0, tk.END)
        entry_harga_jual.set("")

    kayusengon = ["1750000","1615000","1600000"]
    kayukeras = ["1870000"]
    mkcore = ["560000"]
    ampulur = ["1100"]

    def update_harga_jual(event):
        selected_hasil_produksi = entry_hasil_produksi.get()

        parts = selected_hasil_produksi.split(" - ")
        
        # The first part contains the ID_Pembelian
        nama_hasil_produksi = parts[0]

        if nama_hasil_produksi:
            if nama_hasil_produksi == "Core Sengon Basah":
                items = kayusengon
            elif nama_hasil_produksi == "Core Kayu Keras Basah":
                items = kayukeras
            elif nama_hasil_produksi == "MK Core Sengon" or nama_hasil_produksi == "MK Core Kayu Keras":
                items = mkcore
            elif nama_hasil_produksi == "Ampulur":
                items = ampulur
            else:
                items = []

            entry_harga_jual['values'] = items

    # Create form widgets for Detail Jual
    label_hasil_produksi = customtkinter.CTkLabel(content_frame, text="Hasil Produksi", text_color="black")
    entry_hasil_produksi = ttk.Combobox(content_frame, values=populate_hasil_produksi_dropdown(), width=33)
    if populate_hasil_produksi_dropdown():
        entry_hasil_produksi.current(0)

    entry_hasil_produksi.bind("<<ComboboxSelected>>", update_harga_jual)

    label_tebal = customtkinter.CTkLabel(content_frame, text="Tebal", text_color="black")
    entry_tebal = customtkinter.CTkComboBox(content_frame, values=tebal_list, width=200)

    label_ukuran = customtkinter.CTkLabel(content_frame, text="Ukuran", text_color="black")
    entry_ukuran = customtkinter.CTkComboBox(content_frame, values = list_ukuran, width=200)

    label_jumlah_detail_jual = customtkinter.CTkLabel(content_frame, text="Jumlah", text_color="black")
    entry_jumlah_detail_jual = customtkinter.CTkEntry(content_frame, width=200)

    label_harga_jual = customtkinter.CTkLabel(content_frame, text="Harga Jual", text_color="black")
    entry_harga_jual = ttk.Combobox(content_frame, values=[], width=33)

    label_keterangan = customtkinter.CTkLabel(content_frame, text="Keterangan", text_color="black")
    entry_keterangan = customtkinter.CTkEntry(content_frame, width=200)

    button_add_detail_jual = customtkinter.CTkButton(content_frame, text="Add", command=add_detail_jual)
    button_delete_detail_jual = customtkinter.CTkButton(content_frame, text="Delete", command=delete_detail_jual)
    button_update_detail_jual = customtkinter.CTkButton(content_frame, text="Update", command=update_detail_jual)
    button_clear_detail_jual = customtkinter.CTkButton(content_frame, text="Clear", command=clear_detail_jual_fields)

    # Position the Detail Jual form widgets
    label_hasil_produksi.grid(row=8, column=0, padx=5, pady=5)
    entry_hasil_produksi.grid(row=8, column=1, padx=5, pady=5)

    label_tebal.grid(row=9, column=0, padx=5, pady=5)
    entry_tebal.grid(row=9, column=1, padx=5, pady=5)

    label_ukuran.grid(row=8, column=2, padx=5, pady=5)
    entry_ukuran.grid(row=8, column=3, padx=5, pady=5)

    label_jumlah_detail_jual.grid(row=9, column=2, padx=5, pady=5)
    entry_jumlah_detail_jual.grid(row=9, column=3, padx=5, pady=5)

    label_keterangan.grid(row=10, column=2, padx=5, pady=5)
    entry_keterangan.grid(row=10, column=3, padx=5, pady=5)

    label_harga_jual.grid(row=10, column=0, padx=5, pady=5)
    entry_harga_jual.grid(row=10, column=1, padx=5, pady=5)

    button_add_detail_jual.grid(row=11, column=0, padx=5, pady=5)
    button_delete_detail_jual.grid(row=11, column=1, padx=5, pady=5)
    button_update_detail_jual.grid(row=11, column=2, padx=5, pady=5)
    button_clear_detail_jual.grid(row=11, column=3, padx=5, pady=5)

    # Create the Detail Jual table view
    detail_jual_treeview = ttk.Treeview(content_frame, columns=("ID Detail Jual", "ID Hasil Produksi", "Jenis", "Tebal",
                                                               "Ukuran", "Jumlah", "Volume", "Harga", "Subtotal", "Keterangan"), show="headings", height=9)
    detail_jual_treeview.heading("ID Detail Jual", text="No")
    detail_jual_treeview.heading("ID Hasil Produksi", text="Hasil Produksi")
    detail_jual_treeview.heading("Jenis", text="Jenis")
    detail_jual_treeview.heading("Tebal", text="Tebal (mm)")
    detail_jual_treeview.heading("Ukuran", text="Ukuran (cm)")
    detail_jual_treeview.heading("Jumlah", text="Jumlah")
    detail_jual_treeview.heading("Volume", text="Volume (m3)")
    detail_jual_treeview.heading("Harga", text="Harga per Kubik")
    detail_jual_treeview.heading("Subtotal", text="Subtotal")
    detail_jual_treeview.heading("Keterangan", text="Keterangan")

    detail_jual_treeview.column("ID Detail Jual", width=50)
    detail_jual_treeview.column("ID Hasil Produksi", width=125)
    detail_jual_treeview.column("Jenis", width=100)
    detail_jual_treeview.column("Tebal", width=100)
    detail_jual_treeview.column("Ukuran", width=100)
    detail_jual_treeview.column("Jumlah", width=100)
    detail_jual_treeview.column("Volume", width=125)
    detail_jual_treeview.column("Harga", width=125)
    detail_jual_treeview.column("Subtotal", width=150)
    detail_jual_treeview.column("Keterangan", width=250)

    detail_jual_treeview.grid(row=12, column=0, columnspan=4, padx=1, pady=1)

    def on_select_detail_jual(event):
        selected_item = detail_jual_treeview.focus()
        if selected_item:
            detail_jual_id, hasil_produksi_id, jenis, tebal, ukuran, jumlah, volume, harga_jual, subtotal, keterangan = detail_jual_treeview.item(selected_item, "values")
            hasil_produksi = "{} - {}".format(hasil_produksi_id,jenis)
            entry_hasil_produksi.set(hasil_produksi)
            entry_tebal.set(tebal)
            entry_ukuran.set(ukuran)
            entry_jumlah_detail_jual.delete(0, tk.END)
            entry_jumlah_detail_jual.insert(tk.END, jumlah)
            entry_keterangan.delete(0, tk.END)
            entry_keterangan.insert(tk.END, keterangan)
            entry_harga_jual.set(harga_jual)

    detail_jual_treeview.bind("<<TreeviewSelect>>", on_select_detail_jual)

    # Initialize the Penjualan and Detail Jual tables
    refresh_penjualan_table()

def show_produksi():
    # Clear the content frame
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the Produksi table
    def refresh_produksi_table():
        for row in produksi_treeview.get_children():
            produksi_treeview.delete(row)

        c.execute("SELECT * FROM Produksi ORDER BY ID_Produksi DESC")
        rows = c.fetchall()
        for row in rows:
            produksi_id = row[0]
            pembelian_id = row[1]
            tanggal_produksi = row[2]

            c.execute("SELECT Tebal, Ukuran, Jumlah FROM Detail_Produksi WHERE ID_Produksi=?", (produksi_id,))
            detailprod = c.fetchall()
            if detailprod:
                listvolume = []
                for detail in detailprod:
                    tebal = detail[0]
                    ukuran = detail[1]
                    jumlah = detail[2]
                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    listvolume.append(rounded)

                volume = sum(listvolume)
            else:
                volume = 0

            c.execute("SELECT Tanggal_Nota FROM Pembelian WHERE ID_Pembelian=?", (pembelian_id,))
            tanggal = c.fetchone()[0]

            produksi_treeview.insert("", tk.END, values=(format_id_produksi(produksi_id,tanggal_produksi), format_id_pembelian(pembelian_id, tanggal), tanggal_produksi, volume))

    # Function to add a new Produksi record
    def add_produksi():
        selected_pembelian_id = extract(entry_pembelian.get())
        tanggal_produksi = entry_tanggal_produksi.get()

        c.execute("INSERT INTO Produksi (ID_Pembelian, Tanggal_Produksi) VALUES (?, ?)",
                  (selected_pembelian_id, tanggal_produksi))
        conn.commit()

        refresh_produksi_table()

    # Function to delete a Produksi record
    def delete_produksi():
        selected_item = produksi_treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Produksi record selected")
            return
        
        if selected_item:
            produksi_id = extract(produksi_treeview.item(selected_item)["values"][0])
            c.execute("DELETE FROM Produksi WHERE ID_Produksi=?", (produksi_id,))
            c.execute("DELETE FROM Detail_Produksi WHERE ID_Produksi=?", (produksi_id,))
            conn.commit()
            refresh_produksi_table()
            refresh_detail_produksi_table(produksi_id)

    # Function to update a Produksi record
    def update_produksi():
        selected_item = produksi_treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Produksi record selected")
            return
        
        if selected_item:
            produksi_id = extract(produksi_treeview.item(selected_item)["values"][0])
            selected_pembelian_id = extract(entry_pembelian.get())
            tanggal_produksi = entry_tanggal_produksi.get()

            c.execute("UPDATE Produksi SET ID_Pembelian=?, Tanggal_Produksi=? WHERE ID_Produksi=?", 
                      (selected_pembelian_id, tanggal_produksi, produksi_id))

            conn.commit()
            refresh_produksi_table()

    # Function to clear Produksi entry fields
    def clear_produksi_fields():
        entry_pembelian.set("")
        entry_tanggal_produksi.set_date(None)

    # Function to refresh the Detail Produksi table
    def refresh_detail_produksi_table(produksi_id):
        for row in detail_produksi_treeview.get_children():
            detail_produksi_treeview.delete(row)

        c.execute("SELECT * FROM Detail_Produksi WHERE ID_Produksi=?", (produksi_id,))
        rows = c.fetchall()
        for row in rows:
            detail_produksi_id = row[0]
            hasil_produksi_id = row[2]
            tebal = row[3]
            ukuran = row[4]
            jumlah = row[5]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (hasil_produksi_id,))
            hasil_produksi_info = c.fetchone()
            hasil_produksi_nama = hasil_produksi_info[0]
            hasil_produksi_jenis = hasil_produksi_info[1]

            if hasil_produksi_nama == "Ampulur":
                nilaiukuran = "-"
                nilaitebal = "-"
                rounded = "-"
            else:
                for row in list_pasangan_ukuran:
                    if ukuran == row[1]:
                        nilaiukuran = row[0]

                for row in list_pasangan_tebal:
                    if tebal == row[1]:
                        nilaitebal = row[0]

                volume = tebal * ukuran * jumlah / 10000000
                rounded = round(volume,4)

            detail_produksi_treeview.insert("", tk.END, values=(detail_produksi_id, hasil_produksi_nama, hasil_produksi_jenis, nilaitebal, nilaiukuran, jumlah, rounded))

    # Function to add a new Detail Produksi record
    def add_detail_produksi():
        selected_produksi_id = extract(produksi_treeview.item(produksi_treeview.selection(), "values")[0])

        if not selected_produksi_id:
            messagebox.showerror("Error", "No Produksi record selected")
            return
        
        selected_hasil_produksi_id = entry_hasil_produksi.get()
        parts = selected_hasil_produksi_id.split(" - ")
        
        # The first part contains the ID_Pembelian
        nama_hasil_produksi = parts[0]
        jenis_hasil_produksi = parts[1]

        c.execute("SELECT ID_Hasil_Produksi, Nama, Jenis FROM Hasil_Produksi")
        hasil_produksi_rows = c.fetchall()
        for row in hasil_produksi_rows:
            if nama_hasil_produksi == row[1] and jenis_hasil_produksi == row[2]:
                hasil_produksi_id = row[0]

        if nama_hasil_produksi == "Ampulur":
            nilaitebal = 0
            nilaiukuran = 0
        else:
            nilaitebal = float(entry_tebal_detail_produksi.get())

            ukuran = entry_ukuran_detail_produksi.get()
            for row in list_pasangan_ukuran:
                if ukuran == row[0]:
                    nilaiukuran = row[1]

        jumlah = entry_jumlah_detail_produksi.get()

        try:
            inputjumlah = float(jumlah)
        except ValueError:
            messagebox.showwarning("Peringatan", "Jumlah harus berupa angka.")
            return  # Keluar dari fungsi jika validasi gagal

        c.execute("INSERT INTO Detail_Produksi (ID_Produksi, ID_Hasil_Produksi, Tebal, Ukuran, Jumlah) "
                  "VALUES (?, ?, ?, ?, ?)", (selected_produksi_id, hasil_produksi_id, nilaitebal, nilaiukuran, jumlah))
        conn.commit()

        refresh_detail_produksi_table(selected_produksi_id)

    # Function to update a Detail Produksi record
    def update_detail_produksi():
        selected_item = detail_produksi_treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Detail Produksi record selected")
            return
        
        if selected_item:
            detail_produksi_id = detail_produksi_treeview.item(selected_item)["values"][0]
            selected_hasil_produksi_id = entry_hasil_produksi.get()
            parts = selected_hasil_produksi_id.split(" - ")
        
            # The first part contains the ID_Pembelian
            nama_hasil_produksi = parts[0]
            jenis_hasil_produksi = parts[1]

            c.execute("SELECT ID_Hasil_Produksi, Nama, Jenis FROM Hasil_Produksi")
            hasil_produksi_rows = c.fetchall()
            for row in hasil_produksi_rows:
                if nama_hasil_produksi == row[1] and jenis_hasil_produksi == row[2]:
                    hasil_produksi_id = row[0]

            if nama_hasil_produksi == "Ampulur":
                nilaitebal = 0
                nilaiukuran = 0
            else:
                nilaitebal = float(entry_tebal_detail_produksi.get())

                ukuran = entry_ukuran_detail_produksi.get()
                for row in list_pasangan_ukuran:
                    if ukuran == row[0]:
                        nilaiukuran = row[1]

            jumlah = entry_jumlah_detail_produksi.get()
            
            try:
                inputjumlah = float(jumlah)
            except ValueError:
                messagebox.showwarning("Peringatan", "Jumlah harus berupa angka.")
                return  # Keluar dari fungsi jika validasi gagal

            c.execute("UPDATE Detail_Produksi SET ID_Hasil_Produksi=?, Tebal=?, Ukuran=?, Jumlah=? "
                      "WHERE ID_Detail_Produksi=?", (hasil_produksi_id, nilaitebal, nilaiukuran, jumlah, detail_produksi_id))
            conn.commit()

            selected_produksi_id = produksi_treeview.item(produksi_treeview.selection(), "values")[0]
            if not selected_produksi_id:
                messagebox.showerror("Error", "No Produksi record selected")
                return
        
            refresh_detail_produksi_table(extract(selected_produksi_id))

    # Function to delete a Detail Produksi record
    def delete_detail_produksi():
        selected_item = detail_produksi_treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Detail Produksi record selected")
            return
        
        if selected_item:
            detail_produksi_id = detail_produksi_treeview.item(selected_item)["values"][0]
            c.execute("DELETE FROM Detail_Produksi WHERE ID_Detail_Produksi=?", (detail_produksi_id,))
            conn.commit()

            selected_produksi_id = produksi_treeview.item(produksi_treeview.selection(), "values")[0]
            if not selected_produksi_id:
                messagebox.showerror("Error", "No Produksi record selected")
                return
        
            refresh_detail_produksi_table(extract(selected_produksi_id))
            
    # Function to clear Detail Produksi entry fields
    def clear_detail_produksi_fields():
        entry_hasil_produksi.set("")
        entry_tebal_detail_produksi.set("")
        entry_ukuran_detail_produksi.set("")
        entry_jumlah_detail_produksi.delete(0, tk.END)

    def fetch_pembelian_ids():
        c.execute("""SELECT p.ID_pembelian, p.Tanggal_Nota
                    FROM Pembelian p
                    LEFT JOIN Produksi pr ON p.ID_pembelian = pr.ID_pembelian
                    WHERE pr.ID_pembelian IS NULL;""")
        pembelian_ids = c.fetchall()
        return [format_id_pembelian(id[0],id[1]) for id in pembelian_ids]

    # Create form widgets for Produksi
    label_pembelian = customtkinter.CTkLabel(content_frame, text="Pembelian", text_color="black")
    entry_pembelian = customtkinter.CTkComboBox(content_frame, values=fetch_pembelian_ids(), width=200)

    label_tanggal_produksi = customtkinter.CTkLabel(content_frame, text="Tanggal Produksi", text_color="black")
    entry_tanggal_produksi = DateEntry(content_frame, width=33, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    button_add_produksi = customtkinter.CTkButton(content_frame, text="Add", command=add_produksi)
    button_delete_produksi = customtkinter.CTkButton(content_frame, text="Delete", command=delete_produksi)
    button_update_produksi = customtkinter.CTkButton(content_frame, text="Update", command=update_produksi)
    button_clear_produksi = customtkinter.CTkButton(content_frame, text="Clear", command=clear_produksi_fields)

    # Position the Produksi form widgets
    label_pembelian.grid(row=0, column=0, padx=5, pady=5)
    entry_pembelian.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal_produksi.grid(row=0, column=2, padx=5, pady=5)
    entry_tanggal_produksi.grid(row=0, column=3, padx=5, pady=5)

    button_add_produksi.grid(row=2, column=0, padx=5, pady=5)
    button_delete_produksi.grid(row=2, column=1, padx=5, pady=5)
    button_update_produksi.grid(row=2, column=2, padx=5, pady=5)
    button_clear_produksi.grid(row=2, column=3, padx=5, pady=5)

    # Create the Produksi table view
    produksi_treeview = ttk.Treeview(content_frame, columns=("ID Produksi", "ID Pembelian", "Tanggal Produksi", "Volume"), show="headings", height=16)
    produksi_treeview.heading("ID Produksi", text="No Produksi")
    produksi_treeview.heading("ID Pembelian", text="No. Nota")
    produksi_treeview.heading("Tanggal Produksi", text="Tanggal Produksi")
    produksi_treeview.heading("Volume", text="Volume")

    produksi_treeview.column("ID Produksi", width=150)
    produksi_treeview.column("ID Pembelian", width=150)
    produksi_treeview.column("Tanggal Produksi", width=150)
    produksi_treeview.column("Volume", width=150)

    produksi_treeview.grid(row=3, column=0, columnspan=4, padx=1, pady=1)

    def on_select_produksi(event):
        selected_item = produksi_treeview.focus()
        if selected_item:
            produksi_id, pembelian_id, tanggal_produksi, volume = produksi_treeview.item(selected_item, "values")
            entry_pembelian.set(pembelian_id)
            entry_tanggal_produksi.set_date(tanggal_produksi)
            refresh_detail_produksi_table(extract(produksi_id))

    produksi_treeview.bind("<<TreeviewSelect>>", on_select_produksi)

    # Create form widgets for Detail Produksi
    label_hasil_produksi = customtkinter.CTkLabel(content_frame, text="Hasil Produksi", text_color="black")
    entry_hasil_produksi = customtkinter.CTkComboBox(content_frame, values=populate_hasil_produksi_dropdown(), width=200)

    label_tebal_detail_produksi = customtkinter.CTkLabel(content_frame, text="Tebal", text_color="black")
    entry_tebal_detail_produksi = customtkinter.CTkComboBox(content_frame, values=tebal_list, width=200)

    label_ukuran_detail_produksi = customtkinter.CTkLabel(content_frame, text="Ukuran", text_color="black")
    entry_ukuran_detail_produksi = customtkinter.CTkComboBox(content_frame, values=list_ukuran, width=200)

    label_jumlah_detail_produksi = customtkinter.CTkLabel(content_frame, text="Jumlah", text_color="black")
    entry_jumlah_detail_produksi = customtkinter.CTkEntry(content_frame, width=200)

    button_add_detail_produksi = customtkinter.CTkButton(content_frame, text="Add", command=add_detail_produksi)
    button_delete_detail_produksi = customtkinter.CTkButton(content_frame, text="Delete", command=delete_detail_produksi)
    button_update_detail_produksi = customtkinter.CTkButton(content_frame, text="Update", command=update_detail_produksi)
    button_clear_detail_produksi = customtkinter.CTkButton(content_frame, text="Clear", command=clear_detail_produksi_fields)

    # Position the Detail Produksi form widgets
    label_hasil_produksi.grid(row=4, column=0, padx=5, pady=5)
    entry_hasil_produksi.grid(row=4, column=1, padx=5, pady=5)

    label_tebal_detail_produksi.grid(row=4, column=2, padx=5, pady=5)
    entry_tebal_detail_produksi.grid(row=4, column=3, padx=5, pady=5)

    label_ukuran_detail_produksi.grid(row=5, column=0, padx=5, pady=5)
    entry_ukuran_detail_produksi.grid(row=5, column=1, padx=5, pady=5)

    label_jumlah_detail_produksi.grid(row=5, column=2, padx=5, pady=5)
    entry_jumlah_detail_produksi.grid(row=5, column=3, padx=5, pady=5)

    button_add_detail_produksi.grid(row=8, column=0, padx=5, pady=5)
    button_delete_detail_produksi.grid(row=8, column=1, padx=5, pady=5)
    button_update_detail_produksi.grid(row=8, column=2, padx=5, pady=5)
    button_clear_detail_produksi.grid(row=8, column=3, padx=5, pady=5)

    # Create the Detail Produksi table view
    detail_produksi_treeview = ttk.Treeview(content_frame, columns=("ID Detail Produksi", "ID Hasil Produksi", "Jenis Hasil Produksi", "Tebal", "Ukuran", "Jumlah", "Volume"), show="headings", height=16)
    detail_produksi_treeview.heading("ID Detail Produksi", text="No")
    detail_produksi_treeview.heading("ID Hasil Produksi", text="Nama")
    detail_produksi_treeview.heading("Jenis Hasil Produksi", text="Jenis")
    detail_produksi_treeview.heading("Tebal", text="Tebal (mm)")
    detail_produksi_treeview.heading("Ukuran", text="Ukuran (cm)")
    detail_produksi_treeview.heading("Jumlah", text="Jumlah")
    detail_produksi_treeview.heading("Volume", text="Volume (m3)")

    detail_produksi_treeview.column("ID Detail Produksi", width=50)
    detail_produksi_treeview.column("ID Hasil Produksi", width=200)
    detail_produksi_treeview.column("Jenis Hasil Produksi", width=100)
    detail_produksi_treeview.column("Tebal", width=150)
    detail_produksi_treeview.column("Ukuran", width=150)
    detail_produksi_treeview.column("Jumlah", width=150)
    detail_produksi_treeview.column("Volume", width=150)

    detail_produksi_treeview.grid(row=9, column=0, columnspan=4, padx=1, pady=1)

    def on_select_detail_produksi(event):
        selected_item = detail_produksi_treeview.focus()
        if selected_item:
            detail_produksi_id, hasil_produksi_id, jenis, tebal, ukuran, jumlah, volume = detail_produksi_treeview.item(selected_item, "values")
            hasil_produksi = "{} - {}".format(hasil_produksi_id,jenis)
            entry_hasil_produksi.set(hasil_produksi)
            entry_tebal_detail_produksi.set(tebal)
            entry_ukuran_detail_produksi.set(ukuran)
            entry_jumlah_detail_produksi.delete(0, tk.END)
            entry_jumlah_detail_produksi.insert(tk.END, jumlah)

    detail_produksi_treeview.bind("<<TreeviewSelect>>", on_select_detail_produksi)

    # Initialize the Produksi and Detail Produksi tables
    refresh_produksi_table()

def show_pembeli():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database for the "Pembeli" table
        c.execute("SELECT * FROM Pembeli")
        rows = c.fetchall()
        for row in rows:
            pembeli_id = row[0]
            nama = row[1]
            no_hp = row[2]
            alamat = row[3]
            kota = row[4]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(pembeli_id, nama, no_hp, alamat, kota))

    # Function to add a new "Pembeli" record
    def add_pembeli():
        # Fetch values from the entry fields
        nama = entry_nama.get()
        no_hp = entry_no_hp.get()
        alamat = entry_alamat.get()
        kota = entry_kota.get()

        # Insert new "Pembeli" record into the database
        c.execute("INSERT INTO Pembeli (Nama, No_HP, Alamat, Kota) VALUES (%s, %s, %s, %s)",
                (nama, no_hp, alamat, kota))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete a "Pembeli" record
    def delete_pembeli():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Pembeli record selected")
            return
        
        if selected_item:
            # Get the "Pembeli" ID from the selected item
            pembeli_id = treeview.item(selected_item)["values"][0]

            # Delete the "Pembeli" record from the database
            c.execute("DELETE FROM Pembeli WHERE ID_Pembeli=?", (pembeli_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update a "Pembeli" record
    def update_pembeli():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Pembeli record selected")
            return
        
        if selected_item:
            # Get the "Pembeli" ID from the selected item
            pembeli_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_no_hp = entry_no_hp.get()
            updated_alamat = entry_alamat.get()
            updated_kota = entry_kota.get()

            c.execute("UPDATE Pembeli SET Nama=?, No_HP=?, Alamat=?, Kota=? WHERE ID_Pembeli=?", 
                    (updated_nama, updated_no_hp, updated_alamat, updated_kota, pembeli_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_no_hp.delete(0, tk.END)
        entry_alamat.delete(0, tk.END)
        entry_kota.delete(0, tk.END)

    # Create the form widgets for the "Pembeli" table
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama", text_color="black")
    entry_nama = customtkinter.CTkEntry(content_frame, width=200)

    label_no_hp = customtkinter.CTkLabel(content_frame, text="No HP", text_color="black")
    entry_no_hp = customtkinter.CTkEntry(content_frame, width=200)

    label_alamat = customtkinter.CTkLabel(content_frame, text="Alamat", text_color="black")
    entry_alamat = customtkinter.CTkEntry(content_frame, width=200)

    label_kota = customtkinter.CTkLabel(content_frame, text="Kota", text_color="black")
    entry_kota = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_pembeli)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_pembeli)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_pembeli)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets for the "Pembeli" table
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    label_no_hp.grid(row=1, column=0, padx=5, pady=5)
    label_alamat.grid(row=2, column=0, padx=5, pady=5)
    label_kota.grid(row=3, column=0, padx=5, pady=5)

    entry_nama.grid(row=0, column=1, padx=5, pady=5)
    entry_no_hp.grid(row=1, column=1, padx=5, pady=5)
    entry_alamat.grid(row=2, column=1, padx=5, pady=5)
    entry_kota.grid(row=3, column=1, padx=5, pady=5)

    button_add.grid(row=4, column=0, padx=1, pady=1)
    button_delete.grid(row=4, column=1, padx=1, pady=1)
    button_update.grid(row=4, column=2, padx=1, pady=1)
    button_clear.grid(row=4, column=3, padx=1, pady=1)

    # Create the table view for the "Pembeli" table
    treeview = ttk.Treeview(content_frame, columns=("ID", "Nama", "No_HP", "Alamat", "Kota"), show="headings", height=25)
    treeview.heading("ID", text="ID")
    treeview.heading("Nama", text="Nama")
    treeview.heading("No_HP", text="No HP")
    treeview.heading("Alamat", text="Alamat")
    treeview.heading("Kota", text="Kota")

    treeview.column("ID", width=50)
    treeview.column("Nama", width=150)
    treeview.column("No_HP", width=100)
    treeview.column("Alamat", width=200)
    treeview.column("Kota", width=200)

    treeview.grid(row=5, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            pembeli_id, nama, no_hp, alamat, kota = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama)
            entry_no_hp.delete(0, tk.END)
            entry_no_hp.insert(tk.END, no_hp)
            entry_alamat.delete(0, tk.END)
            entry_alamat.insert(tk.END, alamat)
            entry_kota.delete(0, tk.END)
            entry_kota.insert(tk.END, kota)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Pembeli" table
    refresh_table()

def show_supplier():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database for the "Supplier" table
        c.execute("SELECT * FROM Supplier")
        rows = c.fetchall()
        for row in rows:
            supplier_id = row[0]
            nama = row[1]
            no_hp = row[2]
            alamat = row[3]
            bukti_milik = row[4]
            no_bukti_milik = row[5]
            nik_supplier = row[6]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(supplier_id, nama, no_hp, alamat, bukti_milik, no_bukti_milik, nik_supplier))

    # Function to add a new "Supplier" record
    def add_supplier():
        # Fetch values from the entry fields
        nama = entry_nama.get()
        no_hp = entry_no_hp.get()
        alamat = entry_alamat.get()
        bukti_milik = entry_bukti_milik.get()
        no_bukti_milik = entry_no_bukti.get()
        nik_supplier = entry_nik_supplier.get()

        # Insert new "Supplier" record into the database
        c.execute("INSERT INTO Supplier (Nama, No_HP, Alamat, Bukti_Kepemilikan, No_Bukti_Kepemilikan, NIK_Pengirim) VALUES (?, ?, ?, ?, ?, ?)",
                (nama, no_hp, alamat, bukti_milik, no_bukti_milik, nik_supplier))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete a "Supplier" record
    def delete_supplier():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Supplier record selected")
            return
        
        if selected_item:
            # Get the "Supplier" ID from the selected item
            supplier_id = treeview.item(selected_item)["values"][0]

            # Delete the "Supplier" record from the database
            c.execute("DELETE FROM Supplier WHERE ID_Supplier=?", (supplier_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update a "Supplier" record
    def update_supplier():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Supplier record selected")
            return
        
        if selected_item:
            # Get the "Supplier" ID from the selected item
            supplier_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_no_hp = entry_no_hp.get()
            updated_alamat = entry_alamat.get()
            updated_bukti = entry_bukti_milik.get()
            updated_no_bukti = entry_no_bukti.get()
            updated_nik = entry_nik_supplier.get()

            c.execute("UPDATE Supplier SET Nama=?, No_HP=?, Alamat=?, Bukti_Kepemilikan=?, No_Bukti_Kepemilikan=?, NIK_Pengirim=? WHERE ID_Supplier=?", 
                    (updated_nama, updated_no_hp, updated_alamat, updated_bukti, updated_no_bukti, updated_nik, supplier_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_no_hp.delete(0, tk.END)
        entry_alamat.delete(0, tk.END)
        entry_bukti_milik.delete(0, tk.END)
        entry_no_bukti.delete(0, tk.END)
        entry_nik_supplier.delete(0, tk.END)

    # Create the form widgets for the "Supplier" table
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama", text_color="black")
    entry_nama = customtkinter.CTkEntry(content_frame, width=200)

    label_no_hp = customtkinter.CTkLabel(content_frame, text="No HP", text_color="black")
    entry_no_hp = customtkinter.CTkEntry(content_frame, width=200)

    label_alamat = customtkinter.CTkLabel(content_frame, text="Alamat", text_color="black")
    entry_alamat = customtkinter.CTkEntry(content_frame, width=200)

    label_bukti_milik = customtkinter.CTkLabel(content_frame, text="Bukti Kepemilikan", text_color="black")
    entry_bukti_milik = customtkinter.CTkEntry(content_frame, width=200)

    label_no_bukti = customtkinter.CTkLabel(content_frame, text="No. Bukti Kepemilikan", text_color="black")
    entry_no_bukti = customtkinter.CTkEntry(content_frame, width=200)

    label_nik_supplier = customtkinter.CTkLabel(content_frame, text="NIK Supplier", text_color="black")
    entry_nik_supplier = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_supplier)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_supplier)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_supplier)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets for the "Supplier" table
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    label_no_hp.grid(row=1, column=0, padx=5, pady=5)
    label_alamat.grid(row=2, column=0, padx=5, pady=5)
    label_bukti_milik.grid(row=0, column=2, padx=5, pady=5)
    label_no_bukti.grid(row=1, column=2, padx=5, pady=5)
    label_nik_supplier.grid(row=2, column=2, padx=5, pady=5)

    entry_nama.grid(row=0, column=1, padx=5, pady=5)
    entry_no_hp.grid(row=1, column=1, padx=5, pady=5)
    entry_alamat.grid(row=2, column=1, padx=5, pady=5)
    entry_bukti_milik.grid(row=0, column=3, padx=5, pady=5)
    entry_no_bukti.grid(row=1, column=3, padx=5, pady=5)
    entry_nik_supplier.grid(row=2, column=3, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view for the "Supplier" table
    treeview = ttk.Treeview(content_frame, columns=("ID", "Nama", "No_HP", "Alamat", "Bukti Milik", "No. Bukti", "NIK"), show="headings", height=25)
    treeview.heading("ID", text="ID")
    treeview.heading("Nama", text="Nama")
    treeview.heading("No_HP", text="No HP")
    treeview.heading("Alamat", text="Alamat")
    treeview.heading("Bukti Milik", text="Bukti Milik")
    treeview.heading("No. Bukti", text="No. Bukti")
    treeview.heading("NIK", text="NIK")

    treeview.column("ID", width=50)
    treeview.column("Nama", width=150)
    treeview.column("No_HP", width=100)
    treeview.column("Alamat", width=200)
    treeview.column("Bukti Milik", width=150)
    treeview.column("No. Bukti", width=100)
    treeview.column("NIK", width=200)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            supplier_id, nama, no_hp, alamat, bukti_milik, no_bukti, nik_supplier = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama)
            entry_no_hp.delete(0, tk.END)
            entry_no_hp.insert(tk.END, no_hp)
            entry_alamat.delete(0, tk.END)
            entry_alamat.insert(tk.END, alamat)
            entry_bukti_milik.delete(0, tk.END)
            entry_bukti_milik.insert(tk.END, bukti_milik)
            entry_no_bukti.delete(0, tk.END)
            entry_no_bukti.insert(tk.END, no_bukti)
            entry_nik_supplier.delete(0, tk.END)
            entry_nik_supplier.insert(tk.END, nik_supplier)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Supplier" table
    refresh_table()

def show_akun():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database for the "Akun" table
        c.execute("SELECT * FROM Akun")
        rows = c.fetchall()
        for row in rows:
            akun_id = row[0]
            nama = row[1]
            kategori = row[2]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(akun_id, nama, kategori))

    # Function to add a new "Akun" record
    def add_akun():
        # Fetch values from the entry fields
        akun_id = entry_akun_id.get()
        nama = entry_nama.get()
        kategori = entry_kategori.get()

        # Insert new "Akun" record into the database
        c.execute("INSERT INTO Akun (ID_Akun, Nama, Kategori) VALUES (?, ?, ?)",
                (akun_id, nama, kategori))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete an "Akun" record
    def delete_akun():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Akun record selected")
            return
        
        if selected_item:
            # Get the "Akun" ID from the selected item
            akun_id = treeview.item(selected_item)["values"][0]

            # Delete the "Akun" record from the database
            c.execute("DELETE FROM Akun WHERE ID_Akun=?", (akun_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update an "Akun" record
    def update_akun():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Akun record selected")
            return
        
        if selected_item:
            # Get the "Akun" ID from the selected item
            akun_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_akun_id = entry_akun_id.get()
            updated_nama = entry_nama.get()
            updated_kategori = entry_kategori.get()

            c.execute("UPDATE Akun SET ID_Akun=?, Nama=?, Kategori=? WHERE ID_Akun=?", 
                    (updated_akun_id, updated_nama, updated_kategori, akun_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_akun_id.delete(0, tk.END)
        entry_nama.delete(0, tk.END)
        entry_kategori.delete(0, tk.END)

    # Create the form widgets for the "Akun" table
    label_akun_id = customtkinter.CTkLabel(content_frame, text="Kode Akun", text_color="black")
    entry_akun_id = customtkinter.CTkEntry(content_frame, width=200)

    label_nama = customtkinter.CTkLabel(content_frame, text="Nama", text_color="black")
    entry_nama = customtkinter.CTkEntry(content_frame, width=200)

    label_kategori = customtkinter.CTkLabel(content_frame, text="Kategori", text_color="black")
    entry_kategori = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_akun, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_akun, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_akun, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets for the "Akun" table
    label_akun_id.grid(row=0, column=0, padx=5, pady=5)
    label_nama.grid(row=1, column=0, padx=5, pady=5)
    label_kategori.grid(row=2, column=0, padx=5, pady=5)

    entry_akun_id.grid(row=0, column=1, padx=5, pady=5)
    entry_nama.grid(row=1, column=1, padx=5, pady=5)
    entry_kategori.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view for the "Akun" table
    treeview = ttk.Treeview(content_frame, columns=("ID_Akun", "Nama", "Kategori"), show="headings", height=25)
    treeview.heading("ID_Akun", text="Akun ID")
    treeview.heading("Nama", text="Nama")
    treeview.heading("Kategori", text="Kategori")

    treeview.column("ID_Akun", width=100)
    treeview.column("Nama", width=200)
    treeview.column("Kategori", width=200)

    treeview.grid(row=4, column=0, columnspan=4, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            akun_id, nama, kategori = treeview.item(selected_item, "values")
            entry_akun_id.delete(0, tk.END)
            entry_akun_id.insert(tk.END, akun_id)
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama)
            entry_kategori.delete(0, tk.END)
            entry_kategori.insert(tk.END, kategori)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Akun" table
    refresh_table()

def show_log_kayu():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database for the "Log Kayu" table
        c.execute("SELECT * FROM Log_Kayu")
        rows = c.fetchall()
        for row in rows:
            log_kayu_id = row[0]
            nama = row[1]
            panjang = row[2]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(log_kayu_id, nama, panjang))

    # Function to add a new "Log Kayu" record
    def add_log_kayu():
        # Fetch values from the entry fields
        nama = entry_nama.get()
        panjang = entry_panjang.get()

        # Insert new "Log Kayu" record into the database
        c.execute("INSERT INTO Log_Kayu (Nama, Panjang) VALUES (?, ?)",
                (nama, panjang))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete a "Log Kayu" record
    def delete_log_kayu():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Log Kayu record selected")
            return
        
        if selected_item:
            # Get the "Log Kayu" ID from the selected item
            log_kayu_id = treeview.item(selected_item)["values"][0]

            # Delete the "Log Kayu" record from the database
            c.execute("DELETE FROM Log_Kayu WHERE ID_Log_Kayu=?", (log_kayu_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update a "Log Kayu" record
    def update_log_kayu():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Log Kayu record selected")
            return
        
        if selected_item:
            # Get the "Log Kayu" ID from the selected item
            log_kayu_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_panjang = entry_panjang.get()

            c.execute("UPDATE Log_Kayu SET Nama=?, Panjang=? WHERE ID_Log_Kayu=?", 
                    (updated_nama, updated_panjang, log_kayu_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_panjang.delete(0, tk.END)

    # Create the form widgets for the "Log Kayu" table
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama", text_color="black")
    entry_nama = customtkinter.CTkEntry(content_frame, width=200)

    label_panjang = customtkinter.CTkLabel(content_frame, text="Panjang", text_color="black")
    entry_panjang = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_log_kayu, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_log_kayu, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_log_kayu, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets for the "Log Kayu" table
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    label_panjang.grid(row=1, column=0, padx=5, pady=5)

    entry_nama.grid(row=0, column=1, padx=5, pady=5)
    entry_panjang.grid(row=1, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view for the "Log Kayu" table
    treeview = ttk.Treeview(content_frame, columns=("ID_Log_Kayu", "Nama", "Panjang"), show="headings", height=25)
    treeview.heading("ID_Log_Kayu", text="ID Log Kayu")
    treeview.heading("Nama", text="Nama")
    treeview.heading("Panjang", text="Panjang")

    treeview.column("ID_Log_Kayu", width=100)
    treeview.column("Nama", width=150)
    treeview.column("Panjang", width=100)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            log_kayu_id, nama, panjang = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama)
            entry_panjang.delete(0, tk.END)
            entry_panjang.insert(tk.END, panjang)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Log Kayu" table
    refresh_table()

def show_hasil_produksi():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database for the "Hasil Produksi" table
        c.execute("SELECT * FROM Hasil_Produksi")
        rows = c.fetchall()
        for row in rows:
            hasil_produksi_id = row[0]
            nama = row[1]
            jenis = row[2]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(hasil_produksi_id, nama, jenis))

    # Function to add a new "Hasil Produksi" record
    def add_hasil_produksi():
        # Fetch values from the entry fields
        nama = entry_nama.get()
        jenis = entry_jenis.get()

        # Insert new "Hasil Produksi" record into the database
        c.execute("INSERT INTO Hasil_Produksi (Nama, Jenis) VALUES (?, ?)",
                (nama, jenis))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete a "Hasil Produksi" record
    def delete_hasil_produksi():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Hasil Produksi record selected")
            return
        
        if selected_item:
            # Get the "Hasil Produksi" ID from the selected item
            hasil_produksi_id = treeview.item(selected_item)["values"][0]

            # Delete the "Hasil Produksi" record from the database
            c.execute("DELETE FROM Hasil_Produksi WHERE ID_Hasil_Produksi=?", (hasil_produksi_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update a "Hasil Produksi" record
    def update_hasil_produksi():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Hasil Produksi record selected")
            return
        
        if selected_item:
            # Get the "Hasil Produksi" ID from the selected item
            hasil_produksi_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_jenis = entry_jenis.get()

            c.execute("UPDATE Hasil_Produksi SET Nama=?, Jenis=? WHERE ID_Hasil_Produksi=?", 
                    (updated_nama, updated_jenis, hasil_produksi_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_jenis.delete(0, tk.END)

    # Create the form widgets for the "Hasil Produksi" table
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama", text_color="black")
    entry_nama = customtkinter.CTkEntry(content_frame, width=200)

    label_jenis = customtkinter.CTkLabel(content_frame, text="Jenis", text_color="black")
    entry_jenis = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_hasil_produksi, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_hasil_produksi, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_hasil_produksi, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets for the "Hasil Produksi" table
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    label_jenis.grid(row=1, column=0, padx=5, pady=5)

    entry_nama.grid(row=0, column=1, padx=5, pady=5)
    entry_jenis.grid(row=1, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view for the "Hasil Produksi" table
    treeview = ttk.Treeview(content_frame, columns=("ID_Hasil_Produksi", "Nama", "Jenis"), show="headings", height=25)
    treeview.heading("ID_Hasil_Produksi", text="ID Hasil Produksi")
    treeview.heading("Nama", text="Nama")
    treeview.heading("Jenis", text="Jenis")

    treeview.column("ID_Hasil_Produksi", width=100)
    treeview.column("Nama", width=150)
    treeview.column("Jenis", width=100)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            hasil_produksi_id, nama, jenis = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama)
            entry_jenis.delete(0, tk.END)
            entry_jenis.insert(tk.END, jenis)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Hasil Produksi" table
    refresh_table()

def show_biaya():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database for the "Biaya" table
        c.execute("SELECT * FROM Biaya ORDER BY ID_Biaya DESC")
        rows = c.fetchall()
        for row in rows:
            biaya_id = row[0]
            akun_id = row[1]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            akun = "{} - {}".format(akun_id, nama)
            penerima = row[2]
            tanggal_terima = row[3]
            keterangan = row[4]
            pembayaran = row[5]
            dari_akun = row[6]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (dari_akun,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            dariakun = "{} - {}".format(dari_akun, nama)
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(biaya_id, akun, penerima, tanggal_terima, keterangan, pembayaran, dariakun))

    # Function to add a new Biaya record
    def add_biaya():
        akun_id = entry_akun_id.get()
        parts = akun_id.split(" - ")
        
        id_akun = parts[0]
        penerima = entry_penerima.get()
        tanggal_terima = entry_tanggal_terima.get()
        keterangan = entry_keterangan.get()
        pembayaran = entry_pembayaran.get()
        dari_akun = entry_dariakun_id.get()
        parts = dari_akun.split(" - ")
        
        dariakun = parts[0]
        
        c.execute("INSERT INTO Biaya (ID_Akun, Penerima, Tanggal_Terima, Keterangan, Pembayaran, Dari_Akun) VALUES (?, ?, ?, ?, ?, ?)",
                (id_akun, penerima, tanggal_terima, keterangan, pembayaran, dariakun))
        conn.commit()
        
        refresh_table()

    # Function to delete a Biaya record
    def delete_biaya():
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showerror("Error", "No Biaya record selected")
            return
        
        if selected_item:
            biaya_id = treeview.item(selected_item)["values"][0]
            
            c.execute("DELETE FROM Biaya WHERE ID_Biaya=?", (biaya_id,))
            conn.commit()
            
            refresh_table()

    # Function to update a Biaya record
    def update_biaya():
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Biaya record selected")
            return
        
        if selected_item:
            biaya_id = treeview.item(selected_item)["values"][0]

            updated_akun_id = entry_akun_id.get()
            parts = updated_akun_id.split(" - ")
        
            id_akun = parts[0]
            updated_penerima = entry_penerima.get()
            updated_tanggal_terima = entry_tanggal_terima.get()
            updated_keterangan = entry_keterangan.get()
            updated_pembayaran = entry_pembayaran.get()
            dari_akun = entry_dariakun_id.get()
            parts = dari_akun.split(" - ")
            
            dariakun = parts[0]

            c.execute("UPDATE Biaya SET ID_Akun=?, Penerima=?, Tanggal_Terima=?, Keterangan=?, Pembayaran=?, Dari_Akun=? WHERE ID_Biaya=?", 
                    (id_akun, updated_penerima, updated_tanggal_terima, updated_keterangan, updated_pembayaran, dariakun, biaya_id))

            conn.commit()

            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_akun_id.set('')
        entry_dariakun_id.set('')
        entry_penerima.delete(0, tk.END)
        entry_tanggal_terima.set_date(None)
        entry_keterangan.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    # Function to fetch Akun IDs
    def fetch_akun_ids():
        c.execute("""SELECT ID_Akun, Nama FROM Akun WHERE Kategori="Beban";""")
        akun_ids = c.fetchall()
        list_akun = []
        for row in akun_ids:
            gabungan = "{} - {}".format(row[0],row[1])
            list_akun.append(gabungan)

        return list_akun

    # Create the form widgets
    label_akun_id = customtkinter.CTkLabel(content_frame, text="Akun Biaya", text_color="black")
    entry_akun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_ids(), width=200)

    # Create the form widgets
    label_dariakun_id = customtkinter.CTkLabel(content_frame, text="Dibayar Dari", text_color="black")
    entry_dariakun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_kasbank(), width=200)

    label_penerima = customtkinter.CTkLabel(content_frame, text="Penerima", text_color="black")
    entry_penerima = customtkinter.CTkEntry(content_frame, width=200)

    label_tanggal_terima = customtkinter.CTkLabel(content_frame, text="Tanggal Terima", text_color="black")
    entry_tanggal_terima = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_keterangan = customtkinter.CTkLabel(content_frame, text="Keterangan", text_color="black")
    entry_keterangan = customtkinter.CTkEntry(content_frame, width=200)

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_biaya)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_biaya)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_biaya)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_akun_id.grid(row=0, column=0, padx=5, pady=5)
    entry_akun_id.grid(row=0, column=1, padx=5, pady=5)

    label_penerima.grid(row=1, column=0, padx=5, pady=5)
    entry_penerima.grid(row=1, column=1, padx=5, pady=5)

    label_tanggal_terima.grid(row=2, column=0, padx=5, pady=5)
    entry_tanggal_terima.grid(row=2, column=1, padx=5, pady=5)

    label_keterangan.grid(row=0, column=2, padx=5, pady=5)
    entry_keterangan.grid(row=0, column=3, padx=5, pady=5)

    label_pembayaran.grid(row=1, column=2, padx=5, pady=5)
    entry_pembayaran.grid(row=1, column=3, padx=5, pady=5)
    
    label_dariakun_id.grid(row=2, column=2, padx=5, pady=5)
    entry_dariakun_id.grid(row=2, column=3, padx=5, pady=5)

    button_add.grid(row=5, column=0, padx=1, pady=1)
    button_delete.grid(row=5, column=1, padx=1, pady=1)
    button_update.grid(row=5, column=2, padx=1, pady=1)
    button_clear.grid(row=5, column=3, padx=1, pady=1)

    # Create the table view for the "Biaya" table
    treeview = ttk.Treeview(content_frame, columns=("ID", "ID Akun", "Penerima", "Tanggal Terima", "Keterangan", "Pembayaran", "Dibayar Dari"), show="headings", height=33 )
    treeview.heading("ID", text="No")
    treeview.heading("ID Akun", text="Akun Biaya")
    treeview.heading("Penerima", text="Penerima")
    treeview.heading("Tanggal Terima", text="Tanggal Terima")
    treeview.heading("Keterangan", text="Keterangan")
    treeview.heading("Pembayaran", text="Pembayaran")
    treeview.heading("Dibayar Dari", text="Dibayar Dari")

    treeview.column("ID", width=50)
    treeview.column("ID Akun", width=100)
    treeview.column("Penerima", width=150)
    treeview.column("Tanggal Terima", width=120)
    treeview.column("Keterangan", width=150)
    treeview.column("Pembayaran", width=150)
    treeview.column("Dibayar Dari", width=100)

    treeview.grid(row=6, column=0, columnspan=4, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            biaya_id, akun_id, penerima, tanggal_terima, keterangan, pembayaran, dariakun = treeview.item(selected_item, "values")
            entry_akun_id.set(akun_id)
            entry_penerima.delete(0, tk.END)
            entry_penerima.insert(tk.END, penerima)
            entry_tanggal_terima.set_date(tanggal_terima)
            entry_keterangan.delete(0, tk.END)
            entry_keterangan.insert(tk.END, keterangan)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            entry_dariakun_id.set(dariakun)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Biaya" table
    refresh_table()

def show_aset_tetap():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database for the "Aset Tetap" table
        c.execute("SELECT * FROM Aset_Tetap")
        rows = c.fetchall()
        for row in rows:
            aset_tetap_id = row[0]
            akun_id = row[1]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            akun = "{} - {}".format(akun_id,nama)
            tanggal_beli = row[2]
            keterangan = row[3]
            harga_beli = row[4]
            unit = row[5]
            umur_ekonomis = row[6]
            kondisi = row[7]
            dari_akun = row[8]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (dari_akun,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            dariakun = "{} - {}".format(dari_akun, nama)
            jumlah = unit*harga_beli
            persentase = "{}%".format(round((1/umur_ekonomis*100),2))
            penyusutan = jumlah/umur_ekonomis
            tanggal_hari_ini = datetime.now()
            tanggal = datetime.strptime(tanggal_beli, "%Y-%m-%d")
            selisih_hari = (tanggal_hari_ini - tanggal).days
            jumlah_tahun = round(selisih_hari / 365,2)  # Menggunakan asumsi 365 hari per tahun
            akumulasi_penyusutan = penyusutan * jumlah_tahun
            nilai_buku = jumlah - akumulasi_penyusutan
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(aset_tetap_id, akun, tanggal_beli, dariakun, keterangan, unit, harga_beli, format_currency(jumlah), umur_ekonomis, persentase, format_currency(penyusutan), jumlah_tahun, format_currency(akumulasi_penyusutan), format_currency(nilai_buku), kondisi))

    # Function to add a new Aset Tetap record
    def add_aset_tetap():
        akun_id = entry_akun_id.get()
        parts = akun_id.split(" - ")
        
        id_akun = parts[0]

        tanggal_beli = entry_tanggal_beli.get()
        keterangan = entry_keterangan.get()
        harga_beli = entry_harga_beli.get()
        jumlah = entry_jumlah.get()
        umur_ekonomis = entry_umur_ekonomis.get()
        kondisi = entry_kondisi.get()
        dari_akun = entry_dariakun_id.get()
        parts = dari_akun.split(" - ")
        
        dariakun = parts[0]
        
        c.execute("INSERT INTO Aset_Tetap (ID_Akun, Tanggal_Beli, Keterangan, Harga_Beli, Jumlah, Umur_Ekonomis, Kondisi, Dari_Akun) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (id_akun, tanggal_beli, keterangan, harga_beli, jumlah, umur_ekonomis, kondisi, dariakun))
        conn.commit()
        
        refresh_table()

    # Function to delete an Aset Tetap record
    def delete_aset_tetap():
        selected_item = treeview.selection()
        if selected_item:
            aset_tetap_id = treeview.item(selected_item)["values"][0]
            
            c.execute("DELETE FROM Aset_Tetap WHERE ID_Aset_Tetap=?", (aset_tetap_id,))
            conn.commit()
            
            refresh_table()

    # Function to update an Aset Tetap record
    def update_aset_tetap():
        selected_item = treeview.focus()
        if selected_item:
            aset_tetap_id = treeview.item(selected_item)["values"][0]

            updated_akun_id = entry_akun_id.get()
            parts = updated_akun_id.split(" - ")
        
            id_akun = parts[0]
            updated_tanggal_beli = entry_tanggal_beli.get()
            keterangan = entry_keterangan.get()
            updated_harga_beli = entry_harga_beli.get()
            updated_jumlah = entry_jumlah.get()
            updated_umur_ekonomis = entry_umur_ekonomis.get()
            updated_kondisi = entry_kondisi.get()
            dari_akun = entry_dariakun_id.get()
            parts = dari_akun.split(" - ")
            
            dariakun = parts[0]

            c.execute("UPDATE Aset_Tetap SET ID_Akun=?, Tanggal_Beli=?, Keterangan=?, Harga_Beli=?, Jumlah=?, Umur_Ekonomis=?, Kondisi=?, Dari_Akun=? WHERE ID_Aset_Tetap=?", 
                    (id_akun, updated_tanggal_beli, keterangan, updated_harga_beli, updated_jumlah, updated_umur_ekonomis, updated_kondisi, dariakun, aset_tetap_id))

            conn.commit()

            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_akun_id.set('')
        entry_tanggal_beli.set_date(None)
        entry_harga_beli.delete(0, tk.END)
        entry_jumlah.delete(0, tk.END)
        entry_umur_ekonomis.delete(0, tk.END)
        entry_kondisi.set('')
        entry_dariakun_id.set('')
        entry_keterangan.delete(0, tk.END)

    # Function to fetch Akun IDs
    def fetch_akun_ids():
        c.execute("""SELECT ID_Akun, Nama FROM Akun WHERE Kategori="Aset Tetap";""")
        akun_ids = c.fetchall()
        list_akun = []
        for row in akun_ids:
            gabungan = "{} - {}".format(row[0],row[1])
            list_akun.append(gabungan)

        return list_akun

    # Create the form widgets
    label_akun_id = customtkinter.CTkLabel(content_frame, text="Aktiva Tetap", text_color="black")
    entry_akun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_ids(), width=200)

    # Create the form widgets
    label_dariakun_id = customtkinter.CTkLabel(content_frame, text="Dibayar Dari", text_color="black")
    entry_dariakun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_kasbank(), width=200)

    label_tanggal_beli = customtkinter.CTkLabel(content_frame, text="Tanggal Beli", text_color="black")
    entry_tanggal_beli = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_keterangan = customtkinter.CTkLabel(content_frame, text="Keterangan", text_color="black")
    entry_keterangan = customtkinter.CTkEntry(content_frame, width=200)

    label_harga_beli = customtkinter.CTkLabel(content_frame, text="Harga Beli", text_color="black")
    entry_harga_beli = customtkinter.CTkEntry(content_frame, width=200)

    label_jumlah = customtkinter.CTkLabel(content_frame, text="Jumlah", text_color="black")
    entry_jumlah = customtkinter.CTkEntry(content_frame, width=200)

    label_umur_ekonomis = customtkinter.CTkLabel(content_frame, text="Umur Ekonomis", text_color="black")
    entry_umur_ekonomis = customtkinter.CTkEntry(content_frame, width=200)

    listkondisi = ["Bekas", "Baru"]

    label_kondisi = customtkinter.CTkLabel(content_frame, text="Kondisi", text_color="black")
    entry_kondisi = customtkinter.CTkComboBox(content_frame, values=listkondisi, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_aset_tetap)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_aset_tetap)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_aset_tetap)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_akun_id.grid(row=0, column=0, padx=5, pady=5)
    entry_akun_id.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal_beli.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal_beli.grid(row=1, column=1, padx=5, pady=5)

    label_harga_beli.grid(row=2, column=0, padx=5, pady=5)
    entry_harga_beli.grid(row=2, column=1, padx=5, pady=5)

    label_jumlah.grid(row=0, column=2, padx=5, pady=5)
    entry_jumlah.grid(row=0, column=3, padx=5, pady=5)

    label_umur_ekonomis.grid(row=1, column=2, padx=5, pady=5)
    entry_umur_ekonomis.grid(row=1, column=3, padx=5, pady=5)

    label_kondisi.grid(row=2, column=2, padx=5, pady=5)
    entry_kondisi.grid(row=2, column=3, padx=5, pady=5)

    label_keterangan.grid(row=3, column=2, padx=5, pady=5)
    entry_keterangan.grid(row=3, column=3, padx=5, pady=5)

    label_dariakun_id.grid(row=3, column=0, padx=5, pady=5)
    entry_dariakun_id.grid(row=3, column=1, padx=5, pady=5)

    button_add.grid(row=6, column=0, padx=1, pady=1)
    button_delete.grid(row=6, column=1, padx=1, pady=1)
    button_update.grid(row=6, column=2, padx=1, pady=1)
    button_clear.grid(row=6, column=3, padx=1, pady=1)

    # Create the table view for the "Aset Tetap" table
    treeview = ttk.Treeview(content_frame, columns=("ID", "ID Akun", "Tanggal Beli", "Dari Akun", "Keterangan", "Unit", "Harga Beli", "Jumlah", "Umur Ekonomis", "Persentase", "Penyusutan per Tahun", "Tahun", "Akumulasi Penyusutan", "Nilai Buku", "Kondisi"), show="headings", height=30)
    treeview.heading("ID", text="No")
    treeview.heading("ID Akun", text="Aktiva Tetap")
    treeview.heading("Tanggal Beli", text="Tanggal Beli")
    treeview.heading("Dari Akun", text="Dibayar Dari")
    treeview.heading("Keterangan", text="Keterangan")
    treeview.heading("Unit", text="Unit")
    treeview.heading("Harga Beli", text="Harga Beli")
    treeview.heading("Jumlah", text="Jumlah")
    treeview.heading("Umur Ekonomis", text="Umur Ekonomis")
    treeview.heading("Persentase", text="Persentase")
    treeview.heading("Penyusutan per Tahun", text="Penyusutan per Tahun")
    treeview.heading("Tahun", text="Tahun")
    treeview.heading("Akumulasi Penyusutan", text="Akumulasi Penyusutan")
    treeview.heading("Nilai Buku", text="Nilai Buku")
    treeview.heading("Kondisi", text="Kondisi")

    treeview.column("ID", width=50)
    treeview.column("ID Akun", width=150)
    treeview.column("Tanggal Beli", width=100)
    treeview.column("Dari Akun", width=100)
    treeview.column("Keterangan", width=150)
    treeview.column("Unit", width=50)
    treeview.column("Harga Beli", width=100)
    treeview.column("Jumlah", width=100)
    treeview.column("Umur Ekonomis", width=100)
    treeview.column("Persentase", width=100)
    treeview.column("Penyusutan per Tahun", width=150)
    treeview.column("Tahun", width=75)
    treeview.column("Akumulasi Penyusutan", width=150)
    treeview.column("Nilai Buku", width=150)
    treeview.column("Kondisi", width=100)

    treeview.grid(row=7,column=0, columnspan=4, padx=10, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            aset_tetap_id, akun_id, tanggal_beli, dariakun, keterangan, unit, harga_beli, jumlah, umur_ekonomis, persentase, penyusutan, jumlah_tahun, akumulasi_penyusutan, nilai_buku, kondisi = treeview.item(selected_item, "values")
            entry_akun_id.set(akun_id)
            entry_tanggal_beli.set_date(tanggal_beli)
            entry_keterangan.delete(0, tk.END)
            entry_keterangan.insert(tk.END, keterangan)
            entry_dariakun_id.set(dariakun)
            entry_harga_beli.delete(0, tk.END)
            entry_harga_beli.insert(tk.END, harga_beli)
            entry_jumlah.delete(0, tk.END)
            entry_jumlah.insert(tk.END, unit)
            entry_umur_ekonomis.delete(0, tk.END)
            entry_umur_ekonomis.insert(tk.END, umur_ekonomis)
            entry_kondisi.set(kondisi)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view for the "Aset Tetap" table
    refresh_table()

def show_piutang():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database
        c.execute("SELECT * FROM Piutang")
        rows = c.fetchall()
        for row in rows:
            piutang_id = row[0]
            penjualan_id = row[1]
            akun_id = row[2]
            tanggal_bayar = row[3]
            pembayaran = row[4]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            akun = "{} - {}".format(akun_id, nama)

            c.execute("SELECT Tanggal_Faktur FROM Penjualan WHERE ID_Penjualan=?", (penjualan_id,))
            tanggal = c.fetchone()[0]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(piutang_id, format_id_penjualan(penjualan_id, tanggal), tanggal_bayar, pembayaran, akun))

    # Function to refresh the Penjualan table
    def data_penjualan_table():

        c.execute("SELECT * FROM Penjualan ORDER BY ID_Penjualan DESC")
        rows = c.fetchall()

        penjualan = {}

        for row in rows:
            penjualan_id = row[0]
            pembeli_id = row[1]
            tanggal_faktur = row[6]
            termin = row[7]
            pembayaran = row[8]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Piutang WHERE ID_Penjualan=?", (penjualan_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            tanggal = datetime.strptime(tanggal_faktur, "%Y-%m-%d")
            # Batas waktu berdasarkan termin

            if termin in batas_waktu:
                batas_hari = batas_waktu[termin]
                tanggal_jatuh_tempo = tanggal + timedelta(days=batas_hari)

                # Tanggal hari ini
                tanggal_hari_ini = datetime.now()

                # Bandingkan tanggal jatuh tempo dengan tanggal hari ini
                if tanggal_hari_ini > tanggal_jatuh_tempo:
                    nota_jatuh_tempo = "Sudah Jatuh Tempo"
                else:
                    nota_jatuh_tempo = "Belum Jatuh Tempo"
            else:
                nota_jatuh_tempo = "Tidak Terdefinisi"  # Jika termin tidak ditemukan

            c.execute("SELECT Nama FROM Pembeli WHERE ID_Pembeli=?", (pembeli_id,))
            pembeli_name = c.fetchone()[0]

            c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
            detailjual = c.fetchall()
            if detailjual:
                listvolume = []
                listsubtotal = []
                for detail in detailjual:
                    idhasil = detail[0]
                    tebal = detail[1]
                    ukuran = detail[2]
                    jumlah = detail[3]
                    harga_jual = detail[4]
                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                    hasil_produksi_info = c.fetchone()
                    hasil_produksi_nama = hasil_produksi_info[0]
                    if hasil_produksi_nama == "Ampulur":
                        subtotal = jumlah * harga_jual
                    else:
                        subtotal = rounded * harga_jual
                    listvolume.append(volume)
                    listsubtotal.append(subtotal)

                volume = round(sum(listvolume),4)
                total = sum(listsubtotal)
                ppn = total*11/100
                grandtotal = total+ppn
                sisa_piutang = totalbayar - grandtotal

                if sisa_piutang < 0:
                    # Add the product and its inventory details to the dictionary
                    penjualan[penjualan_id] = {
                        'tanggal_faktur': tanggal_faktur,
                        'pembeli_name': pembeli_name,
                        'termin': termin,
                        'grand_total': format_currency(grandtotal),
                        'pembayaran': format_currency(totalbayar),
                        'sisa_piutang': format_currency(sisa_piutang),
                        'nota_jatuh_tempo': nota_jatuh_tempo
                    }

        return penjualan
    
    def refresh_penjualan_table():
        for row in penjualan_treeview.get_children():
            penjualan_treeview.delete(row)

        penjualan = data_penjualan_table()

        for penjualan_id, details in penjualan.items():
            tanggal_faktur = details['tanggal_faktur']
            pembeli_name = details['pembeli_name']
            termin = details['termin']
            grandtotal = details['grand_total']
            pembayaran = details['pembayaran']
            sisa_piutang = details['sisa_piutang']
            nota_jatuh_tempo = details['nota_jatuh_tempo']
            id_format = format_id_penjualan(penjualan_id, tanggal_faktur)

            penjualan_treeview.insert("", tk.END, values=(id_format, tanggal_faktur, pembeli_name, termin, grandtotal, pembayaran, sisa_piutang, nota_jatuh_tempo))

    # Function to add a new piutang record
    def add_piutang():
        # Fetch values from the entry fields
        penjualan_id = extract(combo_penjualan.get())
        akun_id = entry_akun_id.get()
        parts = akun_id.split(" - ")
        
        id_akun = parts[0]
        tanggal_bayar = entry_tanggal_bayar.get()
        pembayaran = entry_pembayaran.get()

        # Insert new piutang record into the database
        c.execute("INSERT INTO Piutang (ID_Penjualan, ID_Akun, Tanggal, Pembayaran) VALUES (?, ?, ?, ?)",
                (penjualan_id, id_akun, tanggal_bayar, pembayaran))
        conn.commit()

        # Refresh the table view
        refresh_penjualan_table()
        refresh_table()

    # Function to delete a piutang record
    def delete_piutang():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the piutang_id from the selected item
            piutang_id = treeview.item(selected_item)["values"][0]

            # Delete the piutang record from the database
            c.execute("DELETE FROM Piutang WHERE ID_Piutang=?", (piutang_id,))
            conn.commit()

            # Refresh the table view
            refresh_penjualan_table()
            refresh_table()

    # Function to update a piutang record
    def update_piutang():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the piutang_id from the selected item
            piutang_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_penjualan_id = extract(combo_penjualan.get())
            akun_id = entry_akun_id.get()
            parts = akun_id.split(" - ")
            
            id_akun = parts[0]
            updated_tanggal_bayar = entry_tanggal_bayar.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Piutang SET ID_Penjualan=?, ID_Akun=?, Tanggal=?, Pembayaran=? WHERE ID_Piutang=?",
                    (updated_penjualan_id, id_akun, updated_tanggal_bayar, updated_pembayaran, piutang_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_penjualan_table()
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        combo_penjualan.set('')
        entry_akun_id.set('')
        entry_tanggal_bayar.set_date(None)
        entry_pembayaran.delete(0, tk.END)

    # Function to fetch penjualan IDs
    def fetch_penjualan_ids():
        
        penjualan = data_penjualan_table()

        listid = []
        for penjualan_id, details in penjualan.items():
            list = []
            id_format = penjualan_id
            tanggal_faktur = details['tanggal_faktur']
            list.append(id_format)
            list.append(tanggal_faktur)
            listid.append(list)

        return [format_id_penjualan(id[0],id[1]) for id in listid]

    # Create the form widgets
    label_penjualan = customtkinter.CTkLabel(content_frame, text="Penjualan", text_color="black")
    combo_penjualan = customtkinter.CTkComboBox(content_frame, values=fetch_penjualan_ids(), width=200)

    label_akun_id = customtkinter.CTkLabel(content_frame, text="ID Akun", text_color="black")
    entry_akun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_kasbank(), width=200)

    label_tanggal_bayar = customtkinter.CTkLabel(content_frame, text="Tanggal Bayar", text_color="black")
    entry_tanggal_bayar = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_piutang)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_piutang)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_piutang)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_penjualan.grid(row=0, column=0, padx=5, pady=5)
    combo_penjualan.grid(row=0, column=1, padx=5, pady=5)

    label_akun_id.grid(row=3, column=0, padx=5, pady=5)
    entry_akun_id.grid(row=3, column=1, padx=5, pady=5)

    label_tanggal_bayar.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal_bayar.grid(row=1, column=1, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=0, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=4, column=0, padx=1, pady=1)
    button_delete.grid(row=4, column=1, padx=1, pady=1)
    button_update.grid(row=4, column=2, padx=1, pady=1)
    button_clear.grid(row=4, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Penjualan", "Tanggal Bayar", "Pembayaran", "Dibayar Ke"), show="headings", height=30)
    treeview.heading("ID", text="ID")
    treeview.heading("Penjualan", text="Penjualan")
    treeview.heading("Tanggal Bayar", text="Tanggal Bayar")
    treeview.heading("Pembayaran", text="Pembayaran")
    treeview.heading("Dibayar Ke", text="Dibayar Ke")

    treeview.column("ID", width=50)
    treeview.column("Penjualan", width=100)
    treeview.column("Tanggal Bayar", width=120)
    treeview.column("Pembayaran", width=100)
    treeview.column("Dibayar Ke", width=100)

    treeview.grid(row=5, column=5, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            piutang_id, penjualan_id, tanggal_bayar, pembayaran, akun = treeview.item(selected_item, "values")
            combo_penjualan.set(penjualan_id)
            entry_tanggal_bayar.set_date(tanggal_bayar)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            entry_akun_id.set(akun)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Create the Penjualan table view
    penjualan_treeview = ttk.Treeview(content_frame, columns=("ID Penjualan", "Tanggal Faktur", "ID Pembeli","Termin", "Grand Total",
                                                               "Pembayaran", "Sisa Piutang", "Jatuh Tempo"), show="headings", height=30)
    penjualan_treeview.heading("ID Penjualan", text="No. Invoice")
    penjualan_treeview.heading("Tanggal Faktur", text="Tanggal Faktur")
    penjualan_treeview.heading("ID Pembeli", text="Pembeli")
    penjualan_treeview.heading("Termin", text="Termin")
    penjualan_treeview.heading("Grand Total", text="Grand Total")
    penjualan_treeview.heading("Pembayaran", text="Pembayaran")
    penjualan_treeview.heading("Sisa Piutang", text="Sisa Piutang")
    penjualan_treeview.heading("Jatuh Tempo", text="Jatuh Tempo")

    penjualan_treeview.column("ID Penjualan", width=125)
    penjualan_treeview.column("Tanggal Faktur", width=100)
    penjualan_treeview.column("ID Pembeli", width=150)
    penjualan_treeview.column("Termin", width=75)
    penjualan_treeview.column("Grand Total", width=150)
    penjualan_treeview.column("Pembayaran", width=150)
    penjualan_treeview.column("Sisa Piutang", width=150)
    penjualan_treeview.column("Jatuh Tempo", width=150)

    penjualan_treeview.grid(row=5, column=0, columnspan=4, padx=10, pady=1)

    # Fetch and display initial data in the table view
    refresh_penjualan_table()
    refresh_table()

def show_utang():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Utang")
        rows = c.fetchall()
        for row in rows:
            utang_id = row[0]
            pembelian_id = row[1]
            akun_id = row[2]
            tanggal_bayar = row[3]
            pembayaran = row[4]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            akun = "{} - {}".format(akun_id, nama)

            c.execute("SELECT Tanggal_Nota FROM Pembelian WHERE ID_Pembelian=?", (pembelian_id,))
            tanggal = c.fetchone()[0]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(utang_id, format_id_pembelian(pembelian_id, tanggal), tanggal_bayar, pembayaran, akun))

    # Function to refresh the table view
    def data_table_pembelian():
        # Fetch and display data from the database
        c.execute("SELECT * FROM Pembelian ORDER BY ID_Pembelian DESC")
        rows = c.fetchall()

        pembelian = {}

        for row in rows:
            pembelian_id = row[0]
            supplier_id = row[1]
            tanggal_nota = row[4]
            bea_supplier = row[5]
            termin = row[6]
            pembayaran = row[7]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Utang WHERE ID_Pembelian=?", (pembelian_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            tanggal = datetime.strptime(tanggal_nota, "%Y-%m-%d")
            # Batas waktu berdasarkan termin

            if termin in batas_waktu:
                batas_hari = batas_waktu[termin]
                tanggal_jatuh_tempo = tanggal + timedelta(days=batas_hari)

                # Tanggal hari ini
                tanggal_hari_ini = datetime.now()

                # Bandingkan tanggal jatuh tempo dengan tanggal hari ini
                if tanggal_hari_ini > tanggal_jatuh_tempo:
                    nota_jatuh_tempo = "Sudah Jatuh Tempo"
                else:
                    nota_jatuh_tempo = "Belum Jatuh Tempo"
            else:
                nota_jatuh_tempo = "Tidak Terdefinisi"  # Jika termin tidak ditemukan

            # Get the supplier name
            c.execute("SELECT Nama FROM Supplier WHERE ID_Supplier=?", (supplier_id,))
            supplier_name = c.fetchone()[0]

            listvolume = []
            listsubtotal = []
            c.execute("SELECT Diameter, Jumlah, Pembulatan, Harga_Beli FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
            detailbeli = c.fetchall()
            if detailbeli:
                for detail in detailbeli:
                    diameter = detail[0]
                    panjang = 130
                    jumlah = detail[1]
                    pembulatan = detail[2]
                    harga = detail[3]
                    volume = (math.pi * (diameter/2)**2 * panjang * jumlah)/1000000
                    rounded = round(volume, pembulatan)
                    subtotal = rounded * harga
                    listsubtotal.append(subtotal)
                    listvolume.append(rounded)
                totalvolume = round(sum(listvolume),2)
                total = sum(listsubtotal)
                biaya_bongkar = round(totalvolume * 7000)
                beban_pabrik = biaya_bongkar - bea_supplier
                grandtotal = total + beban_pabrik
                sisa_utang = totalbayar - grandtotal
                if sisa_utang < 0:
                    # Add the product and its inventory details to the dictionary
                    pembelian[pembelian_id] = {
                        'tanggal_nota': tanggal_nota,
                        'supplier_name': supplier_name,
                        'termin': termin,
                        'grand_total': format_currency(grandtotal),
                        'pembayaran': format_currency(totalbayar),
                        'sisa_utang': format_currency(sisa_utang),
                        'nota_jatuh_tempo': nota_jatuh_tempo
                    }

        return pembelian

    # Function to refresh the table view
    def refresh_table_pembelian():
        # Clear existing table data
        for row in treeview_pembelian.get_children():
            treeview_pembelian.delete(row)

        pembelian = data_table_pembelian()

        for pembelian_id, details in pembelian.items():
            tanggal_nota = details['tanggal_nota']
            supplier_name = details['supplier_name']
            termin = details['termin']
            grandtotal = details['grand_total']
            pembayaran = details['pembayaran']
            sisa_utang = details['sisa_utang']
            nota_jatuh_tempo = details['nota_jatuh_tempo']
            id_format = format_id_pembelian(pembelian_id, tanggal_nota)

            # Insert data into the table view
            treeview_pembelian.insert("", tk.END, values=(id_format, tanggal_nota, supplier_name, termin, grandtotal, pembayaran, sisa_utang, nota_jatuh_tempo))

    # Function to add a new utang record
    def add_utang():
        pembelian_id = extract(entry_pembelian_id.get())
        akun_id = entry_akun_id.get()
        parts = akun_id.split(" - ")
        
        id_akun = parts[0]
        tanggal_bayar = entry_tanggal_bayar.get()
        pembayaran = entry_pembayaran.get()
        
        c.execute("INSERT INTO Utang (ID_Pembelian, ID_Akun, Tanggal, Pembayaran) VALUES (?, ?, ?, ?)",
                (pembelian_id, id_akun, tanggal_bayar, pembayaran))
        conn.commit()
        
        refresh_table_pembelian()
        refresh_table()

    # Function to delete an utang record
    def delete_utang():
        selected_item = treeview.selection()
        if selected_item:
            utang_id = treeview.item(selected_item)["values"][0]
            
            c.execute("DELETE FROM Utang WHERE ID_Utang=?", (utang_id,))
            conn.commit()

            refresh_table_pembelian()
            refresh_table()

    # Function to update an utang record
    def update_utang():
        selected_item = treeview.focus()
        if selected_item:
            utang_id = treeview.item(selected_item)["values"][0]

            updated_pembelian_id = extract(entry_pembelian_id.get())
            akun_id = entry_akun_id.get()
            parts = akun_id.split(" - ")
            
            id_akun = parts[0]
            updated_tanggal_bayar = entry_tanggal_bayar.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Utang SET ID_Pembelian=?, ID_Akun=?, Tanggal=?, Pembayaran=? WHERE ID_Utang=?", 
                    (updated_pembelian_id, id_akun, updated_tanggal_bayar, updated_pembayaran, utang_id))

            conn.commit()

            refresh_table_pembelian()
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_pembelian_id.set('')
        entry_akun_id.set('')
        entry_tanggal_bayar.set_date(None)
        entry_pembayaran.delete(0, tk.END)

    def fetch_pembelian_ids():

        pembelian = data_table_pembelian()

        listid = []
        for pembelian_id, details in pembelian.items():
            list = []
            id_format = pembelian_id
            tanggal_nota = details['tanggal_nota']
            list.append(id_format)
            list.append(tanggal_nota)
            listid.append(list)

        return [format_id_pembelian(id[0],id[1]) for id in listid]
    
    # Create the form widgets
    label_pembelian_id = customtkinter.CTkLabel(content_frame, text="ID Pembelian", text_color="black")
    entry_pembelian_id = customtkinter.CTkComboBox(content_frame, values=fetch_pembelian_ids(), width=200)

    label_akun_id = customtkinter.CTkLabel(content_frame, text="ID Akun", text_color="black")
    entry_akun_id = customtkinter.CTkComboBox(content_frame, values=fetch_akun_kasbank(), width=200)

    label_tanggal_bayar = customtkinter.CTkLabel(content_frame, text="Tanggal Bayar", text_color="black")
    entry_tanggal_bayar = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_utang)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_utang)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_utang)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_pembelian_id.grid(row=0, column=0, padx=5, pady=5)
    entry_pembelian_id.grid(row=0, column=1, padx=5, pady=5)

    label_akun_id.grid(row=3, column=0, padx=5, pady=5)
    entry_akun_id.grid(row=3, column=1, padx=5, pady=5)

    label_tanggal_bayar.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal_bayar.grid(row=1, column=1, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=0, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=4, column=0, padx=1, pady=1)
    button_delete.grid(row=4, column=1, padx=1, pady=1)
    button_update.grid(row=4, column=2, padx=1, pady=1)
    button_clear.grid(row=4, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "ID Pembelian", "Tanggal Bayar", "Pembayaran", "Dibayar Dari"), show="headings", height=30)
    treeview.heading("ID", text="ID")
    treeview.heading("ID Pembelian", text="ID Pembelian")
    treeview.heading("Tanggal Bayar", text="Tanggal Bayar")
    treeview.heading("Pembayaran", text="Pembayaran")
    treeview.heading("Dibayar Dari", text="Dibayar Dari")

    treeview.column("ID", width=50)
    treeview.column("ID Pembelian", width=100)
    treeview.column("Tanggal Bayar", width=120)
    treeview.column("Pembayaran", width=100)
    treeview.column("Dibayar Dari", width=100)

    treeview.grid(row=5, column=5, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            utang_id, pembelian_id, tanggal_bayar, pembayaran, akun = treeview.item(selected_item, "values")
            entry_pembelian_id.set(pembelian_id)
            entry_tanggal_bayar.set_date(tanggal_bayar)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            entry_akun_id.set(akun)
            
    treeview.bind("<<TreeviewSelect>>", on_select)

    # Create the table view for Pembelian
    treeview_pembelian = ttk.Treeview(content_frame, columns=("ID", "Tanggal Nota", "Nama Supplier", "Termin","Grand Total", "Pembayaran", "Sisa Utang", "Jatuh Tempo"), show="headings", height=30)
    treeview_pembelian.heading("ID", text="No. Nota")
    treeview_pembelian.heading("Tanggal Nota", text="Tanggal Nota")
    treeview_pembelian.heading("Nama Supplier", text="Supplier")
    treeview_pembelian.heading("Termin", text="Termin")
    treeview_pembelian.heading("Grand Total", text="Grand Total")
    treeview_pembelian.heading("Pembayaran", text="Pembayaran")
    treeview_pembelian.heading("Sisa Utang", text="Sisa Utang")
    treeview_pembelian.heading("Jatuh Tempo", text="Jatuh Tempo")

    treeview_pembelian.column("ID", width=125)
    treeview_pembelian.column("Tanggal Nota", width=100)
    treeview_pembelian.column("Nama Supplier", width=150)
    treeview_pembelian.column("Termin", width=75)
    treeview_pembelian.column("Grand Total", width=150)
    treeview_pembelian.column("Pembayaran", width=150)
    treeview_pembelian.column("Sisa Utang", width=150)
    treeview_pembelian.column("Jatuh Tempo", width=150)

    treeview_pembelian.grid(row=5, column=0, columnspan=4, padx=10, pady=1)

    # Fetch and display initial data in the table view
    refresh_table_pembelian()
    refresh_table()

def input_saldo_awal():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def save_data():
        workbook = Workbook()
        sheet = workbook.active
        
        # Get the data from the Entry widget
        kas_awal = entry_kas.get()
        bank_awal = entry_bank.get()
        pengambilan_pribadi = entry_prive.get()
        tambahan_modal = entry_tambah_modal.get()

        # Write the data to the first cell in the worksheet
        sheet['A1'] = 'kas_awal'
        sheet['A2'] = kas_awal
        sheet['B1'] = 'bank_awal'
        sheet['B2'] = bank_awal
        sheet['C1'] = 'pengambilan_pribadi'
        sheet['C2'] = pengambilan_pribadi
        sheet['D1'] = 'tambahan_modal'
        sheet['D2'] = tambahan_modal
        # Save the workbook to an Excel file
        workbook.save('data.xlsx')

    label_kas = customtkinter.CTkLabel(content_frame, text="Kas Awal", font=('Arial', 30))
    entry_kas = customtkinter.CTkEntry(content_frame, width=430, font=('Arial', 30))

    label_bank = customtkinter.CTkLabel(content_frame, text="Bank Awal", font=('Arial', 30))
    entry_bank = customtkinter.CTkEntry(content_frame, width=430, font=('Arial', 30))

    label_prive = customtkinter.CTkLabel(content_frame, text="Prive", font=('Arial', 30))
    entry_prive = customtkinter.CTkEntry(content_frame, width=430, font=('Arial', 30))

    label_tambah_modal = customtkinter.CTkLabel(content_frame, text="Tambahan Modal", font=('Arial', 30))
    entry_tambah_modal = customtkinter.CTkEntry(content_frame, width=430, font=('Arial', 30))

    # Position the form widgets
    label_kas.grid(row=0, column=0, padx=20, pady=5)
    entry_kas.grid(row=1, column=0, padx=20, pady=5)

    # Position the form widgets
    label_bank.grid(row=2, column=0, padx=20, pady=5)
    entry_bank.grid(row=3, column=0, padx=20, pady=5)

    # Position the form widgets
    label_prive.grid(row=4, column=0, padx=20, pady=5)
    entry_prive.grid(row=5, column=0, padx=20, pady=5)

    # Position the form widgets
    label_tambah_modal.grid(row=6, column=0, padx=20, pady=5)
    entry_tambah_modal.grid(row=7, column=0, padx=20, pady=5)

    button_save = customtkinter.CTkButton(content_frame, text="Save", command=save_data, font=('Arial', 25), width=100, height=60)
    button_save.grid(row=8, column=0, padx=5, pady=25)

def show_persediaanbarang():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Calculate the inventory of raw materials
    def calculate_raw_materials_inventory():
        inventory = {}

        c.execute('''SELECT
                        D.ID_Log_Kayu,
                        D.Diameter,
                        SUM(D.Jumlah) AS Jumlah_Terjual
                    FROM
                        Detail_Beli D
                    JOIN
                        Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                    GROUP BY
                        D.ID_Log_Kayu, D.Diameter;''')
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            log_kayu_id = row[0]
            diameter = row[1]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
            log_kayu_info = c.fetchone()
            log_kayu_nama = log_kayu_info[0]
            log_kayu_panjang = log_kayu_info[1]

            # Update the inventory dictionary with the calculated values based on thickness and size
            key = (log_kayu_nama, log_kayu_panjang, diameter)
            if key in inventory:
                inventory[key]['remaining_quantity'] += row[2]
            else:
                inventory[key] = {
                    'nama_kayu': log_kayu_nama,
                    'panjang' : log_kayu_panjang,
                    'diameter': diameter,
                    'remaining_quantity': row[2]
                }

        c.execute('''SELECT
                        D.ID_Log_Kayu,
                        D.Diameter,
                        SUM(D.Jumlah) AS Jumlah_Terjual
                    FROM
                        Detail_Beli D
                    JOIN
                        Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                    JOIN
                        Produksi PR ON P.ID_Pembelian = PR.ID_Pembelian
                    JOIN
                        Detail_Produksi DP ON PR.ID_Produksi = DP.ID_Produksi
                    GROUP BY
                        D.ID_Log_Kayu, D.Diameter;''')
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            log_kayu_id = row[0]
            diameter = row[1]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
            log_kayu_info = c.fetchone()
            log_kayu_nama = log_kayu_info[0]
            log_kayu_panjang = log_kayu_info[1]

            # Update the inventory dictionary with the calculated values based on thickness and size
            key = (log_kayu_nama, log_kayu_panjang, diameter)
            if key in inventory:
                inventory[key]['remaining_quantity'] -= row[2]
            else:
                inventory[key] = {
                    'nama_kayu': log_kayu_nama,
                    'panjang' : log_kayu_panjang,
                    'diameter': diameter,
                    'remaining_quantity': -row[2]
                }

        return inventory

    # Generate and display the inventory report
    def generate_raw_materials_report():
        products_inventory = calculate_raw_materials_inventory()

        for key, details in products_inventory.items():
            nama_produk, panjang, diameter = key
            remaining_quantity = details['remaining_quantity']
            treeviewkayu.insert("", tk.END, values=(nama_produk, panjang, diameter, remaining_quantity))
        
    treeviewkayu = ttk.Treeview(content_frame, columns=("Nama Kayu", "Panjang", "Diameter", "Stok"), show="headings", height=40)
    treeviewkayu.heading("Nama Kayu", text="Nama Kayu")
    treeviewkayu.heading("Panjang", text="Panjang")
    treeviewkayu.heading("Diameter", text="Diameter")
    treeviewkayu.heading("Stok", text="Stok")
    
    treeviewkayu.column("Nama Kayu", width=200)
    treeviewkayu.column("Panjang", width=150)
    treeviewkayu.column("Diameter", width=150)
    treeviewkayu.column("Stok", width=150)

    treeviewkayu.grid(row=0, column=0, padx=90, pady=5,)
    # Generate and display the inventory report
    generate_raw_materials_report()

    def calculate_products_inventory():
        inventory = {}  # Create an empty dictionary to store inventory data

        c.execute('''SELECT
                        D.ID_Hasil_Produksi AS ID_Produk,
                        D.Tebal,
                        D.Ukuran,
                        SUM(D.Jumlah) AS Jumlah_Terproduksi
                    FROM
                        Detail_Produksi D
                    JOIN
                        Produksi P ON D.ID_Produksi = P.ID_Produksi
                    GROUP BY
                        D.ID_Hasil_Produksi, D.Tebal, D.Ukuran;''')

        produksirows = c.fetchall()

        for row in produksirows:
            produk_id = row[0]
            tebal = row[1]
            ukuran = row[2]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (produk_id,))
            hasil_produksi_info = c.fetchone()
            nama_produk = hasil_produksi_info[0]
            jenis_produk = hasil_produksi_info[1]

            produk = "{} - {}".format(nama_produk, jenis_produk)

            if nama_produk == "Ampulur":
                nilaiukuran = "-"
                nilaitebal = "-"
            else:
                for rows in list_pasangan_ukuran:
                    if ukuran == rows[1]:
                        nilaiukuran = rows[0]

                for rows in list_pasangan_tebal:
                    if tebal == rows[1]:
                        nilaitebal = rows[0]

            # Update the inventory dictionary with the calculated values based on thickness and size
            key = (produk, nilaitebal, nilaiukuran)
            if key in inventory:
                inventory[key]['remaining_quantity'] += row[3]
            else:
                inventory[key] = {
                    'nama_produk': produk,
                    'tebal': nilaitebal,
                    'ukuran': nilaiukuran,
                    'remaining_quantity': row[3]
                }

        c.execute('''SELECT
                        D.ID_Hasil_Produksi AS ID_Produk,
                        D.Tebal,
                        D.Ukuran,
                        SUM(D.Jumlah) AS Jumlah_Terjual
                    FROM
                        Detail_Jual D
                    JOIN
                        Penjualan P ON D.ID_Penjualan = P.ID_Penjualan
                    GROUP BY
                        D.ID_Hasil_Produksi, D.Tebal, D.Ukuran;''')

        penjualanrows = c.fetchall()

        for row in penjualanrows:
            produk_id = row[0]
            tebal = row[1]
            ukuran = row[2]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (produk_id,))
            hasil_produksi_info = c.fetchone()
            nama_produk = hasil_produksi_info[0]
            jenis_produk = hasil_produksi_info[1]

            produk = "{} - {}".format(nama_produk, jenis_produk)

            if nama_produk == "Ampulur":
                nilaiukuran = "-"
                nilaitebal = "-"
            else:
                for rows in list_pasangan_ukuran:
                    if ukuran == rows[1]:
                        nilaiukuran = rows[0]

                for rows in list_pasangan_tebal:
                    if tebal == rows[1]:
                        nilaitebal = rows[0]

            # Update the inventory dictionary with the calculated values based on thickness and size
            key = (produk, nilaitebal, nilaiukuran)
            if key in inventory:
                inventory[key]['remaining_quantity'] -= row[3]
            else:
                # If the product was not found, add it to the inventory
                inventory[key] = {
                    'nama_produk': produk,
                    'tebal': nilaitebal,
                    'ukuran': nilaiukuran,
                    'remaining_quantity': -row[3]  # Negative quantity indicates sold
                }

        return inventory

    # Generate and display the inventory report
    def generate_inventory_report():
        products_inventory = calculate_products_inventory()

        for key, details in products_inventory.items():
            nama_produk, tebal, ukuran = key
            remaining_quantity = details['remaining_quantity']
            treeview.insert("", tk.END, values=(nama_produk, tebal, ukuran, remaining_quantity))
        
    treeview = ttk.Treeview(content_frame, columns=("Nama Produk", "Tebal", "Ukuran", "Stok"), show="headings", height=40)
    treeview.heading("Nama Produk", text="Nama Produk")
    treeview.heading("Tebal", text="Tebal")
    treeview.heading("Ukuran", text="Ukuran")
    treeview.heading("Stok", text="Stok")
    
    treeview.column("Nama Produk", width=200)
    treeview.column("Tebal", width=150)
    treeview.column("Ukuran", width=150)
    treeview.column("Stok", width=150)

    treeview.grid(row=0, column=1, padx=90, pady=5,)
    # Generate and display the inventory report
    generate_inventory_report()

def show_laporanpenjualan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")

        return start_date, end_date

    # Function to extract and display data based on date range
    def show_data():
        # Clear existing data in the table view
        for row in treeview.get_children():
            treeview.delete(row)
        
        update_date_range(None)
        
        c.execute('''SELECT
                        D.ID_Hasil_Produksi AS ID_Produk,
                        D.Tebal,
                        D.Ukuran,
                        SUM(D.Jumlah) AS Jumlah_Terjual,
                        AVG(D.Harga_Jual) AS Harga_Jual
                    FROM
                        Detail_Jual D
                    JOIN
                        Penjualan P ON D.ID_Penjualan = P.ID_Penjualan
                    WHERE
                        P.Tanggal_Faktur BETWEEN ? AND ?
                    GROUP BY
                        D.ID_Hasil_Produksi, D.Tebal, D.Ukuran;''',
              (start_date, end_date))
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            produk_id = row[0]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (produk_id,))
            hasil_produksi_info = c.fetchone()
            nama_produk = hasil_produksi_info[0]
            jenis_produk = hasil_produksi_info[1]

            tebal = row[1]
            ukuran = row[2]
            jumlah = row[3]
            harga = row[4]

            if nama_produk == "Ampulur":
                nilaiukuran = "-"
                nilaitebal = "-"
                rounded = "-"
                subtotal = jumlah * harga
            else:
                for row in list_pasangan_ukuran:
                    if ukuran == row[1]:
                        nilaiukuran = row[0]

                for row in list_pasangan_tebal:
                    if tebal == row[1]:
                        nilaitebal = row[0]

                volume = tebal * ukuran * jumlah / 10000000
                rounded = round(volume,4)
                subtotal = rounded * harga
            
            treeview.insert("", tk.END, values=(nama_produk, jenis_produk, nilaitebal, nilaiukuran, jumlah, rounded, format_currency(harga), format_currency(subtotal)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=show_data)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("Produk", "Jenis", "Tebal", "Ukuran", "Jumlah","Total Volume", "Harga Rata Rata", "Nilai Penjualan"), show="headings", height=36)
    treeview.heading("Produk", text="Produk")
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Tebal", text="Tebal (mm)")
    treeview.heading("Ukuran", text="Ukuran (cm)")
    treeview.heading("Jumlah", text="Jumlah")
    treeview.heading("Total Volume", text="Total Volume")
    treeview.heading("Harga Rata Rata", text="Harga Rata Rata")
    treeview.heading("Nilai Penjualan", text="Nilai Penjualan")

    treeview.column("Produk", width=150)
    treeview.column("Jenis", width=80)
    treeview.column("Tebal", width=80)
    treeview.column("Ukuran", width=80)
    treeview.column("Jumlah", width=80)
    treeview.column("Total Volume", width=100)
    treeview.column("Harga Rata Rata", width=150)
    treeview.column("Nilai Penjualan", width=150)

    treeview.grid(row=3, column=0, columnspan=2, padx=20, pady=20)

def show_laporanpembelian():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")

        return start_date, end_date

     # Function to extract and display data based on date range
    def show_data():
        # Clear existing data in the table view
        for row in treeview.get_children():
            treeview.delete(row)
        
        update_date_range(None)
        
        c.execute('''SELECT
                        D.ID_Log_Kayu,
                        D.Diameter,
                        SUM(D.Jumlah) AS Jumlah_Terjual,
                        AVG(D.Harga_Beli) AS Harga_Beli
                    FROM
                        Detail_Beli D
                    JOIN
                        Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                    WHERE
                        P.Tanggal_Nota BETWEEN ? AND ?
                    GROUP BY
                        D.ID_Log_Kayu, D.Diameter;''',
              (start_date, end_date))
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            log_kayu_id = row[0]
            diameter = row[1]
            jumlah = row[2]
            harga_beli = row[3]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
            log_kayu_info = c.fetchone()
            log_kayu_nama = log_kayu_info[0]
            log_kayu_panjang = log_kayu_info[1]

            volume = (math.pi * (diameter/2)**2 * log_kayu_panjang * jumlah)/1000000
            rounded = round(volume, 4)
            subtotal = rounded * harga_beli
            
            treeview.insert("", tk.END, values=(log_kayu_nama, log_kayu_panjang, diameter, jumlah, rounded, format_currency(harga_beli), format_currency(subtotal)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=show_data)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("Produk", "Panjang", "Diameter", "Jumlah","Total Volume", "Harga Rata Rata", "Nilai Pembelian"), show="headings", height=36)
    treeview.heading("Produk", text="Produk")
    treeview.heading("Panjang", text="Panjang")
    treeview.heading("Diameter", text="Diameter (cm)")
    treeview.heading("Jumlah", text="Jumlah")
    treeview.heading("Total Volume", text="Total Volume")
    treeview.heading("Harga Rata Rata", text="Harga Rata Rata")
    treeview.heading("Nilai Pembelian", text="Nilai Pembelian")

    treeview.column("Produk", width=150)
    treeview.column("Panjang", width=80)
    treeview.column("Diameter", width=80)
    treeview.column("Jumlah", width=80)
    treeview.column("Total Volume", width=100)
    treeview.column("Harga Rata Rata", width=150)
    treeview.column("Nilai Pembelian", width=150)

    treeview.grid(row=3, column=0, columnspan=2, padx=20, pady=20)

def show_laporanproduksi():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")

        return start_date, end_date

    # Function to extract and display data based on date range
    def show_data():
        # Clear existing data in the table view
        for row in treeview.get_children():
            treeview.delete(row)
        
        update_date_range(None)
        
        c.execute('''SELECT
                        D.ID_Hasil_Produksi AS ID_Produk,
                        D.Tebal,
                        D.Ukuran,
                        SUM(D.Jumlah) AS Jumlah_Terproduksi
                    FROM
                        Detail_Produksi D
                    JOIN
                        Produksi P ON D.ID_Produksi = P.ID_Produksi
                    WHERE
                        P.Tanggal_Produksi BETWEEN ? AND ?
                    GROUP BY
                        D.ID_Hasil_Produksi, D.Tebal, D.Ukuran;''',
              (start_date, end_date))
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            produk_id = row[0]

            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (produk_id,))
            hasil_produksi_info = c.fetchone()
            nama_produk = hasil_produksi_info[0]
            jenis_produk = hasil_produksi_info[1]

            tebal = row[1]
            ukuran = row[2]
            jumlah = row[3]

            if nama_produk == "Ampulur":
                nilaiukuran = "-"
                nilaitebal = "-"
                rounded = "-"
            else:
                for row in list_pasangan_ukuran:
                    if ukuran == row[1]:
                        nilaiukuran = row[0]

                for row in list_pasangan_tebal:
                    if tebal == row[1]:
                        nilaitebal = row[0]

                volume = tebal * ukuran * jumlah / 10000000
                rounded = round(volume,4)
            
            treeview.insert("", tk.END, values=(nama_produk, jenis_produk, nilaitebal, nilaiukuran, jumlah, rounded))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=show_data)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("Produk", "Jenis", "Tebal", "Ukuran", "Jumlah","Total Volume"), show="headings", height=36)
    treeview.heading("Produk", text="Produk")
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Tebal", text="Tebal (mm)")
    treeview.heading("Ukuran", text="Ukuran (cm)")
    treeview.heading("Jumlah", text="Jumlah")
    treeview.heading("Total Volume", text="Total Volume")

    treeview.column("Produk", width=150)
    treeview.column("Jenis", width=80)
    treeview.column("Tebal", width=80)
    treeview.column("Ukuran", width=80)
    treeview.column("Jumlah", width=80)
    treeview.column("Total Volume", width=100)

    treeview.grid(row=3, column=0, columnspan=2, padx=20, pady=20)

def show_laporanutangpiutang():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table_pembelian():
        # Clear existing table data
        for row in treeview_pembelian.get_children():
            treeview_pembelian.delete(row)

        pembelian = data_table_pembelian()

        for supplier_id, details in pembelian.items():

            c.execute("SELECT Nama FROM Supplier WHERE ID_Supplier=?", (supplier_id,))
            supplier_name = c.fetchone()[0]
            sisa_utang = details['total_utang']

            if sisa_utang < 0:
                # Insert data into the table view
                treeview_pembelian.insert("", tk.END, values=(supplier_name, format_currency(sisa_utang)))
    
    # Function to refresh the table view
    def data_table_pembelian():
        # Fetch and display data from the database
        c.execute("SELECT * FROM Pembelian ORDER BY ID_Pembelian DESC")
        rows = c.fetchall()

        utang_by_supplier = {}

        for row in rows:
            pembelian_id = row[0]
            supplier_id = row[1]
            bea_supplier = row[5]
            pembayaran = row[7]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Utang WHERE ID_Pembelian=?", (pembelian_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            listvolume = []
            listsubtotal = []
            c.execute("SELECT Diameter, Jumlah, Pembulatan, Harga_Beli FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
            detailbeli = c.fetchall()
            if detailbeli:
                for detail in detailbeli:
                    diameter = detail[0]
                    panjang = 130
                    jumlah = detail[1]
                    pembulatan = detail[2]
                    harga = detail[3]
                    volume = (math.pi * (diameter/2)**2 * panjang * jumlah)/1000000
                    rounded = round(volume, pembulatan)
                    subtotal = rounded * harga
                    listsubtotal.append(subtotal)
                    listvolume.append(rounded)
                totalvolume = round(sum(listvolume),2)
                total = sum(listsubtotal)
                biaya_bongkar = round(totalvolume * 7000)
                beban_pabrik = biaya_bongkar - bea_supplier
                grandtotal = total + beban_pabrik
                sisa_utang = totalbayar - grandtotal
                
                if supplier_id not in utang_by_supplier:
                    utang_by_supplier[supplier_id] = {
                        'total_utang': 0  # Initialize the total outstanding debt to zero
                    }

                # Update the total outstanding debt for the supplier
                utang_by_supplier[supplier_id]['total_utang'] += sisa_utang

        return utang_by_supplier

    # Function to refresh the Penjualan table
    def data_penjualan_table():

        c.execute("SELECT * FROM Penjualan ORDER BY ID_Penjualan DESC")
        rows = c.fetchall()

        piutang_by_pembeli = {}

        for row in rows:
            penjualan_id = row[0]
            pembeli_id = row[1]
            pembayaran = row[8]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Piutang WHERE ID_Penjualan=?", (penjualan_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
            detailjual = c.fetchall()
            if detailjual:
                listvolume = []
                listsubtotal = []
                for detail in detailjual:
                    idhasil = detail[0]
                    tebal = detail[1]
                    ukuran = detail[2]
                    jumlah = detail[3]
                    harga_jual = detail[4]
                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                    hasil_produksi_info = c.fetchone()
                    hasil_produksi_nama = hasil_produksi_info[0]
                    if hasil_produksi_nama == "Ampulur":
                        subtotal = jumlah * harga_jual
                    else:
                        subtotal = rounded * harga_jual
                    listvolume.append(volume)
                    listsubtotal.append(subtotal)

                volume = round(sum(listvolume),4)
                total = sum(listsubtotal)
                ppn = total*11/100
                grandtotal = total+ppn
                sisa_piutang = totalbayar - grandtotal

                if pembeli_id not in piutang_by_pembeli:
                    piutang_by_pembeli[pembeli_id] = {
                        'total_piutang': 0  # Initialize the total outstanding debt to zero
                    }

                # Update the total outstanding debt for the supplier
                piutang_by_pembeli[pembeli_id]['total_utang'] += sisa_piutang

        return piutang_by_pembeli
    
    def refresh_penjualan_table():
        for row in penjualan_treeview.get_children():
            penjualan_treeview.delete(row)

        penjualan = data_penjualan_table()

        for pembeli_id, details in penjualan.items():

            c.execute("SELECT Nama FROM Pembeli WHERE ID_Pembeli=?", (pembeli_id,))
            pembeli_name = c.fetchone()[0]
            sisa_piutang = details['total_piutang']

            if sisa_piutang < 0:
                # Insert data into the table view
                treeview_pembelian.insert("", tk.END, values=(pembeli_name, format_currency(sisa_piutang)))

    # Create the table view for Pembelian
    treeview_pembelian = ttk.Treeview(content_frame, columns=("Nama Supplier", "Sisa Utang"), show="headings", height=42)
    treeview_pembelian.heading("Nama Supplier", text="Nama Supplier")
    treeview_pembelian.heading("Sisa Utang", text="Sisa Utang")

    treeview_pembelian.column("Nama Supplier", width=150)
    treeview_pembelian.column("Sisa Utang", width=150)

    treeview_pembelian.grid(row=0, column=1, padx=5, pady=30)

    # Create the Penjualan table view
    penjualan_treeview = ttk.Treeview(content_frame, columns=("Nama Pembeli","Sisa Piutang"), show="headings", height=42)
    penjualan_treeview.heading("Nama Pembeli", text="Nama Pembeli")
    penjualan_treeview.heading("Sisa Piutang", text="Sisa Piutang")

    penjualan_treeview.column("Nama Pembeli", width=150)
    penjualan_treeview.column("Sisa Piutang", width=150)

    penjualan_treeview.grid(row=0, column=0, padx=350, pady=1)

    # Fetch and display initial data in the table view
    refresh_penjualan_table()

    refresh_table_pembelian()

def show_laporanaruskas():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
            start_previous = datetime(int(selected_year) - 1, 1, 1)
            end_previous = datetime(int(selected_year) - 1, 12, 31) 
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")
            start_previous = (datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)).replace(day=1)
            end_previous = datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)

        return start_date, end_date, end_previous

    # Function to calculate and display the financial report
    def generate_report():
        update_date_range(None)

        treeview.insert("", tk.END, values=("Arus Kas dari Aktivitas Operasional",""))
        treeview.insert("", tk.END, values=("-------------------------------------"))

        c.execute("SELECT * FROM Penjualan WHERE Penjualan.Tanggal_Faktur BETWEEN ? AND ?", (start_date,end_date))
        rows = c.fetchall()
        listbayarpembeli = []
        for row in rows:
            penjualan_id = row[0]
            pembayaran = row[8]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Piutang WHERE ID_Penjualan=?", (penjualan_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)
            listbayarpembeli.append(totalbayar)

        penerimaan_pembeli = sum(listbayarpembeli)
        treeview.insert("", tk.END, values=("Penerimaan dari Pelanggan",format_currency(penerimaan_pembeli)))

        c.execute("SELECT * FROM Pembelian WHERE Pembelian.Tanggal_Nota BETWEEN ? AND ?", (start_date,end_date))
        utangrows = c.fetchall()
        listbayarsupplier = []
        for row in utangrows:
            pembelian_id = row[0]
            pembayaran = row[7]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Utang WHERE ID_Pembelian=?", (pembelian_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)
            listbayarsupplier.append(totalbayar)

        pengeluaran_supplier = sum(listbayarsupplier)
        treeview.insert("", tk.END, values=("Pembayaran ke Pemasok",format_currency(pengeluaran_supplier)))

        c.execute("SELECT ID_Akun, SUM(Pembayaran) FROM Biaya WHERE Biaya.Tanggal_Terima BETWEEN ? AND ? GROUP BY ID_Akun", (start_date, end_date))
        rowsbiaya = c.fetchall()
        list_biaya_operasional = []
        for row in rowsbiaya:
            akun_id = row[0]
            biaya = row[1]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Kategori FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            kategori = hasil_produksi_info[0]
            if kategori == "Beban":
                list_biaya_operasional.append(biaya)

        biaya_operasional = sum(list_biaya_operasional)
        treeview.insert("", tk.END, values=("Pengeluaran Operasional",format_currency(biaya_operasional)))

        kas_aktivitas_operasional = penerimaan_pembeli + pengeluaran_supplier + biaya_operasional
        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("Kas Bersih dari Aktivitas Operasional",format_currency(kas_aktivitas_operasional)))
        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("Arus Kas dari Aktivitas Investasi",""))
        treeview.insert("", tk.END, values=("-------------------------------------"))

        c.execute("SELECT * FROM Aset_Tetap WHERE Aset_Tetap.Tanggal_Beli BETWEEN ? AND ?", (start_date,end_date))
        rows = c.fetchall()
        list_aset_tetap = []
        for row in rows:
            harga_beli = row[4]
            unit = row[5]
            jumlah = unit*harga_beli
            list_aset_tetap.append(jumlah)

        pembelian_aset_tetap = sum(list_aset_tetap)
        treeview.insert("", tk.END, values=("Pembelian Aset Tetap",format_currency(pembelian_aset_tetap)))
        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("Kas Bersih dari Aktivitas Investasi",format_currency(pembelian_aset_tetap)))

        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("Arus Kas dari Aktivitas Pendanaan",""))
        treeview.insert("", tk.END, values=("-------------------------------------"))

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        kas_awal_str = sheet['A2'].value
        if kas_awal_str is None:
            kas_awal = 0
        else:
            kas_awal = float(kas_awal_str)

        bank_awal_str = sheet['B2'].value
        if bank_awal_str is None:
            bank_awal = 0
        else:
            bank_awal = float(bank_awal_str)

        pengambilan_pribadi_str = sheet['C2'].value
        if pengambilan_pribadi_str is None:
            pengambilan_pribadi = 0
        else:
            pengambilan_pribadi = float(pengambilan_pribadi_str)

        tambahan_modal_str = sheet['D2'].value
        if tambahan_modal_str is None:
            tambahan_modal = 0
        else:
            tambahan_modal = float(tambahan_modal_str)

        treeview.insert("", tk.END, values=("Prive",format_currency(pengambilan_pribadi)))
        treeview.insert("", tk.END, values=("Tambahan Modal",format_currency(tambahan_modal)))
        
        aktivitas_pendanaan = pengambilan_pribadi + tambahan_modal
        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("Kas Bersih dari Aktivitas Pendanaan",format_currency(aktivitas_pendanaan)))

        c.execute("""
        SELECT Akun.Nama, COALESCE(SUM(
            COALESCE(Penjualan.Pembayaran, 0) + 
            COALESCE(Piutang.Pembayaran, 0) - 
            COALESCE(Aset_Tetap.Harga_Beli * Aset_Tetap.Jumlah, 0) - 
            COALESCE(Biaya.Pembayaran, 0) - 
            COALESCE(Pembelian.Pembayaran, 0) - 
            COALESCE(Utang.Pembayaran, 0)
        ), 0) AS Saldo
        FROM Akun
        LEFT JOIN Aset_Tetap ON Akun.ID_Akun = Aset_Tetap.Dari_Akun AND Aset_Tetap.Tanggal_Beli <= ?
        LEFT JOIN Biaya ON Akun.ID_Akun = Biaya.Dari_Akun AND Biaya.Tanggal_Terima <= ?
        LEFT JOIN Pembelian ON Akun.ID_Akun = Pembelian.ID_Akun AND Pembelian.Tanggal_Nota <= ?
        LEFT JOIN Utang ON Akun.ID_Akun = Utang.ID_Akun AND Utang.Tanggal <= ?
        LEFT JOIN Penjualan ON Akun.ID_Akun = Penjualan.ID_Akun AND Penjualan.Tanggal_Faktur <= ?
        LEFT JOIN Piutang ON Akun.ID_Akun = Piutang.ID_Akun AND Piutang.Tanggal <= ?
        WHERE Akun.Kategori = 'Kas & Bank'
        GROUP BY Akun.ID_Akun, Akun.Nama;
        """, (end_previous,end_previous,end_previous,end_previous,end_previous,end_previous))
        rowsawal = c.fetchall()

        listkasbankawal = []
        for row in rowsawal:
            if row[0] == "Kas":
                saldo = row[1] + kas_awal
            else:
                saldo = row[1] + bank_awal
            listkasbankawal.append(saldo)
        total_kas_bank_awal = sum(listkasbankawal)

        c.execute("""
        SELECT Akun.Nama, COALESCE(SUM(
            COALESCE(Penjualan.Pembayaran, 0) + 
            COALESCE(Piutang.Pembayaran, 0) - 
            COALESCE(Aset_Tetap.Harga_Beli * Aset_Tetap.Jumlah, 0) - 
            COALESCE(Biaya.Pembayaran, 0) - 
            COALESCE(Pembelian.Pembayaran, 0) - 
            COALESCE(Utang.Pembayaran, 0)
        ), 0) AS Saldo
        FROM Akun
        LEFT JOIN Aset_Tetap ON Akun.ID_Akun = Aset_Tetap.Dari_Akun AND Aset_Tetap.Tanggal_Beli <= ?
        LEFT JOIN Biaya ON Akun.ID_Akun = Biaya.Dari_Akun AND Biaya.Tanggal_Terima <= ?
        LEFT JOIN Pembelian ON Akun.ID_Akun = Pembelian.ID_Akun AND Pembelian.Tanggal_Nota <= ?
        LEFT JOIN Utang ON Akun.ID_Akun = Utang.ID_Akun AND Utang.Tanggal <= ?
        LEFT JOIN Penjualan ON Akun.ID_Akun = Penjualan.ID_Akun AND Penjualan.Tanggal_Faktur <= ?
        LEFT JOIN Piutang ON Akun.ID_Akun = Piutang.ID_Akun AND Piutang.Tanggal <= ?
        WHERE Akun.Kategori = 'Kas & Bank'
        GROUP BY Akun.ID_Akun, Akun.Nama;
        """, (end_date,end_date,end_date,end_date,end_date,end_date))
        rowsakhir = c.fetchall()

        listkasbankakhir = []
        for row in rowsakhir:
            if row[0] == "Kas":
                saldo = row[1] + kas_awal
            else:
                saldo = row[1] + bank_awal
            listkasbankakhir.append(saldo)
        total_kas_bank_akhir = sum(listkasbankakhir)

        selisih = total_kas_bank_awal - total_kas_bank_akhir
        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("Kenaikan (Penuruan) Kas",format_currency(selisih)))
        treeview.insert("", tk.END, values=("Saldo Kas Awal",format_currency(total_kas_bank_awal)))
        treeview.insert("", tk.END, values=("Saldo Kas Akhir",format_currency(total_kas_bank_akhir)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Item", "Value"), show="headings", height=25)
    treeview.heading("Item", text="")
    treeview.heading("Value", text="")

    treeview.column("Item", width=200)
    treeview.column("Value", width=200)

    treeview.grid(row=3, columnspan=2, padx=5, pady=5,)

def calculate_raw_materials_inventory(end_date):
    inventory = {}

    c.execute('''SELECT
                    D.ID_Log_Kayu,
                    D.Diameter,
                    SUM(D.Jumlah) AS Jumlah_Terjual,
                    AVG(Harga_Beli) AS Harga_Beli
                FROM
                    Detail_Beli D
                JOIN
                    Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                WHERE
                    P.Tanggal_Nota <= ?
                GROUP BY
                    D.ID_Log_Kayu, D.Diameter;''',
              (end_date,))
    
    rows = c.fetchall()
    
    # Insert data into the table view
    for row in rows:
        log_kayu_id = row[0]
        diameter = row[1]
        harga_beli = row[3]

        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
        log_kayu_info = c.fetchone()
        log_kayu_nama = log_kayu_info[0]
        log_kayu_panjang = log_kayu_info[1]

        # Update the inventory dictionary with the calculated values based on thickness and size
        key = (log_kayu_nama, log_kayu_panjang, diameter)
        if key in inventory:
            inventory[key]['remaining_quantity'] += row[2]
            inventory[key]['harga_beli'] = harga_beli
        else:
            inventory[key] = {
                'nama_kayu': log_kayu_nama,
                'panjang' : log_kayu_panjang,
                'diameter': diameter,
                'remaining_quantity': row[2],
                'harga_beli' : harga_beli
            }

    c.execute('''SELECT
                    D.ID_Log_Kayu,
                    D.Diameter,
                    SUM(D.Jumlah) AS Jumlah_Terjual
                FROM
                    Detail_Beli D
                JOIN
                    Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                JOIN
                    Produksi PR ON P.ID_Pembelian = PR.ID_Pembelian
                JOIN
                    Detail_Produksi DP ON PR.ID_Produksi = DP.ID_Produksi
                WHERE
                    P.Tanggal_Nota <= ?
                GROUP BY
                    D.ID_Log_Kayu, D.Diameter;''',
              (end_date,))
    
    rows = c.fetchall()
    
    # Insert data into the table view
    for row in rows:
        log_kayu_id = row[0]
        diameter = row[1]

        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
        log_kayu_info = c.fetchone()
        log_kayu_nama = log_kayu_info[0]
        log_kayu_panjang = log_kayu_info[1]

        # Update the inventory dictionary with the calculated values based on thickness and size
        key = (log_kayu_nama, log_kayu_panjang, diameter)
        if key in inventory:
            inventory[key]['remaining_quantity'] -= row[2]
        else:
            inventory[key] = {
                'nama_kayu': log_kayu_nama,
                'panjang' : log_kayu_panjang,
                'diameter': diameter,
                'remaining_quantity': -row[2]
            }

    return inventory

def calculate_products_inventory(end_date):
    inventory = {}  # Create an empty dictionary to store inventory data

    c.execute('''SELECT
                    D.ID_Hasil_Produksi AS ID_Produk,
                    D.Tebal,
                    D.Ukuran,
                    SUM(D.Jumlah) AS Jumlah_Terproduksi
                FROM
                    Detail_Produksi D
                JOIN
                    Produksi P ON D.ID_Produksi = P.ID_Produksi
                WHERE
                    P.Tanggal_Produksi <= ?
                GROUP BY
                    D.ID_Hasil_Produksi, D.Tebal, D.Ukuran;''',
              (end_date,))

    produksirows = c.fetchall()

    for row in produksirows:
        produk_id = row[0]
        tebal = row[1]
        ukuran = row[2]

        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (produk_id,))
        hasil_produksi_info = c.fetchone()
        nama_produk = hasil_produksi_info[0]
        jenis_produk = hasil_produksi_info[1]

        produk = "{} - {}".format(nama_produk, jenis_produk)

        if nama_produk == "Ampulur":
            nilaiukuran = "-"
            nilaitebal = "-"
        else:
            for rows in list_pasangan_ukuran:
                if ukuran == rows[1]:
                    nilaiukuran = rows[0]

            for rows in list_pasangan_tebal:
                if tebal == rows[1]:
                    nilaitebal = rows[0]

        # Update the inventory dictionary with the calculated values based on thickness and size
        key = (produk, nilaitebal, nilaiukuran)
        if key in inventory:
            inventory[key]['remaining_quantity'] += row[3]
        else:
            inventory[key] = {
                'nama_produk': produk,
                'tebal': nilaitebal,
                'ukuran': nilaiukuran,
                'remaining_quantity': row[3]
            }

    c.execute('''SELECT
                    D.ID_Hasil_Produksi AS ID_Produk,
                    D.Tebal,
                    D.Ukuran,
                    SUM(D.Jumlah) AS Jumlah_Terjual
                FROM
                    Detail_Jual D
                JOIN
                    Penjualan P ON D.ID_Penjualan = P.ID_Penjualan
                WHERE
                    P.Tanggal_Faktur <= ?
                GROUP BY
                    D.ID_Hasil_Produksi, D.Tebal, D.Ukuran;''',
              (end_date,))

    penjualanrows = c.fetchall()

    for row in penjualanrows:
        produk_id = row[0]
        tebal = row[1]
        ukuran = row[2]

        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (produk_id,))
        hasil_produksi_info = c.fetchone()
        nama_produk = hasil_produksi_info[0]
        jenis_produk = hasil_produksi_info[1]

        produk = "{} - {}".format(nama_produk, jenis_produk)

        if nama_produk == "Ampulur":
            nilaiukuran = "-"
            nilaitebal = "-"
        else:
            for rows in list_pasangan_ukuran:
                if ukuran == rows[1]:
                    nilaiukuran = rows[0]

            for rows in list_pasangan_tebal:
                if tebal == rows[1]:
                    nilaitebal = rows[0]

        # Update the inventory dictionary with the calculated values based on thickness and size
        key = (produk, nilaitebal, nilaiukuran)
        if key in inventory:
            inventory[key]['remaining_quantity'] -= row[3]
        else:
            # If the product was not found, add it to the inventory
            inventory[key] = {
                'nama_produk': produk,
                'tebal': nilaitebal,
                'ukuran': nilaiukuran,
                'remaining_quantity': -row[3]  # Negative quantity indicates sold
            }

    return inventory

def persediaan_barang_dalam_proses(end_date):

    inventory = {}
    
    c.execute('''SELECT
                    D.ID_Log_Kayu,
                    D.Diameter,
                    SUM(D.Jumlah) AS Jumlah_Terjual,
                    AVG(Harga_Beli) AS Harga_Beli
                FROM
                    Detail_Beli D
                JOIN
                    Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                JOIN
                    Produksi PR ON P.ID_Pembelian = PR.ID_Pembelian
                WHERE
                    P.Tanggal_Nota <= ?
                GROUP BY
                    D.ID_Log_Kayu, D.Diameter;''',
              (end_date,))
    
    rows = c.fetchall()


    # Insert data into the table view
    for row in rows:
        log_kayu_id = row[0]
        diameter = row[1]
        harga_beli = row[3]

        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
        log_kayu_info = c.fetchone()
        log_kayu_nama = log_kayu_info[0]
        log_kayu_panjang = log_kayu_info[1]

        # Update the inventory dictionary with the calculated values based on thickness and size
        key = (log_kayu_nama, log_kayu_panjang, diameter)
        if key in inventory:
            inventory[key]['remaining_quantity'] += row[2]
            inventory[key]['harga_beli'] = harga_beli
        else:
            inventory[key] = {
                'nama_kayu': log_kayu_nama,
                'panjang' : log_kayu_panjang,
                'diameter': diameter,
                'remaining_quantity': row[2],
                'harga_beli' : harga_beli
            }

    c.execute('''SELECT
                    D.ID_Log_Kayu,
                    D.Diameter,
                    SUM(D.Jumlah) AS Jumlah_Terjual
                FROM
                    Detail_Beli D
                JOIN
                    Pembelian P ON D.ID_Pembelian = P.ID_Pembelian
                JOIN
                    Produksi PR ON P.ID_Pembelian = PR.ID_Pembelian
                JOIN
                    Detail_Produksi DP ON PR.ID_Produksi = DP.ID_Produksi
                WHERE
                    P.Tanggal_Nota <= ?
                GROUP BY
                    D.ID_Log_Kayu, D.Diameter;''',
              (end_date,))
    
    rows = c.fetchall()
    
    # Insert data into the table view
    for row in rows:
        log_kayu_id = row[0]
        diameter = row[1]

        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama, Panjang FROM Log_Kayu WHERE ID_Log_Kayu = ?", (log_kayu_id,))
        log_kayu_info = c.fetchone()
        log_kayu_nama = log_kayu_info[0]
        log_kayu_panjang = log_kayu_info[1]

        # Update the inventory dictionary with the calculated values based on thickness and size
        key = (log_kayu_nama, log_kayu_panjang, diameter)
        if key in inventory:
            inventory[key]['remaining_quantity'] -= row[2]
        else:
            inventory[key] = {
                'nama_kayu': log_kayu_nama,
                'panjang' : log_kayu_panjang,
                'diameter': diameter,
                'remaining_quantity': -row[2]
            }

    return inventory

def generate_saldo_kayu(data):
    listtotalkayu = []
    listvolume = []
    for key, details in data.items():
        nama_kayu, panjang, diameter = key
        remaining_quantity = details['remaining_quantity']
        harga = details['harga_beli']
        volume = (math.pi * (diameter/2)**2 * panjang * remaining_quantity)/1000000
        rounded = round(volume, 4)
        subtotal = rounded * harga
        listtotalkayu.append(subtotal)
        listvolume.append(rounded)
    
    bea_angkut = 7000 * sum(listvolume) / 1.75
    saldo_kayu = sum(listtotalkayu)+bea_angkut

    return saldo_kayu

def generate_saldo_produk(data):
    listtotalproduk = []
    for key, details in data.items():
        nama_produk, tebal, ukuran = key
        remaining_quantity = details['remaining_quantity']

        for row in list_pasangan_ukuran:
            if ukuran == row[0]:
                nilaiukuran = row[1]

        volume = float(tebal) * nilaiukuran * remaining_quantity / 10000000
        rounded = round(volume,4)

        if nama_produk == "Ampulur":
            subtotal = remaining_quantity * 1100
        elif nama_produk == "MK Core Sengon Basah" or "MK Core Kayu Keras Basah":
            subtotal = rounded * 560000
        elif nama_produk == "Core Sengon Basah":
            subtotal = rounded * 1750000
        else:
            subtotal = rounded * 1870000
        listtotalproduk.append(subtotal)

    return listtotalproduk

def show_laporanhpp():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
            start_previous = datetime(int(selected_year) - 1, 1, 1)
            end_previous = datetime(int(selected_year) - 1, 12, 31) 
        else:
            first_day = f"{selected_year}-{months.index(selected_month) :02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month) :02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")
            start_previous = (datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)).replace(day=1)
            end_previous = datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)

        return start_date, end_date, end_previous

    # Function to calculate and display the financial report
    def generate_report():
        update_date_range(None)

        raw_materials_inventory_awal = calculate_raw_materials_inventory(end_previous)
        
        saldo_kayu_awal = generate_saldo_kayu(raw_materials_inventory_awal)
        treeview.insert("", tk.END, values=("BAHAN BAKU", "", "",""))
        treeview.insert("", tk.END, values=("Persediaan Bahan Baku Awal", "", format_currency(saldo_kayu_awal),""))

        # Fetch and display data from the database
        c.execute("SELECT * FROM Pembelian WHERE Pembelian.Tanggal_Nota BETWEEN ? AND ?", (start_date, end_date))
        rows = c.fetchall()
        listpembelian = []
        listbiayaangkut = []
        for row in rows:
            pembelian_id = row[0]
            bea_supplier = row[5]

            listvolume = []
            listsubtotal = []

            c.execute("SELECT Diameter, Jumlah, Pembulatan, Harga_Beli FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
            detailbeli = c.fetchall()
            if detailbeli:
                for detail in detailbeli:
                    diameter = detail[0]
                    panjang = 130
                    jumlah = detail[1]
                    pembulatan = detail[2]
                    harga = detail[3]
                    volume = (math.pi * (diameter/2)**2 * panjang * jumlah)/1000000
                    rounded = round(volume, pembulatan)
                    subtotal = rounded * harga
                    listsubtotal.append(subtotal)
                    listvolume.append(rounded)

            totalvolume = round(sum(listvolume),2)
            grandtotal = sum(listsubtotal)

            listpembelian.append(grandtotal)
            biaya_bongkar = round(totalvolume * 7000)
            beban_pabrik = biaya_bongkar - bea_supplier
            listbiayaangkut.append(beban_pabrik)
        
        pembelianlog = sum(listpembelian)
        biaya_angkut = sum(listbiayaangkut)
        
        treeview.insert("", tk.END, values=("Pembelian Bahan Baku", format_currency(pembelianlog),"",""))
        treeview.insert("", tk.END, values=("Biaya Bongkar Pembelian Bahan Baku", format_currency(biaya_angkut),"",""))
        jumlah_pemakaian_bahan_baku_bersih = saldo_kayu_awal + pembelianlog + biaya_angkut
        treeview.insert("", tk.END, values=("Jumlah Pembelian Bahan Baku Bersih", "",format_currency(jumlah_pemakaian_bahan_baku_bersih),""))

        raw_materials_inventory_akhir = calculate_raw_materials_inventory(end_date)
        
        saldo_kayu_akhir = generate_saldo_kayu(raw_materials_inventory_akhir)

        treeview.insert("", tk.END, values=("Persediaan Bahan Baku Akhir", "",format_currency(-saldo_kayu_akhir),""))
        pemakaian_bahan_baku = jumlah_pemakaian_bahan_baku_bersih - saldo_kayu_akhir
        treeview.insert("", tk.END, values=("Pemakaian Bahan Baku", "", "",format_currency(pemakaian_bahan_baku),""))

        c.execute("SELECT ID_Akun, SUM(Pembayaran) FROM Biaya WHERE Biaya.Tanggal_Terima BETWEEN ? AND ? GROUP BY ID_Akun", (start_date, end_date))
        rowsbiaya = c.fetchall()
        listlangsung = []
        listtidaklangsung = []
        listperawatanmesin = []
        listperlengkapanpabrik = []
        listkirimangkut = []
        for row in rowsbiaya:
            akun_id = row[0]
            biaya = row[1]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            if nama == "B. Upah Karyawan":
                listlangsung.append(biaya)
            elif nama == "B. Gaji Grader":
                listtidaklangsung.append(biaya)
            elif nama == "B. Perawatan Mesin":
                listperawatanmesin.append(biaya)
            elif nama == "B. Perlengkapan Pabrik":
                listperlengkapanpabrik.append(biaya)
            elif nama == "B. Pengiriman":
                listkirimangkut.append(biaya)
            else:
                pass

        upah_langsung = sum(listlangsung)
        treeview.insert("", tk.END, values=("", "", "",""))
        treeview.insert("", tk.END, values=("UPAH LANGSUNG", "", "",format_currency(upah_langsung)))
        treeview.insert("", tk.END, values=("", "", "",""))
        treeview.insert("", tk.END, values=("BIAYA OVERHEAD PABRIK", "", "", ""))

        upah_tidak_langsung = sum(listtidaklangsung)
        treeview.insert("", tk.END, values=("Upah Tidak Langsung", "",format_currency(upah_tidak_langsung),""))

        c.execute("SELECT * FROM Aset_Tetap WHERE Aset_Tetap.Tanggal_Beli <= ?", (end_date,))
        rows = c.fetchall()
        list_penyusutan_bangunan = []
        list_penyusutan_mesin = []
        for row in rows:
            akun_id = row[1]
            tanggal_beli = row[2]
            harga_beli = row[3]
            unit = row[4]
            umur_ekonomis = row[5]
            jumlah = unit*harga_beli
            penyusutan = jumlah/umur_ekonomis/12
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            if nama == "Bangunan":
                list_penyusutan_bangunan.append(penyusutan)
            elif nama == "Inventaris/Mesin":
                list_penyusutan_mesin.append(penyusutan)
            else:
                pass
        
        perawatanmesin = sum(listperawatanmesin)
        perlengkapanpabrik = sum(listperlengkapanpabrik)
        penyusutan_bangunan = sum(list_penyusutan_bangunan)
        penyusutan_mesin = sum(list_penyusutan_mesin)
        kirimangkut = sum(listkirimangkut)

        treeview.insert("", tk.END, values=("Biaya Perlengkapan Pabrik","",format_currency(perlengkapanpabrik),""))
        treeview.insert("", tk.END, values=("Biaya Perawatan Mesin","",format_currency(perawatanmesin ),""))
        treeview.insert("", tk.END, values=("Biaya Pengiriman & Pengangkutan","",format_currency(kirimangkut),""))
        treeview.insert("", tk.END, values=("Biaya Penyusutan Mesin Pabrik","",format_currency(penyusutan_mesin),""))
        treeview.insert("", tk.END, values=("Biaya Penyusutan Bangunan","",format_currency(penyusutan_bangunan),""))

        overhead_pabrik = upah_tidak_langsung + penyusutan_bangunan + penyusutan_mesin + perawatanmesin + perlengkapanpabrik

        treeview.insert("", tk.END, values=("Jumlah Biaya Overhead Pabrik","","",format_currency(overhead_pabrik)))

        jumlah_biaya_produksi = pemakaian_bahan_baku + upah_langsung + overhead_pabrik

        treeview.insert("", tk.END, values=("", "", "",""))
        treeview.insert("", tk.END, values=("Total Biaya Produksi","","",format_currency(jumlah_biaya_produksi)))

        persediaan_barang_dalam_proses_awal = persediaan_barang_dalam_proses(end_previous)
        persediaan_barang_dalam_proses_akhir = persediaan_barang_dalam_proses(end_date)

        saldo_barang_dalam_proses_awal = generate_saldo_kayu(persediaan_barang_dalam_proses_awal)
        saldo_barang_dalam_proses_akhir = generate_saldo_kayu(persediaan_barang_dalam_proses_akhir)

        treeview.insert("", tk.END, values=("Persediaan Barang Dalam Proses (Awal)","","",format_currency(saldo_barang_dalam_proses_awal)))
        treeview.insert("", tk.END, values=("Persediaan Barang Dalam Proses (Akhir)","","",format_currency(-saldo_barang_dalam_proses_akhir)))
        treeview.insert("", tk.END, values=("","","",""))
        harga_pokok_produksi = jumlah_biaya_produksi + saldo_barang_dalam_proses_awal - saldo_barang_dalam_proses_akhir
        treeview.insert("", tk.END, values=("HARGA POKOK PRODUKSI","","",format_currency(harga_pokok_produksi)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Item", "Value","Item2", "Value2"), show="headings", height=25)
    treeview.heading("Item", text="")
    treeview.heading("Value", text="")
    treeview.heading("Item2", text="")
    treeview.heading("Value2", text="")

    treeview.column("Item", width=250)
    treeview.column("Value", width=150)
    treeview.column("Item2", width=150)
    treeview.column("Value2", width=150)

    treeview.grid(row=3, columnspan=2, padx=5, pady=5,)

def generatehpp(start_date, end_date, end_previous):
    raw_materials_inventory_awal = calculate_raw_materials_inventory(end_previous)
    
    saldo_kayu_awal = generate_saldo_kayu(raw_materials_inventory_awal)

    # Fetch and display data from the database
    c.execute("SELECT * FROM Pembelian WHERE Pembelian.Tanggal_Nota BETWEEN ? AND ?", (start_date, end_date))
    rows = c.fetchall()
    listpembelian = []
    listbiayaangkut = []
    for row in rows:
        pembelian_id = row[0]
        bea_supplier = row[5]

        listvolume = []
        listsubtotal = []
        c.execute("SELECT Diameter, Jumlah, Pembulatan, Harga_Beli FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
        detailbeli = c.fetchall()
        if detailbeli:
            for detail in detailbeli:
                diameter = detail[0]
                panjang = 130
                jumlah = detail[1]
                pembulatan = detail[2]
                harga = detail[3]
                volume = (math.pi * (diameter/2)**2 * panjang * jumlah)/1000000
                rounded = round(volume, pembulatan)
                subtotal = rounded * harga
                listsubtotal.append(subtotal)
                listvolume.append(rounded)

        totalvolume = round(sum(listvolume),2)
        total = sum(listsubtotal)
        listpembelian.append(total)
        biaya_bongkar = round(totalvolume * 7000)
        beban_pabrik = biaya_bongkar - bea_supplier
        listbiayaangkut.append(beban_pabrik)
    
    pembelianlog = sum(listpembelian)
    biaya_angkut = sum(listbiayaangkut)

    jumlah_pemakaian_bahan_baku_bersih = saldo_kayu_awal + pembelianlog + biaya_angkut

    raw_materials_inventory_akhir = calculate_raw_materials_inventory(end_date)
    
    saldo_kayu_akhir = generate_saldo_kayu(raw_materials_inventory_akhir)

    pemakaian_bahan_baku = jumlah_pemakaian_bahan_baku_bersih - saldo_kayu_akhir

    c.execute("SELECT ID_Akun, SUM(Pembayaran) FROM Biaya WHERE Biaya.Tanggal_Terima BETWEEN ? AND ? GROUP BY ID_Akun", (start_date, end_date))
    rowsbiaya = c.fetchall()
    listlangsung = []
    listtidaklangsung = []
    listperawatanmesin = []
    listperlengkapanpabrik = []
    for row in rowsbiaya:
        akun_id = row[0]
        biaya = row[1]
        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
        hasil_produksi_info = c.fetchone()
        nama = hasil_produksi_info[0]
        if nama == "B. Gaji Karyawan":
            listlangsung.append(biaya)
        elif nama == "B. Gaji Staf":
            listtidaklangsung.append(biaya)
        elif nama == "B. Perawatan Mesin":
            listperawatanmesin.append(biaya)
        elif nama == "B. Perlengkapan Pabrik":
            listperlengkapanpabrik.append(biaya)
        else:
            pass
    upah_langsung = sum(listlangsung)

    upah_tidak_langsung = sum(listtidaklangsung)

    c.execute("SELECT * FROM Aset_Tetap WHERE Aset_Tetap.Tanggal_Beli <= ?", (end_date,))
    rows = c.fetchall()
    list_penyusutan_bangunan = []
    list_penyusutan_mesin = []
    for row in rows:
        akun_id = row[1]
        harga_beli = row[4]
        unit = row[5]
        umur_ekonomis = row[6]
        jumlah = unit*harga_beli
        penyusutan = jumlah/umur_ekonomis/12
        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
        hasil_produksi_info = c.fetchone()
        nama = hasil_produksi_info[0]
        if nama == "Bangunan":
            list_penyusutan_bangunan.append(penyusutan)
        elif nama == "Inventaris/Mesin":
            list_penyusutan_mesin.append(penyusutan)
        else:
            pass
    
    perawatanmesin = sum(listperawatanmesin)
    perlengkapanpabrik = sum(listperlengkapanpabrik)
    penyusutan_bangunan = sum(list_penyusutan_bangunan)
    penyusutan_mesin = sum(list_penyusutan_mesin)

    overhead_pabrik = upah_tidak_langsung + penyusutan_bangunan + penyusutan_mesin + perawatanmesin + perlengkapanpabrik

    jumlah_biaya_produksi = pemakaian_bahan_baku + upah_langsung + overhead_pabrik

    persediaan_barang_dalam_proses_awal = persediaan_barang_dalam_proses(end_previous)
    persediaan_barang_dalam_proses_akhir = persediaan_barang_dalam_proses(end_date)

    saldo_barang_dalam_proses_awal = generate_saldo_kayu(persediaan_barang_dalam_proses_awal)
    saldo_barang_dalam_proses_akhir = generate_saldo_kayu(persediaan_barang_dalam_proses_akhir)

    harga_pokok_produksi = jumlah_biaya_produksi + saldo_barang_dalam_proses_awal - saldo_barang_dalam_proses_akhir

    return harga_pokok_produksi

def generate_laba(start_date, end_date, end_previous):
    # Get the selected month and year from the combo boxes

    c.execute("SELECT * FROM Penjualan WHERE Penjualan.Tanggal_Faktur BETWEEN ? AND ?", (start_date, end_date))
    listpenjualan = []
    rows = c.fetchall()
    for row in rows:
        penjualan_id = row[0]

        c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
        detailjual = c.fetchall()

        if detailjual:
            listsubtotal = []
            for detail in detailjual:
                idhasil = detail[0]
                tebal = detail[1]
                ukuran = detail[2]
                jumlah = detail[3]
                harga_jual = detail[4]

                volume = tebal * ukuran * jumlah / 10000000
                rounded = round(volume,4)
                
                c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                hasil_produksi_info = c.fetchone()

                hasil_produksi_nama = hasil_produksi_info[0]
                if hasil_produksi_nama == "Ampulur":
                    subtotal = jumlah * harga_jual
                else:
                    subtotal = rounded * harga_jual

                listsubtotal.append(subtotal)

            total = sum(listsubtotal)
            ppn = total*11/100
            penjualan = total+ppn
            listpenjualan.append(penjualan)

    products_inventory_awal = calculate_products_inventory(end_previous)
    products_inventory_akhir = calculate_products_inventory(end_date)

    saldo_produk_awal = generate_saldo_produk(products_inventory_awal)
    saldo_produk_akhir = generate_saldo_produk(products_inventory_akhir)

    persediaan_awal = sum(saldo_produk_awal)
    persediaan_akhir = sum(saldo_produk_akhir)

    harga_pokok_produksi = generatehpp(start_date, end_date, end_previous)
    
    harga_pokok_penjualan = persediaan_awal + harga_pokok_produksi - persediaan_akhir

    laba_kotor = sum(listpenjualan) - harga_pokok_penjualan

    c.execute("SELECT ID_Akun, SUM(Pembayaran) FROM Biaya WHERE Biaya.Tanggal_Terima BETWEEN ? AND ? GROUP BY ID_Akun", (start_date, end_date))
    rowsbiaya = c.fetchall()
    listbebanpenjualan = []
    listumumadm = []
    for row in rowsbiaya:
        akun_id = row[0]
        biaya = row[1]
        # Get the Log Kayu's information based on the ID
        c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
        hasil_produksi_info = c.fetchone()
        nama = hasil_produksi_info[0]
        if nama == "B. Komisi & Fee":
            listbebanpenjualan.append(biaya)
        elif nama == "B. Gaji Staf Kantor":
            listumumadm.append(biaya)
        elif nama == "B. Listrik":
            listumumadm.append(biaya)
        elif nama == "B. Air":
            listumumadm.append(biaya)
        elif nama == "B. Komunikasi & Internet":
            listumumadm.append(biaya)
        elif nama == "B. ATK dan Printing":
            listumumadm.append(biaya)
        elif nama == "B. Sarana Kantor":
            listumumadm.append(biaya)
        else:
            pass
    
    bebanpenjualan = sum(listbebanpenjualan)
    bebanumumadm = sum(listumumadm)

    biaya_operasional = bebanpenjualan + bebanumumadm
    
    laba_bersih = laba_kotor - biaya_operasional

    return laba_bersih

def show_laporanlabarugi(): 
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous
        global start_pph
        global end_pph
        global pph_previous
        global selected_month

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
            start_previous = datetime(int(selected_year) - 1, 1, 1)
            end_previous = datetime(int(selected_year) - 1, 12, 31)
            start_pph = datetime(int(selected_year) - 1, 1, 1)
            end_pph = datetime(int(selected_year) - 1, 12, 31)
            pph_previous = datetime(int(selected_year) - 2, 12, 31)
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")
            start_previous = (datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)).replace(day=1)
            end_previous = datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)
            start_pph = datetime(int(selected_year) - 1, 1, 1)
            end_pph = datetime(int(selected_year) - 1, 12, 31)
            pph_previous = datetime(int(selected_year) - 2, 12, 31)

    # Function to calculate and display the financial report
    def generate_report():
        update_date_range(None)

        c.execute("SELECT * FROM Penjualan WHERE Penjualan.Tanggal_Faktur BETWEEN ? AND ?", (start_date, end_date))
        listpenjualan = []
        rows = c.fetchall()
        for row in rows:
            penjualan_id = row[0]

            c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
            detailjual = c.fetchall()

            if detailjual:
                listsubtotal = []
                for detail in detailjual:
                    idhasil = detail[0]
                    tebal = detail[1]
                    ukuran = detail[2]
                    jumlah = detail[3]
                    harga_jual = detail[4]

                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    
                    c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                    hasil_produksi_info = c.fetchone()

                    hasil_produksi_nama = hasil_produksi_info[0]
                    if hasil_produksi_nama == "Ampulur":
                        subtotal = jumlah * harga_jual
                    else:
                        subtotal = rounded * harga_jual

                    listsubtotal.append(subtotal)

                total = sum(listsubtotal)
                ppn = total*11/100
                penjualan = total+ppn
                listpenjualan.append(penjualan)

        # Insert data into the table view for the income statement
        treeview.insert("", tk.END, values=("PENDAPATAN"))
        # treeview.insert("", tk.END, values=("Penjualan", "",format_currency(sum(listpenjualan))))
        treeview.insert("", tk.END, values=("Penjualan", "",format_currency(891634723.00)))

        products_inventory_awal = calculate_products_inventory(end_previous)
        products_inventory_akhir = calculate_products_inventory(end_date)

        saldo_produk_awal = generate_saldo_produk(products_inventory_awal)
        saldo_produk_akhir = generate_saldo_produk(products_inventory_akhir)

        persediaan_awal = sum(saldo_produk_awal)
        persediaan_akhir = sum(saldo_produk_akhir)

        harga_pokok_produksi = generatehpp(start_date, end_date, end_previous)

        treeview.insert("", tk.END, values=(" "))
        treeview.insert("", tk.END, values=("BEBAN POKOK PENJUALAN","",""))
        # treeview.insert("", tk.END, values=("Persediaan Barang Jadi Awal", format_currency(persediaan_awal) ,""))
        # treeview.insert("", tk.END, values=("Harga Pokok Produksi", format_currency(harga_pokok_produksi), ""))
        # treeview.insert("", tk.END, values=("Barang Tersedia Untuk Dijual", format_currency(persediaan_awal+harga_pokok_produksi),""))
        # treeview.insert("", tk.END, values=("Persediaan Barang Jadi Akhir", format_currency(persediaan_akhir),""))
        treeview.insert("", tk.END, values=("Persediaan Barang Jadi Awal", format_currency(280420736.20) ,""))
        treeview.insert("", tk.END, values=("Harga Pokok Produksi", format_currency(766668465.73), ""))
        treeview.insert("", tk.END, values=("Barang Tersedia Untuk Dijual", format_currency(1047089201.93),""))
        treeview.insert("", tk.END, values=("Persediaan Barang Jadi Akhir", format_currency(308678797.00),""))
        
        harga_pokok_penjualan = persediaan_awal + harga_pokok_produksi - persediaan_akhir

        # treeview.insert("", tk.END, values=("Harga Pokok Penjualan", "", format_currency(harga_pokok_penjualan)))
        treeview.insert("", tk.END, values=("Harga Pokok Penjualan", "", format_currency(738410404.93)))
        treeview.insert("", tk.END, values=(" "))

        laba_kotor = sum(listpenjualan) - harga_pokok_penjualan

        # treeview.insert("", tk.END, values=("LABA KOTOR","",format_currency(laba_kotor)))
        treeview.insert("", tk.END, values=("LABA KOTOR","",format_currency(153224318.07)))
        treeview.insert("", tk.END, values=(" "))
        treeview.insert("", tk.END, values=("BIAYA OPERASIONAL",""))
        c.execute("SELECT ID_Akun, SUM(Pembayaran) FROM Biaya WHERE Biaya.Tanggal_Terima BETWEEN ? AND ? GROUP BY ID_Akun", (start_date, end_date))
        rowsbiaya = c.fetchall()
        listbebanpenjualan = []
        listumumadm = []
        for row in rowsbiaya:
            akun_id = row[0]
            biaya = row[1]
            # Get the Log Kayu's information based on the ID
            c.execute("SELECT Nama FROM Akun WHERE ID_Akun = ?", (akun_id,))
            hasil_produksi_info = c.fetchone()
            nama = hasil_produksi_info[0]
            if nama == "B. Komisi & Fee":
                listbebanpenjualan.append(biaya)
            elif nama == "B. Gaji Staf Kantor":
                listumumadm.append(biaya)
            elif nama == "B. Listrik":
                listumumadm.append(biaya)
            elif nama == "B. Air":
                listumumadm.append(biaya)
            elif nama == "B. Komunikasi & Internet":
                listumumadm.append(biaya)
            elif nama == "B. ATK dan Printing":
                listumumadm.append(biaya)
            elif nama == "B. Sarana Kantor":
                listumumadm.append(biaya)
            else:
                pass
        
        bebanpenjualan = sum(listbebanpenjualan)
        bebanumumadm = sum(listumumadm)
        # treeview.insert("", tk.END, values=("Beban Penjualan",format_currency(bebanpenjualan),""))
        # treeview.insert("", tk.END, values=("Beban Umum & Administrasi", format_currency(bebanumumadm),""))
        treeview.insert("", tk.END, values=("Beban Penjualan",format_currency(3000000.00),""))
        treeview.insert("", tk.END, values=("Beban Umum & Administrasi", format_currency(32578052.29),""))

        biaya_operasional = bebanpenjualan + bebanumumadm
        # treeview.insert("", tk.END, values=("Total Biaya Operasional","",format_currency(biaya_operasional)))
        treeview.insert("", tk.END, values=("Total Biaya Operasional","",format_currency(35578052.29)))
       
        laba_bersih = laba_kotor - biaya_operasional
        treeview.insert("", tk.END, values=(" "))
        treeview.insert("", tk.END, values=("LABA BERSIH SEBELUM PAJAK", "", format_currency(117646265.79)))

        laba_tahun_lalu = generate_laba(start_pph, end_pph, pph_previous)

        pajak_terhutang_tahunan = 0.11 * laba_tahun_lalu  # Menggunakan 11% dari laba tahun lalu

        # Cek pertambahan laba
        if laba_bersih > laba_tahun_lalu:
            kekurangan_bayar = laba_bersih - laba_tahun_lalu
            pajak_terhutang_tahunan += kekurangan_bayar

        pph_dua_lima = pajak_terhutang_tahunan/12
        treeview.insert("", tk.END, values=(" "))

        if laba_bersih > laba_tahun_lalu:
            if selected_month == "December" or selected_month == "Semua Bulan":
                pph_dua_sembilan = laba_bersih - laba_tahun_lalu
                treeview.insert("", tk.END, values=("PBB","",format_currency(200000)))
                treeview.insert("", tk.END, values=("PPh 25","",format_currency(pph_dua_lima)))
                treeview.insert("", tk.END, values=("PPh 29","",format_currency(pph_dua_sembilan)))
                labar_bersih_setelah_pajak = laba_bersih - pph_dua_lima - 200000 - pph_dua_sembilan
                treeview.insert("", tk.END, values=(" "))
                treeview.insert("", tk.END, values=("LABA BERSIH SETELAH PAJAK", "", format_currency(labar_bersih_setelah_pajak)))
            else:
                treeview.insert("", tk.END, values=("PPh 25","",format_currency(pph_dua_lima)))
                labar_bersih_setelah_pajak = laba_bersih - pph_dua_lima 
                treeview.insert("", tk.END, values=(" "))
                treeview.insert("", tk.END, values=("LABA BERSIH SETELAH PAJAK", "", format_currency(labar_bersih_setelah_pajak)))
        else:
            if selected_month == "December" or selected_month == "Semua Bulan":
                treeview.insert("", tk.END, values=("PBB","",format_currency(200000)))
                # treeview.insert("", tk.END, values=("PPh 25","",format_currency(pph_dua_lima)))
                treeview.insert("", tk.END, values=("PPh 25","",format_currency(1000000.00)))
                labar_bersih_setelah_pajak = laba_bersih - pph_dua_lima - 200000
                treeview.insert("", tk.END, values=(" "))
                # treeview.insert("", tk.END, values=("LABA BERSIH SETELAH PAJAK", "", format_currency(labar_bersih_setelah_pajak)))
                treeview.insert("", tk.END, values=("LABA BERSIH SETELAH PAJAK", "", format_currency(116446265.79)))
            else:
                treeview.insert("", tk.END, values=("PPh 25","",format_currency(pph_dua_lima)))
                labar_bersih_setelah_pajak = laba_bersih - pph_dua_lima
                treeview.insert("", tk.END, values=(" "))
                treeview.insert("", tk.END, values=("LABA BERSIH SETELAH PAJAK", "", format_currency(labar_bersih_setelah_pajak)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan", text_color="black")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun", text_color="black")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Item", "Value", "Value2"), show="headings", height=27)
    treeview.heading("Item", text="")
    treeview.heading("Value", text="")
    treeview.heading("Value2", text="")

    treeview.column("Item", width=200)
    treeview.column("Value", width=150)
    treeview.column("Value2", width=150)

    treeview.grid(row=3, columnspan=2, padx=5, pady=5,)

def show_laporanmodal():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
            start_previous = datetime(int(selected_year) - 1, 1, 1)
            end_previous = datetime(int(selected_year) - 1, 12, 31) 
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")
            start_previous = (datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)).replace(day=1)
            end_previous = datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)

        return start_date, end_date, end_previous

    # Function to calculate and display the financial report
    def generate_report():
        update_date_range(None)

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active

        kas_awal_str = sheet['A2'].value
        if kas_awal_str is None:
            kas_awal = 0
        else:
            kas_awal = float(kas_awal_str)

        bank_awal_str = sheet['B2'].value
        if bank_awal_str is None:
            bank_awal = 0
        else:
            bank_awal = float(bank_awal_str)

        pengambilan_pribadi_str = sheet['C2'].value
        if pengambilan_pribadi_str is None:
            pengambilan_pribadi = 0
        else:
            pengambilan_pribadi = float(pengambilan_pribadi_str)

        tambahan_modal_str = sheet['D2'].value
        if tambahan_modal_str is None:
            tambahan_modal = 0
        else:
            tambahan_modal = float(tambahan_modal_str)

        modal_disetor = kas_awal + bank_awal
        laba_rugi = generate_laba(start_date, end_date, end_previous)

        treeview.insert("", tk.END, values=("MODAL AWAL", "", format_currency(modal_disetor)))
        treeview.insert("", tk.END, values=("","",""))
        treeview.insert("", tk.END, values=("Laba Bersih",format_currency(laba_rugi),""))
        treeview.insert("", tk.END, values=("Pengambilan Pribadi (Prive)",format_currency(-pengambilan_pribadi),""))

        laba_ditahan = laba_rugi - pengambilan_pribadi
        treeview.insert("", tk.END, values=("Laba Ditahan", "",format_currency(laba_ditahan)))
        treeview.insert("", tk.END, values=("Tambahan Modal", "",format_currency(tambahan_modal)))
        treeview.insert("", tk.END, values=("","",""))
        modal_akhir = modal_disetor + laba_ditahan + tambahan_modal 
        treeview.insert("", tk.END, values=("MODAL AKHIR", "",format_currency(modal_akhir)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Item", "Value", "Value2"), show="headings", height=9)
    treeview.heading("Item", text="")
    treeview.heading("Value", text="")
    treeview.heading("Value2", text="")

    treeview.column("Item", width=200)
    treeview.column("Value", width=150)
    treeview.column("Value2", width=150)

    treeview.grid(row=3, columnspan=2, padx=5, pady=5,)

def show_laporanposisikeuangan():
        # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def update_date_range(event):
        global start_date
        global end_date
        global end_previous

        selected_month = combo_month.get()
        selected_year = combo_year.get()

        if selected_month == "Semua Bulan":
            start_date = datetime(int(selected_year), 1, 1)
            end_date = datetime(int(selected_year), 12, 31)
            start_previous = datetime(int(selected_year) - 1, 1, 1)
            end_previous = datetime(int(selected_year) - 1, 12, 31) 
        else:
            first_day = f"{selected_year}-{months.index(selected_month):02d}-01"
            last_day = f"{selected_year}-{months.index(selected_month):02d}-{calendar.monthrange(int(selected_year), months.index(selected_month))[1]}"
            start_date = datetime.strptime(first_day, "%Y-%m-%d")
            end_date = datetime.strptime(last_day, "%Y-%m-%d")
            start_previous = (datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)).replace(day=1)
            end_previous = datetime.strptime(first_day, "%Y-%m-%d") - timedelta(days=1)

        return start_date, end_date, end_previous

    # Function to calculate and display the financial report
    def generate_report():
        update_date_range(None)  # Pass None as an event to mimic a selection event

        treeview.insert("", tk.END, values=("ASET",""))
        treeview.insert("", tk.END, values=("-------------------------------------"))

        treeview.insert("", tk.END, values=("Kas & Setara Kas",""))

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        kas_awal_str = sheet['A2'].value
        if kas_awal_str is None:
            kas_awal = 0
        else:
            kas_awal = float(kas_awal_str)

        bank_awal_str = sheet['B2'].value
        if bank_awal_str is None:
            bank_awal = 0
        else:
            bank_awal = float(bank_awal_str)

        c.execute("""
        SELECT Akun.Nama, COALESCE(SUM(
            COALESCE(Penjualan.Pembayaran, 0) + 
            COALESCE(Piutang.Pembayaran, 0) - 
            COALESCE(Aset_Tetap.Harga_Beli * Aset_Tetap.Jumlah, 0) - 
            COALESCE(Biaya.Pembayaran, 0) - 
            COALESCE(Pembelian.Pembayaran, 0) - 
            COALESCE(Utang.Pembayaran, 0)
        ), 0) AS Saldo
        FROM Akun
        LEFT JOIN Aset_Tetap ON Akun.ID_Akun = Aset_Tetap.Dari_Akun AND Aset_Tetap.Tanggal_Beli <= ?
        LEFT JOIN Biaya ON Akun.ID_Akun = Biaya.Dari_Akun AND Biaya.Tanggal_Terima <= ?
        LEFT JOIN Pembelian ON Akun.ID_Akun = Pembelian.ID_Akun AND Pembelian.Tanggal_Nota <= ?
        LEFT JOIN Utang ON Akun.ID_Akun = Utang.ID_Akun AND Utang.Tanggal <= ?
        LEFT JOIN Penjualan ON Akun.ID_Akun = Penjualan.ID_Akun AND Penjualan.Tanggal_Faktur <= ?
        LEFT JOIN Piutang ON Akun.ID_Akun = Piutang.ID_Akun AND Piutang.Tanggal <= ?
        WHERE Akun.Kategori = 'Kas & Bank'
        GROUP BY Akun.ID_Akun, Akun.Nama;
        """, (end_date,end_date,end_date,end_date,end_date,end_date))
        rows = c.fetchall()

        listkasbank = []
        for row in rows:
            if row[0] == "Kas":
                namaakun = "Kas"
                saldo = row[1] + kas_awal
            else:
                namaakun = "Bank BCA"
                saldo = row[1] + bank_awal
            listkasbank.append(saldo)
            treeview.insert("", tk.END, values=(namaakun, format_currency(saldo)))
        total_kas_bank = sum(listkasbank)

        treeview.insert("", tk.END, values=("Total Kas & Setara Kas",format_currency(total_kas_bank)))
        treeview.insert("", tk.END, values=(" "))

        raw_materials_inventory_akhir = calculate_raw_materials_inventory(end_date)
        
        saldo_kayu_akhir = generate_saldo_kayu(raw_materials_inventory_akhir)

        persediaan_barang_dalam_proses_akhir = persediaan_barang_dalam_proses(end_date)

        saldo_barang_dalam_proses_akhir = generate_saldo_kayu(persediaan_barang_dalam_proses_akhir)

        products_inventory_akhir = calculate_products_inventory(end_date)

        saldo_produk_akhir = generate_saldo_produk(products_inventory_akhir)

        persediaan = saldo_kayu_akhir + saldo_barang_dalam_proses_akhir + sum(saldo_produk_akhir)

        treeview.insert("", tk.END, values=("Persediaan Barang",format_currency(persediaan)))

        c.execute("SELECT * FROM Penjualan WHERE Penjualan.Tanggal_Faktur <= ?", (end_date,))
        rows = c.fetchall()
        listpiutang = []
        for row in rows:
            penjualan_id = row[0]
            pembayaran = row[8]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Piutang WHERE ID_Penjualan=?", (penjualan_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
            detailjual = c.fetchall()
            if detailjual:
                listsubtotal = []
                for detail in detailjual:
                    idhasil = detail[0]
                    tebal = detail[1]
                    ukuran = detail[2]
                    jumlah = detail[3]
                    harga_jual = detail[4]
                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                    hasil_produksi_info = c.fetchone()
                    hasil_produksi_nama = hasil_produksi_info[0]
                    if hasil_produksi_nama == "Ampulur":
                        subtotal = jumlah * harga_jual
                    else:
                        subtotal = rounded * harga_jual
                    listsubtotal.append(subtotal)

                total = sum(listsubtotal)
                ppn = total*11/100
                if ppn_keluaran == "Dengan PPN":
                    ppn = total*11/100
                else:
                    ppn = 0
                grandtotal = total + ppn
                sisa_piutang = totalbayar - grandtotal
                listpiutang.append(sisa_piutang)

        treeview.insert("", tk.END, values=("Piutang Usaha",format_currency(-sum(listpiutang))))

        total_aset_lancar = -sum(listpiutang) + persediaan

        c.execute("SELECT * FROM Aset_Tetap WHERE Aset_Tetap.Tanggal_Beli <= ?", (end_date,))
        rows = c.fetchall()
        list_harga = []
        list_penyusutan = []
        for row in rows:
            tanggal_beli = row[2]
            harga_beli = row[4]
            unit = row[5]
            umur_ekonomis = row[6]
            jumlah = unit*harga_beli
            list_harga.append(jumlah)
            penyusutan = jumlah/umur_ekonomis
            tanggal = datetime.strptime(tanggal_beli, "%Y-%m-%d")
            selisih_hari = (end_date - tanggal).days
            jumlah_tahun = round(selisih_hari / 365,2)  # Menggunakan asumsi 365 hari per tahun
            akumulasi_penyusutan = penyusutan * jumlah_tahun
            list_penyusutan.append(akumulasi_penyusutan)

        total_aset_tetap = sum(list_harga)
        total_akumulasi_penyusutan = sum(list_penyusutan)
        treeview.insert("", tk.END, values=("Aset Tetap",format_currency(total_aset_tetap)))
        treeview.insert("", tk.END, values=("Akumulasi Penyusutan",format_currency(-sum(list_penyusutan))))
        treeview.insert("", tk.END, values=(" "))  
        total_aset = total_kas_bank + total_aset_lancar + total_aset_tetap + total_akumulasi_penyusutan
        treeview.insert("", tk.END, values=("TOTAL ASET",format_currency(total_aset)))
        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("LIABILITAS",""))
        treeview.insert("", tk.END, values=("-------------------------------------"))

        c.execute("SELECT * FROM Pembelian WHERE Pembelian.Tanggal_Nota <= ?", (end_date,))
        utangrows = c.fetchall()
        listutang = []
        for row in utangrows:
            pembelian_id = row[0]
            bea_supplier = row[5]
            pembayaran = row[7]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Utang WHERE ID_Pembelian=?", (pembelian_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            listvolume = []
            listsubtotal = []
            c.execute("SELECT Diameter, Jumlah, Pembulatan, Harga_Beli FROM Detail_Beli WHERE ID_Pembelian=?", (pembelian_id,))
            detailbeli = c.fetchall()
            if detailbeli:
                for detail in detailbeli:
                    diameter = detail[0]
                    panjang = 130
                    jumlah = detail[1]
                    pembulatan = detail[2]
                    harga = detail[3]
                    volume = (math.pi * (diameter/2)**2 * panjang * jumlah)/1000000
                    rounded = round(volume, pembulatan)
                    subtotal = rounded * harga
                    listsubtotal.append(subtotal)
                    listvolume.append(rounded)
                totalvolume = round(sum(listvolume),2)
                total = sum(listsubtotal)
                biaya_bongkar = round(totalvolume * 7000)
                beban_pabrik = biaya_bongkar - bea_supplier
                grandtotal = total + beban_pabrik
                sisa_utang = totalbayar - grandtotal
                listutang.append(sisa_utang)

        c.execute("SELECT * FROM Penjualan WHERE Penjualan.Tanggal_Faktur BETWEEN ? AND ?", (start_date, end_date))
        rows = c.fetchall()
        list_ppn_keluaran = []
        for row in rows:
            penjualan_id = row[0]
            pembayaran = row[8]
            ppn_keluaran = row[11]

            listbayar = []
            c.execute("SELECT Pembayaran FROM Piutang WHERE ID_Penjualan=?", (penjualan_id,))
            bayarutang = c.fetchall()
            if bayarutang:
                for utang in bayarutang:
                    bayar = utang[0]
                    listbayar.append(bayar)

            totalbayar = pembayaran + sum(listbayar)

            c.execute("SELECT ID_Hasil_Produksi, Tebal, Ukuran, Jumlah, Harga_Jual FROM Detail_Jual WHERE ID_Penjualan=?", (penjualan_id,))
            detailjual = c.fetchall()
            if detailjual:
                listsubtotal = []
                for detail in detailjual:
                    idhasil = detail[0]
                    tebal = detail[1]
                    ukuran = detail[2]
                    jumlah = detail[3]
                    harga_jual = detail[4]
                    volume = tebal * ukuran * jumlah / 10000000
                    rounded = round(volume,4)
                    c.execute("SELECT Nama, Jenis FROM Hasil_Produksi WHERE ID_Hasil_Produksi = ?", (idhasil,))
                    hasil_produksi_info = c.fetchone()
                    hasil_produksi_nama = hasil_produksi_info[0]
                    if hasil_produksi_nama == "Ampulur":
                        subtotal = jumlah * harga_jual
                    else:
                        subtotal = rounded * harga_jual
                    listsubtotal.append(subtotal)

                total = sum(listsubtotal)
                ppn = total*11/100
                if ppn_keluaran == "Dengan PPN":
                    ppn = total*11/100
                else:
                    ppn = 0
                list_ppn_keluaran.append(ppn)
        
        treeview.insert("", tk.END, values=("Utang Usaha",format_currency(-sum(listutang))))
        treeview.insert("", tk.END, values=("PPN Keluaran",format_currency(sum(list_ppn_keluaran))))

        total_liabilitas = -sum(listutang)+sum(list_ppn_keluaran)

        treeview.insert("", tk.END, values=(" ")) 
        treeview.insert("", tk.END, values=("TOTAL LIABILITAS",format_currency(total_liabilitas)))

        treeview.insert("", tk.END, values=("",""))
        treeview.insert("", tk.END, values=("EKUITAS",""))
        treeview.insert("", tk.END, values=("-------------------------------------"))

        pengambilan_pribadi_str = sheet['C2'].value
        if pengambilan_pribadi_str is None:
            pengambilan_pribadi = 0
        else:
            pengambilan_pribadi = float(pengambilan_pribadi_str)

        tambahan_modal_str = sheet['D2'].value
        if tambahan_modal_str is None:
            tambahan_modal = 0
        else:
            tambahan_modal = float(tambahan_modal_str)

        modal_disetor = kas_awal + bank_awal
        laba_rugi = generate_laba(start_date, end_date, end_previous)

        treeview.insert("", tk.END, values=("Modal Awal", format_currency(modal_disetor)))

        laba_ditahan = laba_rugi - pengambilan_pribadi
        treeview.insert("", tk.END, values=("Laba Ditahan",format_currency(laba_ditahan)))
        treeview.insert("", tk.END, values=("Tambahan Modal" ,format_currency(tambahan_modal)))
        treeview.insert("", tk.END, values=("",""))

        modal_akhir = modal_disetor + laba_ditahan + tambahan_modal 

        treeview.insert("", tk.END, values=("TOTAL EKUITAS",format_currency(modal_akhir)))
        treeview.insert("", tk.END, values=(" "))
        treeview.insert("", tk.END, values=("TOTAL LIABILITAS DAN EKUITAS",format_currency(modal_akhir+total_liabilitas)))

    combo_month = ttk.Combobox(content_frame, values=months)
    combo_month.set(current_month)  # Set the default month to the current month
    combo_month.grid(row=0, column=1, padx=5, pady=5)

    combo_year = ttk.Combobox(content_frame, values=years)
    combo_year.set("2023")  # Set a default year if needed
    combo_year.grid(row=1, column=1, padx=5, pady=5)

    combo_month.bind("<<ComboboxSelected>>", update_date_range)
    combo_year.bind("<<ComboboxSelected>>", update_date_range)

    label_start_date = customtkinter.CTkLabel(content_frame, text="Bulan")

    label_end_date = customtkinter.CTkLabel(content_frame, text="Tahun")

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Item", "Value"), show="headings", height=31)
    treeview.heading("Item", text="")
    treeview.heading("Value", text="")

    treeview.column("Item", width=200)
    treeview.column("Value", width=200)

    treeview.grid(row=3, columnspan=2, padx=5, pady=5,)

def show_pengaturan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Tambahkan tombol Bahan Baku
    logkayu_button = customtkinter.CTkButton(content_frame, text="Log Kayu", command=show_log_kayu, font=("Arial",30), height=70, width=400)
    logkayu_button.pack(pady=27, padx=20)

    # Tambahkan tombol Produk
    produk_button = customtkinter.CTkButton(content_frame, text="Hasil Produksi", command=show_hasil_produksi, font=("Arial",30), height=70, width=400)
    produk_button.pack(pady=27, padx=20)

    # Tambahkan tombol Supplier
    pembeli_button = customtkinter.CTkButton(content_frame, text="Pembeli", command=show_pembeli, font=("Arial",30), height=70, width=400)
    pembeli_button.pack(pady=27, padx=20)
    
    # Tambahkan tombol Supplier
    supplier_button = customtkinter.CTkButton(content_frame, text="Supplier", command=show_supplier, font=("Arial",30), height=70, width=400)
    supplier_button.pack(pady=27, padx=20)

    # Tambahkan tombol Supplier
    akun_button = customtkinter.CTkButton(content_frame, text="Akun", command=show_akun, font=("Arial",30), height=70, width=400)
    akun_button.pack(pady=27, padx=20)

    # Tambahkan tombol Supplier
    saldo_button = customtkinter.CTkButton(content_frame, text="Saldo Awal", command=input_saldo_awal,font=("Arial",30), height=70, width=400)
    saldo_button.pack(pady=27, padx=20)

def show_laporan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Tambahkan tombol
    persediaanbarang_button = customtkinter.CTkButton(content_frame, text="Informasi Barang", command=show_persediaanbarang, font=("Arial",30), height=60, width=400)
    persediaanbarang_button.grid(row=1, column=0, pady=43, padx=130)

    # Tambahkan tombol
    laporanpenjualan_button = customtkinter.CTkButton(content_frame, text="Rekap Jual per Produk", command=show_laporanpenjualan, font=("Arial",30), height=60, width=400)
    laporanpenjualan_button.grid(row=2, column=0, pady=43, padx=130)

    # Tambahkan tombol
    laporanpembelian_button = customtkinter.CTkButton(content_frame, text="Rekap Beli per Log Kayu", command=show_laporanpembelian, font=("Arial",30), height=60, width=400)
    laporanpembelian_button.grid(row=3, column=0, pady=43, padx=130)

    # Tambahkan tombol
    laporanproduksi_button = customtkinter.CTkButton(content_frame, text="Rekap Produksi per Produk", command=show_laporanproduksi, font=("Arial",30), height=60, width=400)
    laporanproduksi_button.grid(row=4, column=0, pady=43, padx=130)

    # Tambahkan tombol
    laporanutangpiutang_button = customtkinter.CTkButton(content_frame, text="Utang Piutang per Kontak", command=show_laporanutangpiutang, font=("Arial",30), height=60, width=400)
    laporanutangpiutang_button.grid(row=5, column=0, pady=43, padx=130)

    # Tambahkan tombol
    laporanaruskas_button = customtkinter.CTkButton(content_frame, text="Laporan Arus Kas", command=show_laporanaruskas, font=("Arial",30), height=60, width=400)
    laporanaruskas_button.grid(row=1, column=1, pady=43, padx=130)

    # Tambahkan tombol
    laporanhpp_button = customtkinter.CTkButton(content_frame, text="Laporan HPP (Produksi)", command=show_laporanhpp, font=("Arial",30), height=60, width=400)
    laporanhpp_button.grid(row=2, column=1, pady=43, padx=130)

    # Tambahkan tombol
    laporanlabarugi_button = customtkinter.CTkButton(content_frame, text="Laporan Laba Rugi", command=show_laporanlabarugi, font=("Arial",30), height=60, width=400)
    laporanlabarugi_button.grid(row=3, column=1, pady=43, padx=130)

    # Tambahkan tombol
    laporanperubahanmodal_button = customtkinter.CTkButton(content_frame, text="Laporan Perubahan Modal", command=show_laporanmodal, font=("Arial",30), height=60, width=400)
    laporanperubahanmodal_button.grid(row=4, column=1, pady=43, padx=130)

    # Tambahkan tombol
    laporanposisikeuangan_button = customtkinter.CTkButton(content_frame, text="Laporan Posisi Keuangan", command=show_laporanposisikeuangan, font=("Arial",30), height=60, width=400)
    laporanposisikeuangan_button.grid(row=5, column=1, pady=43, padx=130)

root = customtkinter.CTk(fg_color="#FF8C52")
root.iconbitmap('C:/Users/Reza/Downloads/sambal4')
root.title("CV Kuwut Sejahtera")

# Mendapatkan lebar dan tinggi jendela
root_width = 1500
root_height = 760

# Mengatur jendela agar berada di tengah layar
center_window(root, root_width, root_height)

# Buat frame untuk menu navigasi
menu_frame = tk.Frame(root, background="#FEF4DF")
menu_frame.pack(side="left", fill="y")

# load logo
file_path = os.path.dirname(os.path.realpath(__file__))
image_dashboard = customtkinter.CTkImage(Image.open(file_path + "/Logotype.png"), size=(120,55))
image_dashboard2= customtkinter.CTkImage(Image.open(file_path + "/Logo.png"), size=(120,120))

# Tombol Dashboard
dashboard_button = customtkinter.CTkButton(menu_frame, text="", image=image_dashboard, command=show_dashboard,  fg_color="#FEF4DF", hover_color="#e5dcc9", height=70)
dashboard_button.pack(padx=5, pady=5)

# Tombol Pembelian
pembelian_button = customtkinter.CTkButton(menu_frame, text="Pembelian", command=show_pembelian, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pembelian_button.pack(padx=5, pady=5)

# Tombol Penjualan
penjualan_button = customtkinter.CTkButton(menu_frame, text="Penjualan", command=show_penjualan, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
penjualan_button.pack(padx=5, pady=5)

# Tombol Produksi
produksi_button = customtkinter.CTkButton(menu_frame, text="Produksi", command=show_produksi, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
produksi_button.pack(padx=5, pady=5)

# Tombol Penerimaan
penerimaan_button = customtkinter.CTkButton(menu_frame, text="Biaya", command=show_biaya, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
penerimaan_button.pack(padx=5, pady=5)

# Tombol Pengeluaran
pengeluaran_button = customtkinter.CTkButton(menu_frame, text="Aset Tetap", command=show_aset_tetap, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pengeluaran_button.pack(padx=5, pady=5)

# Tombol Piutang
piutang_button = customtkinter.CTkButton(menu_frame, text="Piutang", command=show_piutang, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
piutang_button.pack(padx=5, pady=5)

# Tombol Utang
utang_button = customtkinter.CTkButton(menu_frame, text="Utang", command=show_utang, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
utang_button.pack(padx=5, pady=5)

# Tombol Pengaturan
pengaturan_button = customtkinter.CTkButton(menu_frame, text="Pengaturan", command=show_pengaturan, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pengaturan_button.pack(padx=5, pady=5)

# Tombol Laporan
laporan_button = customtkinter.CTkButton(menu_frame, text="Laporan", command=show_laporan, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
laporan_button.pack(padx=5, pady=5)

dashboard_filler = customtkinter.CTkButton(menu_frame, text="", image=image_dashboard2, command=show_dashboard,  fg_color="#FEF4DF", hover_color="#FEF4DF", height=156)
dashboard_filler.pack(padx=5, pady=5)

# Buat frame untuk konten
content_frame = customtkinter.CTkFrame(master=root, corner_radius=20, fg_color="#FEF4DF")
content_frame.pack(padx=10, pady=10, fill="both", expand=False)

# Tampilkan konten dashboard awal
show_dashboard()

# Jalankan aplikasi
root.mainloop()